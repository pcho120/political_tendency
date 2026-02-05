#!/usr/bin/env python3
"""
Firm Website Finder - Desktop App
Standalone desktop application with file dialogs (no browser needed)
Uses openpyxl only - no pandas/numpy to avoid PyInstaller conflicts
"""

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import time
import re
import os
import sys
import threading
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote_plus, urlparse
from openpyxl import load_workbook, Workbook
from pathlib import Path

# Configuration
BLOCKED_DOMAINS = [
    "wikipedia.org", "linkedin.com", "facebook.com", "instagram.com", "twitter.com",
    "law.com", "vault.com", "chambers.com", "bloomberg.com", "crunchbase.com"
]

SEARCH_DELAY = 0.5

# Global variables
current_file = None
output_file = None
processing_thread = None


class FirmWebsiteFinderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Firm Website Finder")
        self.root.geometry("650x500")
        self.root.resizable(True, True)
        
        # Configure grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # Main frame
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid for main frame
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=0)
        main_frame.rowconfigure(1, weight=0)
        main_frame.rowconfigure(2, weight=0)
        main_frame.rowconfigure(3, weight=1)
        main_frame.rowconfigure(4, weight=0)
        
        # Header
        header = ttk.Label(main_frame, text="Firm Website Finder", font=("Arial", 16, "bold"))
        header.grid(row=0, column=0, pady=(0, 20))
        
        # File selection area
        file_frame = ttk.LabelFrame(main_frame, text="Input File", padding="10")
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        
        # File path label
        self.file_path_var = tk.StringVar()
        self.file_path_var.set("No file selected")
        self.file_path_label = ttk.Label(file_frame, textvariable=self.file_path_var, 
                                          relief="sunken", anchor="w", padding="5")
        self.file_path_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Browse button
        browse_btn = ttk.Button(file_frame, text="Browse for Excel File...", command=self.browse_file)
        browse_btn.grid(row=1, column=0, pady=(0, 5))
        
        # Settings area
        settings_frame = ttk.LabelFrame(main_frame, text="Settings", padding="10")
        settings_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        settings_frame.columnconfigure(1, weight=1)
        
        # Delay setting
        delay_label = ttk.Label(settings_frame, text="Delay between searches (seconds):")
        delay_label.grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        self.delay_var = tk.DoubleVar(value=0.5)
        delay_spin = ttk.Spinbox(settings_frame, from_=0.2, to=3.0, increment=0.1, 
                                  textvariable=self.delay_var, width=8)
        delay_spin.grid(row=0, column=1, sticky=tk.W)
        
        # Max firms setting
        max_label = ttk.Label(settings_frame, text="Max firms to process (0 = all):")
        max_label.grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        
        self.max_var = tk.IntVar(value=0)
        max_spin = ttk.Spinbox(settings_frame, from_=0, to=10000, increment=10, 
                                textvariable=self.max_var, width=8)
        max_spin.grid(row=1, column=1, sticky=tk.W, pady=(5, 0))
        
        # Progress area
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        progress_frame.rowconfigure(0, weight=1)
        
        # Progress text
        self.progress_text = tk.Text(progress_frame, height=12, state="disabled", wrap="word")
        self.progress_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(progress_frame, orient="vertical", command=self.progress_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.progress_text.config(yscrollcommand=scrollbar.set)
        
        # Button area
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, pady=(0, 0))
        
        # Process button
        self.process_btn = ttk.Button(button_frame, text="Find Official Websites", 
                                       command=self.start_processing)
        self.process_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Clear button
        clear_btn = ttk.Button(button_frame, text="Clear", command=self.clear)
        clear_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Exit button
        exit_btn = ttk.Button(button_frame, text="Exit", command=self.root.quit)
        exit_btn.pack(side=tk.LEFT)
        
        # Center window
        self.center_window()
        
        # Show welcome message
        self.update_progress("Welcome to Firm Website Finder!\n\n")
        self.update_progress("Instructions:\n")
        self.update_progress("1. Click 'Browse for Excel File...' to select your file\n")
        self.update_progress("2. Your Excel file must have a 'Firm' column\n")
        self.update_progress("3. Click 'Find Official Websites' to start\n")
        self.update_progress("4. Output will be saved automatically\n\n")
    
    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    def browse_file(self):
        """Browse for Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            self.load_file(file_path)
    
    def load_file(self, file_path):
        """Load the selected file"""
        global current_file
        
        try:
            # Check if it's a valid Excel file using openpyxl
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Find header row and Firm column
            firm_col = None
            headers = []
            for col_idx, cell in enumerate(ws[1], start=1):
                val = cell.value
                headers.append(val)
                if val and str(val).strip().lower() == "firm":
                    firm_col = col_idx
            
            wb.close()
            
            if firm_col is None:
                messagebox.showerror("Invalid File", 
                    "The Excel file must have a 'Firm' column.\n\n" +
                    f"Found columns: {', '.join(str(h) for h in headers if h)}")
                return
            
            # Count rows
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            row_count = sum(1 for row in ws.iter_rows(min_row=2) if row[firm_col-1].value)
            wb.close()
            
            current_file = file_path
            self.file_path_var.set(file_path)
            
            self.progress_text.config(state="normal")
            self.progress_text.delete(1.0, tk.END)
            self.progress_text.insert(tk.END, f"Loaded: {os.path.basename(file_path)}\n")
            self.progress_text.insert(tk.END, f"Firms found: {row_count}\n\n")
            self.progress_text.insert(tk.END, "Ready to process. Click 'Find Official Websites' to start.\n")
            self.progress_text.config(state="disabled")
            
            messagebox.showinfo("File Loaded", 
                f"Successfully loaded {os.path.basename(file_path)}\n\nFirms found: {row_count}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file:\n\n{e}")
            current_file = None
            self.file_path_var.set("No file selected")
    
    def start_processing(self):
        """Start processing the file"""
        global current_file, processing_thread
        
        if not current_file:
            messagebox.showwarning("No File", "Please select an Excel file first")
            return
        
        if processing_thread and processing_thread.is_alive():
            messagebox.showwarning("Processing", "Already processing a file")
            return
        
        # Disable button during processing
        self.process_btn.config(state="disabled")
        
        # Start processing in a separate thread
        processing_thread = threading.Thread(target=self.process_file)
        processing_thread.daemon = True
        processing_thread.start()
    
    def process_file(self):
        """Process the Excel file to find websites"""
        global current_file, output_file
        
        try:
            # Load workbook
            wb = load_workbook(current_file)
            ws = wb.active
            
            # Find Firm column
            firm_col = None
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value and str(cell.value).strip().lower() == "firm":
                    firm_col = col_idx
                    break
            
            if firm_col is None:
                raise ValueError("Could not find 'Firm' column")
            
            # Add Official Website column
            last_col = ws.max_column + 1
            ws.cell(row=1, column=last_col, value="Official Website")
            
            # Get settings
            delay = self.delay_var.get()
            max_firms = self.max_var.get()
            
            # Count total firms
            total_firms = 0
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=firm_col).value:
                    total_firms += 1
            
            if max_firms > 0:
                total_firms = min(total_firms, max_firms)
            
            # Process each firm
            websites_found = 0
            failed_firms = []
            processed = 0
            
            # Clear and start progress
            self.root.after(0, lambda: self.clear_progress())
            self.root.after(0, lambda: self.update_progress(f"Processing {total_firms} firms...\n\n"))
            
            for row in range(2, ws.max_row + 1):
                firm_name = ws.cell(row=row, column=firm_col).value
                
                if not firm_name or not str(firm_name).strip():
                    continue
                
                processed += 1
                if max_firms > 0 and processed > max_firms:
                    break
                
                firm_name = str(firm_name).strip()
                
                # Update progress
                self.root.after(0, lambda fn=firm_name, p=processed, t=total_firms: 
                    self.update_progress(f"[{p}/{t}] {fn}... "))
                
                # Search for website
                try:
                    website = search_firm_website(firm_name)
                    
                    if website:
                        ws.cell(row=row, column=last_col, value=website)
                        websites_found += 1
                        self.root.after(0, lambda: self.update_progress("Found!\n"))
                    else:
                        failed_firms.append(firm_name)
                        self.root.after(0, lambda: self.update_progress("Not found\n"))
                except Exception as e:
                    failed_firms.append(firm_name)
                    self.root.after(0, lambda err=str(e): self.update_progress(f"Error: {err}\n"))
                
                # Delay between searches
                time.sleep(delay)
            
            # Save output file
            input_dir = os.path.dirname(current_file)
            base_name = os.path.splitext(os.path.basename(current_file))[0]
            output_file = os.path.join(input_dir, f"{base_name}_with_websites.xlsx")
            
            wb.save(output_file)
            wb.close()
            
            # Show results
            self.root.after(0, lambda: self.update_progress(f"\n{'='*50}\n"))
            self.root.after(0, lambda: self.update_progress("PROCESSING COMPLETE!\n"))
            self.root.after(0, lambda: self.update_progress(f"{'='*50}\n"))
            self.root.after(0, lambda: self.update_progress(f"Total firms processed: {processed}\n"))
            self.root.after(0, lambda: self.update_progress(f"Websites found: {websites_found}\n"))
            self.root.after(0, lambda: self.update_progress(f"Not found: {len(failed_firms)}\n"))
            self.root.after(0, lambda: self.update_progress(f"\nOutput saved to:\n{output_file}\n"))
            
            # Show success message
            self.root.after(0, lambda: messagebox.showinfo(
                "Success", 
                f"Processing complete!\n\n" +
                f"Websites found: {websites_found}/{processed}\n\n" +
                f"Output file:\n{output_file}"
            ))
            
        except Exception as e:
            self.root.after(0, lambda: self.update_progress(f"\nERROR: {e}\n"))
            self.root.after(0, lambda: messagebox.showerror("Error", f"Processing failed:\n\n{e}"))
        
        finally:
            # Re-enable button
            self.root.after(0, lambda: self.process_btn.config(state="normal"))
    
    def clear_progress(self):
        """Clear progress text"""
        self.progress_text.config(state="normal")
        self.progress_text.delete(1.0, tk.END)
        self.progress_text.config(state="disabled")
    
    def update_progress(self, text):
        """Update progress text (thread-safe)"""
        self.progress_text.config(state="normal")
        self.progress_text.insert(tk.END, text)
        self.progress_text.see(tk.END)
        self.progress_text.config(state="disabled")
        self.root.update_idletasks()
    
    def clear(self):
        """Clear all fields"""
        global current_file, output_file
        
        current_file = None
        output_file = None
        self.file_path_var.set("No file selected")
        self.progress_text.config(state="normal")
        self.progress_text.delete(1.0, tk.END)
        self.progress_text.insert(tk.END, "Cleared. Select a new file to begin.\n")
        self.progress_text.config(state="disabled")


def search_firm_website(firm_name: str) -> str:
    """Search for firm website using pattern matching and web search."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    
    try:
        clean_name = firm_name.lower()
        clean_name = clean_name.replace("&", "and")
        clean_name = clean_name.replace(",", "")
        clean_name = re.sub(r'[^a-z0-9\s]', '', clean_name)
        
        words = clean_name.split()
        primary_name = words[0] if words else clean_name
        full_name_no_spaces = clean_name.replace(" ", "")
        
        # Known law firm mappings
        known_mappings = {
            "kirkland": "https://www.kirkland.com",
            "latham": "https://www.lw.com",
            "skadden": "https://www.skadden.com",
            "gibson": "https://www.gibsondunn.com",
            "sidley": "https://www.sidley.com",
            "sullivan": "https://www.sullcrom.com",
            "davis polk": "https://www.davispolk.com",
            "weil": "https://www.weil.com",
            "simpson": "https://www.simpsonthacher.com",
            "cleary": "https://www.clearygottlieb.com",
        }
        
        firm_lower = firm_name.lower()
        for key, url in known_mappings.items():
            if key in firm_lower:
                return url
        
        # Try potential domains
        potential_domains = [
            f"https://www.{full_name_no_spaces}.com",
            f"https://{full_name_no_spaces}.com",
            f"https://www.{full_name_no_spaces}law.com",
            f"https://www.{primary_name}law.com",
            f"https://www.{primary_name}.com",
        ]
        
        if len(words) >= 2:
            first_two = "".join(words[:2])
            potential_domains.extend([
                f"https://www.{first_two}.com",
                f"https://www.{first_two}law.com",
            ])
        
        tested = set()
        for domain in potential_domains[:10]:
            if domain in tested:
                continue
            tested.add(domain)
            
            try:
                response = requests.head(domain, headers=headers, timeout=3, allow_redirects=True)
                if response.status_code in [200, 301, 302]:
                    return str(response.url)
            except:
                continue
                
    except:
        pass
    
    # Fallback: DuckDuckGo search
    try:
        search_url = f"https://duckduckgo.com/html/?q={quote_plus(f'{firm_name} law firm official website')}"
        
        response = requests.get(search_url, headers=headers, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, "html.parser")
            
            for link in soup.find_all("a", href=True):
                href = str(link.get("href", ""))
                
                if not href or "duckduckgo" in href:
                    continue
                
                if href.startswith("//"):
                    href = "https:" + href
                
                # Skip blocked domains
                if any(b in href.lower() for b in BLOCKED_DOMAINS):
                    continue
                
                # Look for law-related URLs
                if any(kw in href.lower() for kw in ["law", "llp", "firm"]):
                    try:
                        parsed = urlparse(href)
                        if parsed.scheme and parsed.netloc:
                            return f"{parsed.scheme}://{parsed.netloc}"
                    except:
                        pass
    except:
        pass
    
    return ""


def main():
    """Main entry point."""
    root = tk.Tk()
    app = FirmWebsiteFinderApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
