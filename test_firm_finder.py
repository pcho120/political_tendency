#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test script for firm_finder_desktop.py core functionality
Tests source discovery and validation without GUI
"""

import sys
import os
from pathlib import Path
from openpyxl import load_workbook, Workbook

# Force UTF-8 output on Windows
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Import from discovery module
sys.path.insert(0, str(Path(__file__).parent))
from discovery.firm_finder_desktop import SourceDiscovery, SourceValidator, BLOCKLIST

def test_firm_finder(input_file: str, max_firms: int = 3, timeout: int = 45):
    """Test source discovery and validation"""
    
    print("=" * 60)
    print("TESTING FIRM FINDER - SOURCE DISCOVERY & VALIDATION")
    print("=" * 60)
    print(f"\nInput: {input_file}")
    print(f"Max Firms: {max_firms}")
    print(f"Timeout per firm: {timeout}s\n")
    
    # Load input
    wb = load_workbook(input_file, read_only=True)
    ws = wb.active
    
    # Find columns
    headers = [cell.value for cell in ws[1]]
    firm_col = headers.index("Firm") + 1
    url_col = headers.index("Official Website") + 1
    
    print(f"Columns found: {headers}\n")
    
    # Initialize validator
    validator = SourceValidator()
    
    # Prepare output
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Validated Sources"
    out_ws.append([
        "Firm", "official_base_url", "attorney_list_sources", "profile_sources",
        "source_types", "sample_pass_fields", "notes", "is_valid"
    ])
    
    # Process firms
    processed = 0
    validated = 0
    
    for row in range(2, min(ws.max_row + 1, max_firms + 2)):
        firm_name = ws.cell(row, firm_col).value
        base_url = ws.cell(row, url_col).value
        
        if not firm_name or not base_url:
            continue
        
        processed += 1
        
        print(f"\n{'='*60}")
        print(f"[{processed}] {firm_name}")
        print(f"URL: {base_url}")
        print(f"{'='*60}")
        
        # Discover sources
        discovery = SourceDiscovery(str(base_url), str(firm_name), timeout)
        candidates = discovery.discover_sources()
        
        print(f"\n✓ Discovery complete: {len(candidates)} candidate sources")
        
        # Show candidates
        for i, candidate in enumerate(candidates, 1):
            print(f"  {i}. [{candidate['type']}] {candidate['url'][:80]}...")
        
        # Validate sources
        valid_sources = []
        source_types = []
        sample_fields = set()
        
        print(f"\n→ Validating candidates...")
        
        for candidate in candidates:
            print(f"\n  Testing: [{candidate['type']}] {candidate['url'][:70]}...")
            
            is_valid, result = validator.validate_source(
                candidate["url"], candidate["type"], str(base_url)
            )
            
            print(f"    Profiles tested: {result['tested_profiles']}")
            print(f"    Names found: {result['names_found']}")
            print(f"    Fields found: {result['fields_found']}")
            
            if is_valid:
                print(f"    ✓ VALID - Fields: {result['sample_pass_fields']}")
                valid_sources.append(candidate["url"])
                source_types.append(candidate["type"])
                sample_fields.update(result["sample_pass_fields"])
            else:
                print(f"    ✗ INVALID - {result.get('failure_reason', 'unknown')}")
        
        # Summary
        is_firm_valid = len(valid_sources) > 0
        if is_firm_valid:
            validated += 1
        
        print(f"\n{'='*60}")
        print(f"RESULT: {len(valid_sources)} valid sources")
        if valid_sources:
            for src in valid_sources:
                print(f"  • {src}")
        print(f"{'='*60}")
        
        # Save result
        out_ws.append([
            firm_name,
            base_url,
            "|".join(valid_sources) if valid_sources else "",
            "",  # profile_sources (same as list sources for now)
            "|".join(source_types) if source_types else "",
            ",".join(sorted(sample_fields)) if sample_fields else "",
            f"{len(valid_sources)} validated sources" if is_firm_valid else "No valid sources found",
            "TRUE" if is_firm_valid else "FALSE"
        ])
    
    wb.close()
    
    # Save outputs
    outputs_dir = Path(input_file).parent / "outputs"
    outputs_dir.mkdir(exist_ok=True)
    
    sources_file = outputs_dir / "sources_test.xlsx"
    out_wb.save(sources_file)
    
    # Save failure report
    if validator.failure_log:
        fail_wb = Workbook()
        fail_ws = fail_wb.active
        fail_ws.title = "Source Failures"
        fail_ws.append(["source_url", "source_type", "failure_type", "http_status", "reason", "timestamp"])
        
        for failure in validator.failure_log:
            fail_ws.append([
                failure["source_url"],
                failure["source_type"],
                failure["failure_type"],
                failure.get("http_status", ""),
                failure["reason"],
                failure["timestamp"]
            ])
        
        fail_file = outputs_dir / "source_failure_report_test.xlsx"
        fail_wb.save(fail_file)
        print(f"\n✓ Failure report saved: {fail_file}")
    
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Processed: {processed} firms")
    print(f"Validated: {validated} firms ({validated/max(processed,1)*100:.1f}%)")
    print(f"Output: {sources_file}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    input_file = "Company list_with_websites.xlsx"
    max_firms = 3  # Test first 3 firms: Kirkland, Latham, Skadden
    
    test_firm_finder(input_file, max_firms=max_firms, timeout=45)
