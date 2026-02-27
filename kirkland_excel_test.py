"""
End-to-end Excel output test for Kirkland.
Fetches 10 real URLs from letter A scroll, enriches them, saves to Excel.
"""
import asyncio
import time
import openpyxl
from pathlib import Path
from playwright.async_api import async_playwright

EXCEL_FILE = "AmLaw200_2025 Rank_gross revenue_with_websites.xlsx"
TAB_NAME = "Kirkland_Test"

COLUMNS = [
    "full_name", "title", "offices", "department",
    "practice_areas", "industries", "bar_admissions", "education",
    "profile_url", "extraction_status",
]


async def get_real_urls(n=10):
    """Scroll Kirkland /lawyers?letter=A and return the first N real profile URLs."""
    urls = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        await page.goto("https://www.kirkland.com/lawyers?letter=A", wait_until="domcontentloaded")
        await page.wait_for_selector(".person-result", timeout=15000)

        while len(urls) < n:
            cards = await page.query_selector_all(".person-result__name a")
            for card in cards:
                href = await card.get_attribute("href")
                if href and href not in urls:
                    full = f"https://www.kirkland.com{href}" if href.startswith("/") else href
                    if full not in urls:
                        urls.append(full)
            if len(urls) >= n:
                break
            btn = await page.query_selector(".search-results__load-more")
            if btn and await btn.is_visible():
                await btn.click()
                await page.wait_for_timeout(2000)
            else:
                break

        await browser.close()
    return urls[:n]


def fmt(val):
    if val is None:
        return ""
    if isinstance(val, list):
        parts = []
        for item in val:
            if hasattr(item, '__iter__') and not isinstance(item, str):
                parts.append(" | ".join(str(x) for x in item))
            else:
                parts.append(str(item))
        return "; ".join(parts)
    return str(val)


def main():
    print("Fetching 10 real Kirkland profile URLs...")
    urls = asyncio.run(get_real_urls(10))
    print(f"Got {len(urls)} URLs:")
    for u in urls:
        print(f"  {u}")

    from multi_mode_extractor import MultiModeExtractor
    extractor = MultiModeExtractor(enable_playwright=True)
    results = []

    print(f"\nEnriching {len(urls)} profiles...\n")
    for url in urls:
        t0 = time.time()
        profile = extractor.extract_profile("Kirkland & Ellis", url, force_playwright=True)
        elapsed = time.time() - t0
        results.append(profile)
        edu_str = "; ".join(
            f"{e.school} ({e.degree}, {e.year})" for e in (profile.education or [])
        )
        status = profile.extraction_status or "UNKNOWN"
        name = profile.full_name or "(no name)"
        print(f"[{status:8s}] {name:45s} offices={profile.offices}  bar={profile.bar_admissions}")
        if edu_str:
            print(f"           EDU: {edu_str}")

    # --- Write to Excel ---
    xlsx_path = Path(EXCEL_FILE)
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()

    if TAB_NAME in wb.sheetnames:
        del wb[TAB_NAME]

    ws = wb.create_sheet(TAB_NAME)
    ws.append(COLUMNS)

    for p in results:
        edu_formatted = fmt(
            [(e.school, e.degree, e.year) for e in (p.education or [])]
        )
        row = [
            fmt(p.full_name),
            fmt(p.title),
            fmt(p.offices),
            fmt(p.department),
            fmt(p.practice_areas),
            fmt(getattr(p, 'industries', None)),
            fmt(p.bar_admissions),
            edu_formatted,
            fmt(p.profile_url),
            fmt(p.extraction_status),
        ]
        ws.append(row)

    wb.save(xlsx_path)

    success = sum(1 for p in results if p.extraction_status == "SUCCESS")
    partial = sum(1 for p in results if p.extraction_status == "PARTIAL")
    failed  = sum(1 for p in results if p.extraction_status not in ("SUCCESS", "PARTIAL"))
    print(f"\nSaved {len(results)} rows to '{EXCEL_FILE}' tab '{TAB_NAME}'")
    print(f"SUCCESS: {success}  PARTIAL: {partial}  FAILED: {failed}")

    print(f"\n--- Excel Tab Preview ({TAB_NAME}) ---")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            print("  " + " | ".join(str(c)[:20] for c in row))
            print("  " + "-" * 130)
        else:
            print("  " + " | ".join(str(c or "")[:20] for c in row))


if __name__ == "__main__":
    main()
