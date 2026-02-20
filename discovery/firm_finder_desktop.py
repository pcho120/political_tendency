#!/usr/bin/env python3
"""
SUCCESS-FIRST Source Discovery and Validation
Discovers, tests, and validates attorney list sources before extraction
Outputs only SCRAPABLE sources with proof of extraction capability
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json
import time
import re
import os
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from urllib.parse import urlparse, urljoin
import xml.etree.ElementTree as ET

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# Load configuration
CONFIG_DIR = Path(__file__).parent.parent / "config"
with open(CONFIG_DIR / "blocklist.json") as f:
    BLOCKLIST = json.load(f)
with open(CONFIG_DIR / "known_patterns.json") as f:
    KNOWN_PATTERNS = json.load(f)

# Constants
DEFAULT_FIRM_TIMEOUT = 45  # seconds per firm
DEFAULT_PROFILE_SAMPLE_SIZE = 3
VALIDATION_THRESHOLD_NAME = 0.67  # 2/3 profiles must have name
VALIDATION_THRESHOLD_FIELD = 0.33  # 1/3 must have title/office/practice

class SourceValidator:
    """Validates sources by sampling profile extraction"""
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        self.failure_log = []
    
    def validate_source(self, source_url: str, source_type: str, base_url: str) -> Tuple[bool, Dict]:
        """
        Validate source by sampling 3 profiles
        Returns: (is_valid, validation_results)
        """
        result = {
            "source_url": source_url,
            "source_type": source_type,
            "tested_profiles": 0,
            "names_found": 0,
            "fields_found": 0,  # title, office, or practice
            "sample_pass_fields": set(),
            "failure_reason": None
        }
        
        try:
            # Get profile URLs from source
            profile_urls = self._get_sample_profiles(source_url, source_type, base_url)
            
            if not profile_urls:
                result["failure_reason"] = "no_profiles_found"
                self._log_failure(source_url, source_type, "NO_PROFILES", 200, "Source yielded 0 profile URLs")
                return False, result
            
            # Test each profile
            for profile_url in profile_urls[:DEFAULT_PROFILE_SAMPLE_SIZE]:
                result["tested_profiles"] += 1
                
                # Try extraction
                extracted = self._extract_profile_quick(profile_url)
                
                if extracted.get("name"):
                    result["names_found"] += 1
                
                if any(extracted.get(f) for f in ["title", "office", "practice"]):
                    result["fields_found"] += 1
                    for field in ["title", "office", "practice"]:
                        if extracted.get(field):
                            result["sample_pass_fields"].add(field)
            
            # Check thresholds
            name_ratio = result["names_found"] / max(result["tested_profiles"], 1)
            field_ratio = result["fields_found"] / max(result["tested_profiles"], 1)
            
            is_valid = (name_ratio >= VALIDATION_THRESHOLD_NAME and 
                       field_ratio >= VALIDATION_THRESHOLD_FIELD)
            
            if not is_valid:
                reason = f"low_extraction_rate (name:{name_ratio:.1%}, fields:{field_ratio:.1%})"
                result["failure_reason"] = reason
                self._log_failure(source_url, source_type, "LOW_EXTRACTION", 200, reason)
            
            return is_valid, result
            
        except Exception as e:
            result["failure_reason"] = f"validation_error: {str(e)}"
            self._log_failure(source_url, source_type, "VALIDATION_ERROR", None, str(e))
            return False, result
    
    def _get_sample_profiles(self, source_url: str, source_type: str, base_url: str) -> List[str]:
        """Extract sample profile URLs from source"""
        profiles = []
        
        try:
            if source_type == "sitemap":
                profiles = self._extract_from_sitemap(source_url, base_url)
            elif source_type == "directory":
                profiles = self._extract_from_directory(source_url, base_url)
            elif source_type == "alphabet":
                profiles = self._extract_from_alphabet(source_url, base_url)
        except Exception as e:
            self._log_failure(source_url, source_type, "SAMPLE_ERROR", None, str(e))
        
        return profiles[:10]  # Return up to 10 for sampling
    
    def _extract_from_sitemap(self, sitemap_url: str, base_url: str) -> List[str]:
        """Extract profile URLs from sitemap (handles sitemap indexes)"""
        profiles = []
        
        try:
            resp = self.session.get(sitemap_url, timeout=10)
            if resp.status_code != 200:
                return profiles
            
            root = ET.fromstring(resp.content)
            
            # Check if this is a sitemap index
            is_sitemap_index = any(elem.tag.endswith("}sitemap") or elem.tag == "sitemap" 
                                   for elem in root.iter())
            
            if is_sitemap_index:
                # Extract sub-sitemap URLs and look for people/professionals sitemaps
                for url_elem in root.iter():
                    if url_elem.tag.endswith("}loc") or url_elem.tag == "loc":
                        sub_sitemap_url = url_elem.text
                        if sub_sitemap_url and any(kw in sub_sitemap_url.lower() 
                                                   for kw in ['people', 'professional', 'attorney', 'lawyer']):
                            # Recursively extract from sub-sitemap
                            profiles.extend(self._extract_from_sitemap(sub_sitemap_url, base_url))
            else:
                # Regular sitemap - extract profile URLs
                for url_elem in root.iter():
                    if url_elem.tag.endswith("}loc") or url_elem.tag == "loc":
                        url = url_elem.text
                        if url:
                            # Canonicalize BEFORE profile check
                            canonical = self._canonicalize_source(url, base_url)
                            url_to_check = canonical if canonical else url
                            
                            if self._is_profile_url(url_to_check, base_url):
                                profiles.append(url_to_check)
        except Exception:
            pass
        
        return profiles
    
    def _extract_from_directory(self, dir_url: str, base_url: str) -> List[str]:
        """Extract profile URLs from directory page"""
        profiles = []
        
        try:
            resp = self.session.get(dir_url, timeout=10)
            if resp.status_code != 200:
                return profiles
            
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            for link in soup.find_all('a', href=True):
                url = urljoin(base_url, link['href'])
                
                # Canonicalize BEFORE profile check
                canonical = self._canonicalize_source(url, base_url)
                url_to_check = canonical if canonical else url
                
                if self._is_profile_url(url_to_check, base_url):
                    profiles.append(url_to_check)
        except Exception:
            pass
        
        return profiles
    
    def _extract_from_alphabet(self, alpha_url: str, base_url: str) -> List[str]:
        """Extract profiles from alphabet page"""
        return self._extract_from_directory(alpha_url, base_url)
    
    def _is_profile_url(self, url: str, base_url: str) -> bool:
        """Check if URL looks like a profile"""
        try:
            parsed = urlparse(url)
            base_parsed = urlparse(base_url)
            
            # Same domain
            if parsed.netloc != base_parsed.netloc:
                return False
            
            url_lower = url.lower()
            
            # Must have profile keywords
            if not any(kw in url_lower for kw in KNOWN_PATTERNS["profile_keywords"]):
                return False
            
            # Must not have junk patterns
            if any(junk in url_lower for junk in BLOCKLIST["junk_url_patterns"]):
                return False
            
            return True
        except:
            return False
    
    def _extract_profile_quick(self, url: str) -> Dict:
        """Quick extraction attempt from profile page"""
        extracted = {"name": None, "title": None, "office": None, "practice": None}
        
        try:
            resp = self.session.get(url, timeout=5)
            if resp.status_code != 200:
                return extracted
            
            html = resp.text
            soup = BeautifulSoup(html, 'html.parser')
            
            # Extract name (h1 or title)
            h1 = soup.find('h1')
            if h1:
                name = h1.get_text(strip=True)
                if self._looks_like_name(name):
                    extracted["name"] = name
            
            # Extract title (common patterns)
            for pattern in [r'(Partner|Associate|Counsel|Of Counsel)', r'class="title"']:
                if re.search(pattern, html, re.IGNORECASE):
                    extracted["title"] = "found"
                    break
            
            # Extract office (location patterns)
            if re.search(r'(Office|Location|New York|California|Texas)', html, re.IGNORECASE):
                extracted["office"] = "found"
            
            # Extract practice (practice area patterns)
            if re.search(r'(Practice|Litigation|Corporate|Tax)', html, re.IGNORECASE):
                extracted["practice"] = "found"
        
        except Exception:
            pass
        
        return extracted
    
    def _looks_like_name(self, text: str) -> bool:
        """Check if text looks like a person's name"""
        if not text or len(text) < 4 or len(text) > 100:
            return False
        parts = text.split()
        return len(parts) >= 2 and all(p[0].isupper() for p in parts if p)
    
    def _canonicalize_source(self, source_url: str, base_url: str) -> Optional[str]:
        """
        Replace bad domains (cm-prod, staging, CDN) with official domain
        Example: www-cm-prod.lw.com → www.lw.com
        Returns: canonical URL if replacement needed, None otherwise
        """
        parsed = urlparse(source_url)
        base_parsed = urlparse(base_url)
        
        # Check if domain contains bad patterns
        has_bad_domain = any(bad in parsed.netloc for bad in BLOCKLIST["bad_domains"])
        
        if has_bad_domain:
            # Replace netloc with official domain
            canonical = source_url.replace(parsed.netloc, base_parsed.netloc)
            return canonical
        
        return None
    
    def _log_failure(self, source_url: str, source_type: str, failure_type: str, 
                     http_status: Optional[int], reason: str):
        """Log source failure"""
        self.failure_log.append({
            "source_url": source_url,
            "source_type": source_type,
            "failure_type": failure_type,
            "http_status": http_status,
            "reason": reason,
            "timestamp": datetime.now().isoformat()
        })


class SourceDiscovery:
    """Discovers candidate sources for attorney listings"""
    
    def __init__(self, base_url: str, firm_name: str, timeout: int = DEFAULT_FIRM_TIMEOUT):
        self.base_url = base_url
        self.firm_name = firm_name
        self.timeout = timeout
        self.start_time = None
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
    
    def discover_sources(self) -> List[Dict]:
        """Discover all candidate sources within time budget"""
        self.start_time = time.time()
        candidates = []
        
        # Phase 1: robots.txt sitemaps
        if not self._timeout_exceeded():
            candidates.extend(self._discover_from_robots())
        
        # Phase 2: Known sitemap paths
        if not self._timeout_exceeded():
            candidates.extend(self._discover_known_sitemaps())
        
        # Phase 3: Directory paths
        if not self._timeout_exceeded():
            candidates.extend(self._discover_directories())
        
        # Phase 4: Alphabet patterns
        if not self._timeout_exceeded():
            candidates.extend(self._discover_alphabet_patterns())
        
        return candidates
    
    def _timeout_exceeded(self) -> bool:
        """Check if firm timeout exceeded"""
        if not self.start_time:
            return False
        return (time.time() - self.start_time) > self.timeout
    
    def _discover_from_robots(self) -> List[Dict]:
        """Discover sitemaps from robots.txt"""
        candidates = []
        
        try:
            robots_url = urljoin(self.base_url, "/robots.txt")
            resp = self.session.get(robots_url, timeout=5)
            
            if resp.status_code == 200:
                for line in resp.text.split('\n'):
                    if line.lower().startswith('sitemap:'):
                        sitemap_url = line.split(':', 1)[1].strip()
                        
                        # Canonicalize if needed (replace bad domains)
                        parsed = urlparse(sitemap_url)
                        base_parsed = urlparse(self.base_url)
                        has_bad_domain = any(bad in parsed.netloc for bad in BLOCKLIST["bad_domains"])
                        
                        if has_bad_domain:
                            sitemap_url = sitemap_url.replace(parsed.netloc, base_parsed.netloc)
                        
                        candidates.append({
                            "url": sitemap_url,
                            "type": "sitemap",
                            "source": "robots.txt"
                        })
        except Exception:
            pass
        
        return candidates
    
    def _discover_known_sitemaps(self) -> List[Dict]:
        """Try known sitemap paths"""
        candidates = []
        
        for path in KNOWN_PATTERNS["sitemap_paths"]:
            url = urljoin(self.base_url, path)
            
            try:
                resp = self.session.head(url, timeout=5, allow_redirects=True)
                if resp.status_code == 200:
                    candidates.append({
                        "url": url,
                        "type": "sitemap",
                        "source": "known_path"
                    })
            except Exception:
                pass
            
            if self._timeout_exceeded():
                break
        
        return candidates
    
    def _discover_directories(self) -> List[Dict]:
        """Try known directory paths"""
        candidates = []
        
        for path in KNOWN_PATTERNS["directory_paths"]:
            url = urljoin(self.base_url, path)
            
            try:
                resp = self.session.head(url, timeout=5, allow_redirects=True)
                if resp.status_code == 200:
                    candidates.append({
                        "url": url,
                        "type": "directory",
                        "source": "known_directory"
                    })
            except Exception:
                pass
            
            if self._timeout_exceeded():
                break
        
        return candidates
    
    def _discover_alphabet_patterns(self) -> List[Dict]:
        """Detect alphabet navigation patterns"""
        candidates = []
        
        # Check if site uses alphabet navigation
        for base_pattern in ["/professionals/", "/lawyers/", "/attorneys/", "/people/"]:
            test_url = urljoin(self.base_url, f"{base_pattern}a")
            
            try:
                resp = self.session.head(test_url, timeout=5, allow_redirects=True)
                if resp.status_code == 200:
                    # Add alphabet URLs
                    for letter in "abcdefghijklmnopqrstuvwxyz":
                        candidates.append({
                            "url": urljoin(self.base_url, f"{base_pattern}{letter}"),
                            "type": "alphabet",
                            "source": "alphabet_enumeration"
                        })
                    break  # Only need one alphabet pattern
            except Exception:
                pass
            
            if self._timeout_exceeded():
                break
        
        return candidates


class FirmSourceFinderApp:
    """GUI Application for source discovery and validation"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Law Firm Source Finder & Validator")
        self.root.geometry("700x550")
        self.root.resizable(True, True)
        
        self.validator = SourceValidator()
        self.current_file = None
        self.processing_thread = None
        
        self._setup_ui()
    
    def _setup_ui(self):
        """Setup UI components"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
        # Header
        header = ttk.Label(main_frame, text="Source Discovery & Validation", 
                          font=("Arial", 16, "bold"))
        header.grid(row=0, column=0, pady=(0, 20))
        
        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="Input File", padding="10")
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        
        self.file_path_var = tk.StringVar(value="No file selected")
        file_label = ttk.Label(file_frame, textvariable=self.file_path_var,
                               relief="sunken", anchor="w", padding="5")
        file_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        browse_btn = ttk.Button(file_frame, text="Browse...", command=self.browse_file)
        browse_btn.grid(row=1, column=0)
        
        # Settings
        settings_frame = ttk.LabelFrame(main_frame, text="Settings", padding="10")
        settings_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(settings_frame, text="Timeout per firm (seconds):").grid(row=0, column=0, sticky=tk.W)
        self.timeout_var = tk.IntVar(value=45)
        ttk.Spinbox(settings_frame, from_=15, to=120, textvariable=self.timeout_var, width=8).grid(row=0, column=1, padx=10)
        
        ttk.Label(settings_frame, text="Max firms (0=all):").grid(row=1, column=0, sticky=tk.W, pady=(5,0))
        self.max_firms_var = tk.IntVar(value=0)
        ttk.Spinbox(settings_frame, from_=0, to=1000, textvariable=self.max_firms_var, width=8).grid(row=1, column=1, padx=10, pady=(5,0))
        
        # Progress
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        progress_frame.rowconfigure(0, weight=1)
        
        self.progress_text = tk.Text(progress_frame, height=15, state="disabled", wrap="word")
        self.progress_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        scrollbar = ttk.Scrollbar(progress_frame, orient="vertical", command=self.progress_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.progress_text.config(yscrollcommand=scrollbar.set)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0)
        
        self.process_btn = ttk.Button(button_frame, text="Discover & Validate Sources",
                                      command=self.start_processing)
        self.process_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(button_frame, text="Clear", command=self.clear).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Exit", command=self.root.quit).pack(side=tk.LEFT)
        
        self._log("Ready. Select input Excel file with 'Firm' and 'Official Website' columns.\n")
    
    def browse_file(self):
        """Browse for input file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                wb.close()
                
                if "Firm" not in headers or "Official Website" not in headers:
                    messagebox.showerror("Invalid File",
                        "File must have 'Firm' and 'Official Website' columns")
                    return
                
                self.current_file = file_path
                self.file_path_var.set(os.path.basename(file_path))
                self._log(f"Loaded: {os.path.basename(file_path)}\n")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load file:\n{e}")
    
    def start_processing(self):
        """Start source discovery"""
        if not self.current_file:
            messagebox.showwarning("No File", "Please select an Excel file first")
            return
        
        if self.processing_thread and self.processing_thread.is_alive():
            messagebox.showwarning("Processing", "Already processing")
            return
        
        self.process_btn.config(state="disabled")
        self.processing_thread = threading.Thread(target=self.process_firms)
        self.processing_thread.daemon = True
        self.processing_thread.start()
    
    def process_firms(self):
        """Process all firms"""
        try:
            # Load input
            wb = load_workbook(self.current_file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            firm_col = headers.index("Firm") + 1
            url_col = headers.index("Official Website") + 1
            
            # Create output workbook
            out_wb = Workbook()
            out_ws = out_wb.active
            out_ws.title = "Sources"
            out_ws.append([
                "Firm", "official_base_url", "attorney_list_sources", "profile_sources",
                "source_types", "sample_pass_fields", "notes", "is_valid"
            ])
            
            # Process firms
            max_firms = self.max_firms_var.get()
            timeout = self.timeout_var.get()
            processed = 0
            validated = 0
            
            for row in range(2, ws.max_row + 1):
                firm_name = ws.cell(row, firm_col).value
                base_url = ws.cell(row, url_col).value
                
                if not firm_name or not base_url:
                    continue
                
                processed += 1
                if max_firms > 0 and processed > max_firms:
                    break
                
                self._log(f"\n[{processed}] {firm_name}...\n")
                
                # Discover sources
                discovery = SourceDiscovery(str(base_url), str(firm_name), timeout)
                candidates = discovery.discover_sources()
                
                self._log(f"  Found {len(candidates)} candidate sources\n")
                
                # Validate sources
                valid_sources = []
                source_types = []
                sample_fields = set()
                
                for candidate in candidates:
                    is_valid, result = self.validator.validate_source(
                        candidate["url"], candidate["type"], str(base_url)
                    )
                    
                    if is_valid:
                        valid_sources.append(candidate["url"])
                        source_types.append(candidate["type"])
                        sample_fields.update(result["sample_pass_fields"])
                        self._log(f"  ✓ Validated: {candidate['type']} - {candidate['url'][:60]}...\n")
                
                # Save result
                is_valid = len(valid_sources) > 0
                if is_valid:
                    validated += 1
                
                out_ws.append([
                    firm_name,
                    base_url,
                    "|".join(valid_sources) if valid_sources else "",
                    "",  # profile_sources (same as list sources for now)
                    "|".join(source_types) if source_types else "",
                    ",".join(sorted(sample_fields)) if sample_fields else "",
                    f"{len(valid_sources)} validated sources" if is_valid else "No valid sources found",
                    "TRUE" if is_valid else "FALSE"
                ])
                
                self._log(f"  Result: {len(valid_sources)} valid sources\n")
            
            # Save outputs
            outputs_dir = Path(self.current_file).parent / "outputs"
            outputs_dir.mkdir(exist_ok=True)
            
            sources_file = outputs_dir / "sources.xlsx"
            out_wb.save(sources_file)
            
            # Save failure report
            if self.validator.failure_log:
                fail_wb = Workbook()
                fail_ws = fail_wb.active
                fail_ws.title = "Source Failures"
                fail_ws.append(["source_url", "source_type", "failure_type", "http_status", "reason", "timestamp"])
                
                for failure in self.validator.failure_log:
                    fail_ws.append([
                        failure["source_url"],
                        failure["source_type"],
                        failure["failure_type"],
                        failure.get("http_status", ""),
                        failure["reason"],
                        failure["timestamp"]
                    ])
                
                fail_file = outputs_dir / "source_failure_report.xlsx"
                fail_wb.save(fail_file)
                self._log(f"\nFailure report saved: {fail_file}\n")
            
            self._log(f"\n{'='*50}\n")
            self._log(f"COMPLETE!\n")
            self._log(f"Processed: {processed} firms\n")
            self._log(f"Validated: {validated} firms\n")
            self._log(f"Output: {sources_file}\n")
            
            messagebox.showinfo("Complete",
                f"Processed {processed} firms\n"
                f"Validated {validated} firms\n\n"
                f"Output: {sources_file}")
            
        except Exception as e:
            self._log(f"\nERROR: {e}\n")
            messagebox.showerror("Error", f"Processing failed:\n{e}")
            import traceback
            traceback.print_exc()
        
        finally:
            self.root.after(0, lambda: self.process_btn.config(state="normal"))
    
    def clear(self):
        """Clear progress"""
        self.progress_text.config(state="normal")
        self.progress_text.delete(1.0, tk.END)
        self.progress_text.config(state="disabled")
        self._log("Cleared. Ready for new processing.\n")
    
    def _log(self, msg: str):
        """Log message to progress text"""
        self.root.after(0, lambda: self._log_impl(msg))
    
    def _log_impl(self, msg: str):
        """Implementation of logging"""
        self.progress_text.config(state="normal")
        self.progress_text.insert(tk.END, msg)
        self.progress_text.see(tk.END)
        self.progress_text.config(state="disabled")


def main():
    root = tk.Tk()
    app = FirmSourceFinderApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
