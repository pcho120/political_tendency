import time
import re
import streamlit as st
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote_plus, urlparse, parse_qs, unquote
from openpyxl import load_workbook
import pandas as pd

st.set_page_config(page_title="Firm Website Finder", layout="centered")
st.title("Firm Official Website Auto-Filler")

BLOCKED_DOMAINS = [
    "wikipedia.org", "linkedin.com", "facebook.com", "instagram.com", "twitter.com",
    "law.com", "vault.com", "chambers.com", "bloomberg.com", "crunchbase.com"
]

def normalize_homepage(url: str) -> str:
    """Return scheme://netloc (homepage)"""
    try:
        p = urlparse(url)
        if not p.scheme or not p.netloc:
            return ""
        return f"{p.scheme}://{p.netloc}"
    except Exception:
        return ""

def looks_blocked(url: str) -> bool:
    try:
        netloc = urlparse(url).netloc.lower()
        return any(b in netloc for b in BLOCKED_DOMAINS)
    except Exception:
        return True

def search_firm_website(firm_name: str) -> str:
    """
    Search for firm website using pattern matching and minimal web requests.
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "DNT": "1",
    }
    
    # Method 1: Pattern matching based on firm name
    try:
        # Clean the firm name for domain generation
        clean_name = firm_name.lower()
        clean_name = clean_name.replace("&", "and")
        clean_name = clean_name.replace(",", "")
        clean_name = re.sub(r'[^a-z0-9\s]', '', clean_name)
        
        # Extract key parts of the name
        words = clean_name.split()
        primary_name = words[0] if words else clean_name
        full_name_no_spaces = clean_name.replace(" ", "")
        
        # Known law firm domains mapping
        known_mappings = {
            "kirkland & ellis": "https://www.kirkland.com",
            "kirkland": "https://www.kirkland.com", 
            "latham & watkins": "https://www.lw.com",
            "latham": "https://www.lw.com",
            "skadden arps": "https://www.skadden.com",
            "skadden": "https://www.skadden.com",
            "gibson dunn": "https://www.gibsondunn.com",
            "gibson": "https://www.gibsondunn.com",
            "sidley austin": "https://www.sidley.com",
            "sidley": "https://www.sidley.com",
        }
        
        # Check known mappings first
        firm_lower = firm_name.lower()
        for key, url in known_mappings.items():
            if key in firm_lower:
                return url
        
        # Generate potential domains in order of likelihood
        potential_domains = []
        
        # Try combinations with full name
        potential_domains.extend([
            f"https://www.{full_name_no_spaces}.com",
            f"https://{full_name_no_spaces}.com",
            f"https://www.{full_name_no_spaces}law.com",
            f"https://{full_name_no_spaces}law.com",
            f"https://www.{full_name_no_spaces}llp.com", 
            f"https://{full_name_no_spaces}llp.com",
            f"https://www.{full_name_no_spaces}firm.com",
            f"https://{full_name_no_spaces}firm.com",
        ])
        
        # Try with primary name
        potential_domains.extend([
            f"https://www.{primary_name}law.com",
            f"https://{primary_name}law.com",
            f"https://www.{primary_name}llp.com",
            f"https://{primary_name}llp.com",
            f"https://www.{primary_name}.com",
            f"https://{primary_name}.com",
        ])
        
        # Try with first two words
        if len(words) >= 2:
            first_two = "".join(words[:2])
            potential_domains.extend([
                f"https://www.{first_two}law.com",
                f"https://{first_two}law.com",
                f"https://www.{first_two}.com",
                f"https://{first_two}.com",
            ])
        
        # Test each domain (limit to avoid too many requests)
        tested = set()
        for domain in potential_domains[:15]:  # Limit to 15 requests max
            if domain in tested:
                continue
            tested.add(domain)
            
            try:
                # Quick HEAD request to check if domain exists
                response = requests.head(domain, headers=headers, timeout=3, allow_redirects=True)
                if response.status_code in [200, 301, 302]:
                    print(f"Found: {domain}")
                    return normalize_homepage(domain)
            except requests.RequestException:
                continue
            except Exception:
                continue
                
    except Exception as e:
        print(f"Pattern matching failed: {e}")
    
    # Method 2: If pattern matching fails, try one web search
    try:
        print(f"Fallback search for: {firm_name}")
        search_url = f"https://duckduckgo.com/html/?q={quote_plus(f'\"{firm_name}\" law firm website')}"
        
        response = requests.get(search_url, headers=headers, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, "html.parser")
            links = soup.find_all("a")
            
            for link in links:
                href = str(link.get("href", ""))
                text = link.get_text(strip=True)
                
                if not href or "duckduckgo.com" in href:
                    continue
                    
                # Convert to absolute URL
                if href.startswith("//"):
                    href = "https:" + href
                elif href.startswith("/"):
                    href = "https://duckduckgo.com" + href
                
                if looks_blocked(href):
                    continue
                
                # Look for law firm patterns
                if any(keyword in text.lower() for keyword in ["law", "firm", "llp", "attorney"]) or \
                   any(keyword in href.lower() for keyword in ["law", "firm", "llp"]):
                    homepage = normalize_homepage(href)
                    if homepage:
                        print(f"Search found: {homepage}")
                        return homepage
                        
    except Exception as e:
        print(f"Search fallback failed: {e}")
    
    print(f"Not found: {firm_name}")
    return ""

def process_excel_file(uploaded_file):
    """Process the uploaded Excel file."""
    try:
        # Read Excel with pandas first to understand structure
        df = pd.read_excel(uploaded_file)
        
        if 'Firm' not in df.columns:
            st.error("Could not find the 'Firm' column in the Excel file.")
            return None
        
        # Load with openpyxl to preserve formatting
        wb = load_workbook(uploaded_file)
        ws = wb.active
        
        # Find the Firm column
        firm_col = None
        header_row = 1
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if isinstance(cell_value, str) and cell_value.strip().lower() == "firm":
                firm_col = col
                break
        
        if not firm_col:
            st.error("Could not find the 'Firm' column header.")
            return None
        
        # Add Official Website column at the end
        last_col = ws.max_column + 1
        ws.cell(row=header_row, column=last_col, value="Official Website")
        
        # Process each firm
        total_processed = 0
        websites_found = 0
        failed_firms = []
        
        progress = st.progress(0)
        status_text = st.empty()
        
        for row in range(header_row + 1, ws.max_row + 1):
            firm_cell = ws.cell(row=row, column=firm_col)
            firm_name = firm_cell.value
            
            if firm_name is None or str(firm_name).strip() == "":
                continue
            
            total_processed += 1
            firm_name = str(firm_name).strip()
            
            status_text.text(f"Searching: {firm_name} ({total_processed})")
            progress.progress(min(1.0, total_processed / ws.max_row))
            
            # Search for website
            website = search_firm_website(firm_name)
            
            if website:
                ws.cell(row=row, column=last_col, value=website)
                websites_found += 1
            else:
                ws.cell(row=row, column=last_col, value="")  # Leave blank if not found
                failed_firms.append(firm_name)
            
            # Small delay to be respectful
            time.sleep(0.5)
        
        status_text.text(f"Done! Found {websites_found} websites out of {total_processed} firms.")
        
        # Show failed firms if any
        if failed_firms:
            with st.expander("Firms without websites found"):
                st.write(failed_firms)
        
        return wb
        
    except Exception as e:
        st.error(f"Error processing file: {e}")
        return None

# Streamlit UI
uploaded_file = st.file_uploader("Upload Excel (.xlsx) with 'Firm' column", type=["xlsx"])

rate_limit = st.slider("Delay between searches (seconds)", 0.2, 3.0, 0.5, 0.1)
max_rows = st.number_input("Max firms to process (for testing)", min_value=1, value=50, step=10)

run_btn = st.button("Find Official Websites")

if uploaded_file and run_btn:
    # Process the file
    processed_wb = process_excel_file(uploaded_file)
    
    if processed_wb:
        # Save the processed file
        output_filename = "company_list_with_websites.xlsx"
        processed_wb.save(output_filename)
        
        # Provide download button
        with open(output_filename, "rb") as f:
            st.download_button(
                label="Download Updated Excel File",
                data=f,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        st.success(f"Processing complete! Download your updated file above.")