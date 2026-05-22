#!/usr/bin/env python3
"""
Firm Website Finder - Deterministic, cached, auditable discovery

Goal: near-100% official website discovery for law firms.
"""

from __future__ import annotations

import argparse
import difflib
import json
import os
import re
import socket
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook


CACHE_DIR = Path("cache")
CACHE_DIR.mkdir(exist_ok=True)
CACHE_FILE = CACHE_DIR / "firm_domain_cache.json"
ALIAS_FILE = CACHE_DIR / "firm_aliases.json"

OUTPUT_FAILURES = "discovery_failures.xlsx"

GOOGLE_CSE_API_KEY = os.getenv("GOOGLE_CSE_API_KEY", "").strip()
GOOGLE_CSE_CX = os.getenv("GOOGLE_CSE_CX", "").strip()
SERPAPI_KEY = os.getenv("SERPAPI_KEY", "").strip()

GOOGLE_RATE_LIMIT_SECONDS = 1.0
SITE_RATE_LIMIT_SECONDS = 0.5

DEFAULT_TIMEOUT = 5

BLOCKED_DOMAINS = {
    "wikipedia.org",
    "linkedin.com",
    "facebook.com",
    "instagram.com",
    "twitter.com",
    "law.com",
    "vault.com",
    "chambers.com",
    "bloomberg.com",
    "crunchbase.com",
    "justia.com",
    "martindale.com",
    "superlawyers.com",
    "lawyer.com",
    "lawyers.com",
    "amlaw.com",
    "law360.com",
    "law360",
}

NOISE_SUFFIXES = {
    "llp",
    "lllp",
    "pllc",
    "plc",
    "pc",
    "p.c.",
    "ltd.",
    "l.l.p",
    "l.l.c",
    "law firm",
    "verein",
    "international",
    "global",
    "the",
}

# Words stripped BEFORE building domain tokens (not from display name)
TOKEN_STOPWORDS = {"law", "firm", "lawyers", "legal", "group", "and", "of", "at", "llp", "pllc"}

# Known domains for BigLaw firms with non-obvious domains
KNOWN_DOMAINS: dict[str, str] = {
    "latham watkins": "lw.com",
    "latham": "lw.com",
    "dla piper": "dlapiper.com",
    "skadden arps slate meagher flom": "skadden.com",
    "skadden arps": "skadden.com",
    "skadden": "skadden.com",
    "sidley austin": "sidley.com",
    "sidley": "sidley.com",
    "ropes gray": "ropesgray.com",
    "morgan lewis bockius": "morganlewis.com",
    "morgan lewis": "morganlewis.com",
    "hogan lovells": "hoganlovells.com",
    "dentons": "dentons.com",
    "goodwin procter": "goodwinlaw.com",
    "goodwin": "goodwinlaw.com",
    "cooley": "cooley.com",
    "kirkland ellis": "kirkland.com",
    "kirkland": "kirkland.com",
    "paul weiss rifkind wharton garrison": "paulweiss.com",
    "paul weiss": "paulweiss.com",
    "weil gotshal manges": "weil.com",
    "weil": "weil.com",
    "sullivan cromwell": "sullcrom.com",
    "cravath swaine moore": "cravath.com",
    "cravath": "cravath.com",
    "simpson thacher bartlett": "stblaw.com",
    "simpson thacher": "stblaw.com",
    "davis polk wardwell": "davispolk.com",
    "davis polk": "davispolk.com",
    "cleary gottlieb steen hamilton": "clearygottlieb.com",
    "cleary gottlieb": "clearygottlieb.com",
    "white case": "whitecase.com",
    "jones day": "jonesday.com",
    "baker mckenzie": "bakermckenzie.com",
    "wilmer cutler pickering hale dorr": "wilmerhale.com",
    "wilmerhale": "wilmerhale.com",
    "wilmer cutler": "wilmerhale.com",
    "wilmer": "wilmerhale.com",
    "willkie farr gallagher": "willkie.com",
    "willkie": "willkie.com",
    "proskauer rose": "proskauer.com",
    "proskauer": "proskauer.com",
    "akin gump strauss hauer feld": "akingump.com",
    "akin gump": "akingump.com",
    "mayer brown": "mayerbrown.com",
    "o melveny myers": "omm.com",
    "omelveny": "omm.com",
    "gibson dunn crutcher": "gibsondunn.com",
    "gibson dunn": "gibsondunn.com",
    "orrick herrington sutcliffe": "orrick.com",
    "orrick": "orrick.com",
    "paul hastings": "paulhastings.com",
    "quinn emanuel urquhart sullivan": "quinnemanuel.com",
    "quinn emanuel": "quinnemanuel.com",
    "shearman sterling": "shearman.com",
    "freshfields bruckhaus deringer": "freshfields.com",
    "freshfields": "freshfields.com",
    "linklaters": "linklaters.com",
    "clifford chance": "cliffordchance.com",
    "allen overy": "allenovery.com",
    "norton rose fulbright": "nortonrosefulbright.com",
    "greenberg traurig": "gtlaw.com",
    "reed smith": "reedsmith.com",
    "crowell moring": "crowell.com",
    "foley lardner": "foley.com",
    "foley": "foley.com",
    "perkins coie": "perkinscoie.com",
    "morrison foerster": "mofo.com",
    "fenwick west": "fenwick.com",
    "wilson sonsini goodrich rosati": "wsgr.com",
    "wilson sonsini": "wsgr.com",
    "mintz levin": "mintz.com",
    "mintz": "mintz.com",
    "king spalding": "kslaw.com",
    "troutman pepper": "troutman.com",
    "troutman sanders": "troutman.com",
    "hunton andrews kurth": "huntonak.com",
    "dechert": "dechert.com",
    "squire patton boggs": "squirepb.com",
    "ballard spahr": "ballardspahr.com",
    "steptoe johnson": "steptoe.com",
    "steptoe": "steptoe.com",
    "covington burling": "cov.com",
    "covington": "cov.com",
    "venable": "venable.com",
    "vinson elkins": "velaw.com",
    "haynes boone": "haynesboone.com",
    "bracewell": "bracewell.com",
    "baker botts": "bakerbotts.com",
    "locke lord": "lockelord.com",
    "jackson walker": "jw.com",
    # --- Additional AmLaw 200 firms (all return 403 to scrapers) ---
    "debevoise plimpton": "debevoise.com",
    "debevoise": "debevoise.com",
    "alston bird": "alston.com",
    "alston and bird": "alston.com",
    "alston": "alston.com",
    "sheppard mullin": "sheppardmullin.com",
    "wachtell lipton rosen katz": "wlrk.com",
    "wachtell lipton": "wlrk.com",
    "wachtell": "wlrk.com",
    "mcguirewoods": "mcguirewoods.com",
    "mcguire woods": "mcguirewoods.com",
    "faegre drinker": "faegredrinker.com",
    "polsinelli": "polsinelli.com",
    "jenner block": "jenner.com",
    "jenner and block": "jenner.com",
    "womble bond dickinson": "wbd.com",
    "womble bond": "wbd.com",
    "kilpatrick townsend": "kilpatricktownsend.com",
    "kilpatrick": "kilpatricktownsend.com",
    "susman godfrey": "susmangodfrey.com",
    "fish richardson": "fr.com",
    "fish": "fr.com",
    "dorsey whitney": "dorsey.com",
    "dorsey": "dorsey.com",
    "bradley arant": "bradley.com",
    "bradley": "bradley.com",
    "loeb loeb": "loeb.com",
    "loeb and loeb": "loeb.com",
    "choate hall": "choate.com",
    "choate": "choate.com",
    "vedder price": "vedderprice.com",
    "vedder": "vedderprice.com",
    "finnegan henderson": "finnegan.com",
    "finnegan": "finnegan.com",
    "buchalter": "buchalter.com",
    "kutak rock": "kutakrock.com",
    "kutak": "kutakrock.com",
    "patterson belknap webb tyler": "pbwt.com",
    "patterson belknap": "pbwt.com",
    "kelley drye": "kelleydrye.com",
    "carlton fields": "carltonfields.com",
    "lathrop gpm": "lathropgpm.com",
    "shumaker loop": "shumaker.com",
    "shumaker": "shumaker.com",
    "fennemore craig": "fennemorelaw.com",
    "fennemore": "fennemorelaw.com",
    "greenspoon marder": "greenspoonlaw.com",
    "lewis roca": "lewisroca.com",
    "arnall golden gregory": "agg.com",
    "arnall golden": "agg.com",
    "grayrobinson": "gray-robinson.com",
    "gray robinson": "gray-robinson.com",
    "offit kurman": "offitkurman.com",
    "morris manning": "mmmlaw.com",
    "bond schoeneck": "bsk.com",
    "ub greensfelder": "ubglaw.com",
    "miles stockbridge": "milesstockbridge.com",
    "miles and stockbridge": "milesstockbridge.com",
    "herrick feinstein": "herrick.com",
    "herrick": "herrick.com",
}



def _log(msg: str) -> None:
    print(f"[firm_finder] {msg}", flush=True)


@dataclass
class SearchResult:
    title: str
    link: str
    snippet: str
    display_link: str


def _load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_json(path: Path, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def normalize_firm_name(name: str) -> dict:
    if not name:
        return {
            'clean_name': '', 'tokens': [], 'full_tokens': [], 'content_tokens': [],
            'joined': '', 'hyphenated': '', 'first_token': '', 'second_token': '',
            'token_count': 0, 'acronym': '',
        }
    text = unicodedata.normalize('NFKC', str(name)).strip()
    text = text.lower()
    text = re.sub(r'\s*&\s*', ' and ', text)
    text = re.sub(r'\([^\)]*\)$', '', text).strip()
    text = re.sub(r"[\.,\'\"]", '', text)
    text = re.sub(r'\s+', ' ', text)
    for suffix in NOISE_SUFFIXES:
        text = re.sub(r'\b' + re.escape(suffix) + r'\b', '', text).strip()
    text = re.sub(r'\s+', ' ', text).strip()
    clean_name = text

    full_tokens = [t for t in clean_name.split() if t]
    content_tokens = [t for t in full_tokens if t not in TOKEN_STOPWORDS]
    first_token = content_tokens[0] if content_tokens else (full_tokens[0] if full_tokens else '')
    second_token = content_tokens[1] if len(content_tokens) >= 2 else ''
    token_count = len(full_tokens)
    joined = ''.join(full_tokens)
    hyphenated = '-'.join(full_tokens)
    acronym = ''.join(t[0] for t in content_tokens) if content_tokens else ''

    return {
        'clean_name': clean_name,
        'tokens': full_tokens,
        'full_tokens': full_tokens,
        'content_tokens': content_tokens,
        'joined': joined,
        'hyphenated': hyphenated,
        'first_token': first_token,
        'second_token': second_token,
        'token_count': token_count,
        'acronym': acronym,
    }


def generate_domain_candidates(norm: dict) -> list[str]:
    joined = norm['joined']
    full_tokens = norm['full_tokens']
    content_tokens = norm['content_tokens']
    first_token = norm['first_token']
    second_token = norm['second_token']
    acronym = norm['acronym']

    content_joined = ''.join(content_tokens)
    content_hyphenated = '-'.join(content_tokens)
    first_two_content = ''.join(content_tokens[:2]) if len(content_tokens) >= 2 else ''

    seen = []
    def add(d):
        if d and d not in seen and len(d) > 4:
            seen.append(d)

    add(f"{joined}.com")
    if content_joined != joined:
        add(f"{content_joined}.com")
    if len(content_tokens) >= 2 and '-' not in content_joined:
        add(f"{content_hyphenated}.com")
    if first_two_content and first_two_content != joined and first_two_content != content_joined:
        add(f"{first_two_content}.com")
    if first_token and second_token:
        combo = first_token + second_token
        if combo != joined and combo != content_joined:
            add(f"{combo}.com")
    add(f"{first_token}.com")
    if len(content_tokens) >= 2 and acronym:
        add(f"{acronym}.com")
    if first_token:
        add(f"{first_token}law.com")
    if content_joined and content_joined != first_token:
        add(f"{content_joined}law.com")
    if joined != content_joined and joined != first_token:
        add(f"{joined}law.com")
    add(f"{joined}.law")
    if first_token:
        add(f"{first_token}.law")

    return seen[:12]


def generate_aliases(normalized: str) -> list[str]:
    if not normalized:
        return []
    aliases = set()
    aliases.add(normalized)
    aliases.add(normalized.replace(' and ', ' '))
    aliases.add(normalized.replace(' ', ''))

    tokens = normalized.split()
    if len(tokens) >= 2:
        aliases.add(f"{tokens[0]} {tokens[-1]}")
        aliases.add(tokens[0])
    if 'bockius' in tokens:
        aliases.add(' '.join([t for t in tokens if t != 'bockius']))

    return list({a.strip() for a in aliases if a.strip()})[:10]


def _expected_domains(aliases: list[str]) -> set[str]:
    expected = set()
    for alias in aliases:
        token = alias.replace(' ', '')
        if len(token) < 4:
            continue
        expected.add(f"{token}.com")
        expected.add(f"www.{token}.com")
    return expected


def _rate_limit(state: dict, key: str, delay: float) -> None:
    last = state.get(key, 0.0)
    elapsed = time.time() - last
    if elapsed < delay:
        time.sleep(delay - elapsed)
    state[key] = time.time()


def google_cse_search(query: str, rate_state: dict) -> list[SearchResult]:
    if not GOOGLE_CSE_API_KEY or not GOOGLE_CSE_CX:
        return []
    _rate_limit(rate_state, 'google', GOOGLE_RATE_LIMIT_SECONDS)
    url = 'https://www.googleapis.com/customsearch/v1'
    params = {'key': GOOGLE_CSE_API_KEY, 'cx': GOOGLE_CSE_CX, 'q': query}
    try:
        resp = requests.get(url, params=params, timeout=DEFAULT_TIMEOUT)
    except Exception:
        return []
    if resp.status_code != 200:
        return []
    data = resp.json()
    items = data.get('items', [])[:5]
    results = []
    for item in items:
        results.append(SearchResult(
            title=item.get('title', ''),
            link=item.get('link', ''),
            snippet=item.get('snippet', ''),
            display_link=item.get('displayLink', ''),
        ))
    return results


def serpapi_search(query: str, rate_state: dict) -> list[SearchResult]:
    if not SERPAPI_KEY:
        return []
    _rate_limit(rate_state, 'google', GOOGLE_RATE_LIMIT_SECONDS)
    url = 'https://serpapi.com/search'
    params = {'engine': 'google', 'q': query, 'api_key': SERPAPI_KEY, 'num': 5}
    try:
        resp = requests.get(url, params=params, timeout=DEFAULT_TIMEOUT)
    except Exception:
        return []
    if resp.status_code != 200:
        return []
    data = resp.json()
    items = data.get('organic_results', [])[:5]
    results = []
    for item in items:
        results.append(SearchResult(
            title=item.get('title', ''),
            link=item.get('link', ''),
            snippet=item.get('snippet', ''),
            display_link=item.get('displayed_link', ''),
        ))
    return results


def _extract_domain(url: str) -> str:
    try:
        parsed = urlparse(url)
        netloc = parsed.netloc.lower()
        if netloc.startswith('www.'):
            netloc = netloc[4:]
        return netloc
    except Exception:
        return ''


def _looks_like_directory(domain: str) -> bool:
    return any(blocked in domain for blocked in BLOCKED_DOMAINS)


def score_candidate(
    domain: str,
    results: list[SearchResult],
    query_type: str,
    tokens: list[str],
    expected_domains: set[str],
) -> tuple[float, str]:
    score = 0.0
    evidence = []

    if any(tok in domain for tok in tokens):
        score += 0.10
        evidence.append('domain contains firm token')

    if domain in expected_domains or ('www.' + domain) in expected_domains:
        score += 0.10
        evidence.append('exact domain match')

    if results:
        top = results[0]
        if query_type == 'official' and domain == _extract_domain(top.link):
            score += 0.20
            evidence.append('top result for official query')

    appearances = sum(1 for r in results if _extract_domain(r.link) == domain)
    if appearances > 1:
        score += 0.15
        evidence.append('domain appears multiple times')

    joined_text = ' '.join([r.title + ' ' + r.snippet for r in results]).lower()
    if any(tok in joined_text for tok in tokens):
        score += 0.50
        evidence.append('brand tokens in title/snippet')

    if _looks_like_directory(domain):
        score -= 0.50
        evidence.append('blocked directory/news site')

    if domain.count('.') > 2:
        score -= 0.30
        evidence.append('subdomain on another platform')

    if domain.startswith('http'):
        score -= 0.20
        evidence.append('domain looks like full url')

    return score, '; '.join(evidence)


def _check_dns(domain: str) -> bool:
    import socket
    try:
        socket.getaddrinfo(domain, 80)
        return True
    except Exception:
        return False


def _is_parked_domain(html: str) -> bool:
    parked_signals = [
        'this domain is for sale',
        'domain parking',
        'buy this domain',
        'parked by',
        'godaddy.com/parking',
        'sedoparking.com',
        'domainmarket.com',
    ]
    lower = html.lower()
    return any(sig in lower for sig in parked_signals)


def verify_domain(domain: str, tokens: list[str], rate_state: dict) -> tuple[bool, str, str]:
    if not domain:
        return False, '', 'empty domain'
    if _looks_like_directory(domain):
        return False, '', 'blocked directory/news domain'
    if not _check_dns(domain):
        return False, '', 'dns resolution failed'

    _rate_limit(rate_state, domain, SITE_RATE_LIMIT_SECONDS)
    url = f'https://{domain}'
    try:
        resp = requests.get(url, timeout=DEFAULT_TIMEOUT, allow_redirects=True)
    except Exception as e:
        return False, '', f'request failed: {e}'

    if resp.status_code not in (200, 301, 302):
        return False, '', f'bad status {resp.status_code}'

    final_domain = _extract_domain(resp.url)
    if _looks_like_directory(final_domain):
        return False, '', 'redirected to directory/news site'

    html = resp.text
    if _is_parked_domain(html):
        return False, '', 'parked domain'

    html_lower = html.lower()
    title_match = re.search(r'<title[^>]*>(.*?)</title>', html_lower, re.IGNORECASE | re.DOTALL)
    title = title_match.group(1).strip() if title_match else ''
    nav_keywords = ['professionals', 'people', 'lawyers', 'attorneys', 'offices']
    nav_hit = any(k in html_lower for k in nav_keywords)

    if any(tok in title or tok in html_lower for tok in tokens):
        return True, resp.url, 'token evidence in homepage'

    if nav_hit:
        return True, resp.url, 'nav keywords present on homepage'

    for path in ['/about', '/people', '/professionals']:
        _rate_limit(rate_state, final_domain, SITE_RATE_LIMIT_SECONDS)
        try:
            sub = requests.get(f'https://{final_domain}{path}', timeout=DEFAULT_TIMEOUT)
            if sub.status_code == 200:
                if any(tok in sub.text.lower() for tok in tokens):
                    return True, sub.url, f'token evidence in {path}'
        except Exception:
            continue

    return False, resp.url, 'no brand evidence'


def score_domain(
    domain: str,
    html: str,
    final_url: str,
    norm: dict,
) -> int:
    score = 0
    first_token = norm['first_token']
    second_token = norm['second_token']
    full_tokens = norm['full_tokens']
    content_tokens = norm['content_tokens']

    html_lower = html.lower()
    domain_base = domain.split('.')[0].lower()

    title_match = re.search(r'<title[^>]*>(.*?)</title>', html_lower, re.DOTALL)
    title = title_match.group(1).strip() if title_match else ''

    meta_match = re.search(r'name=["\x27]description["\x27]\s+content=["\x27]([^"\x27]+)', html_lower)
    meta_desc = meta_match.group(1) if meta_match else ''

    og_match = re.search(r'property=["\x27]og:site_name["\x27]\s+content=["\x27]([^"\x27]+)', html_lower)
    og_site = og_match.group(1) if og_match else ''

    first_token_in_title = bool(first_token and first_token in title)
    if first_token_in_title:
        score += 4
    if second_token and second_token in title:
        score += 2
    if first_token and (first_token in meta_desc or first_token in og_site):
        score += 2

    law_signals = ['law firm', 'attorneys', 'lawyers', 'legal services', 'counsel']
    combined = title + ' ' + meta_desc
    if any(sig in combined for sig in ['law firm', 'attorneys', 'lawyers']):
        score += 2
    if any(sig in html_lower for sig in ['practice areas', 'our practice', 'areas of practice']):
        score += 2
    if any(sig in html_lower for sig in ['attorneys', 'lawyers']):
        score += 1

    if re.search(r'href=["\x27][^"\x27]*/about["\x27]', html_lower):
        score += 1
    if re.search(r'href=["\x27][^"\x27]*/contact["\x27]', html_lower):
        score += 1

    footer_text = ''
    footer_match = re.search(r'<footer[^>]*>(.*?)</footer>', html_lower, re.DOTALL)
    if footer_match:
        footer_text = footer_match.group(1)
    if first_token and first_token in footer_text:
        score += 2

    unrelated = ['real estate agency', 'restaurant', 'hospital', 'bank of', 'university of']
    if any(u in html_lower for u in unrelated):
        score -= 3

    if len(domain_base) <= 4:
        law_signals_present = any(sig in html_lower for sig in law_signals)
        if first_token_in_title and law_signals_present:
            pass
        elif not first_token_in_title:
            score -= 5
    else:
        import difflib
        best_ratio = max(
            (difflib.SequenceMatcher(None, domain_base, tok).ratio() for tok in full_tokens),
            default=0.0,
        )
        if best_ratio < 0.35 and not first_token_in_title:
            score -= 5

    return score


def _lookup_known_domain(norm: dict) -> str:
    clean = norm['clean_name']
    full_tokens = norm['full_tokens']
    content_tokens = norm['content_tokens']

    if clean in KNOWN_DOMAINS:
        return KNOWN_DOMAINS[clean]

    for n in range(len(full_tokens), 0, -1):
        prefix = ' '.join(full_tokens[:n])
        if prefix in KNOWN_DOMAINS:
            return KNOWN_DOMAINS[prefix]

    for n in range(len(content_tokens), 0, -1):
        prefix = ' '.join(content_tokens[:n])
        if prefix in KNOWN_DOMAINS:
            return KNOWN_DOMAINS[prefix]

    return ''


def _run_search(firm_name: str, rate_state: dict) -> list[str]:
    queries = [
        (f"{firm_name} official website", 'official'),
        (f"{firm_name} law firm", 'lawfirm'),
    ]
    seen = []
    for q, qtype in queries:
        if GOOGLE_CSE_API_KEY and GOOGLE_CSE_CX:
            results = google_cse_search(q, rate_state)
        elif SERPAPI_KEY:
            results = serpapi_search(q, rate_state)
        else:
            results = []
        for r in results:
            d = _extract_domain(r.link)
            if d and not _looks_like_directory(d) and d not in seen:
                seen.append(d)
    return seen


def _validate_candidate(domain: str, tokens: list[str], norm: dict, rate_state: dict) -> tuple[int, str, str]:
    if not domain or _looks_like_directory(domain):
        return -1, '', 'blocked'
    if not _check_dns(domain):
        return -1, '', 'dns failed'
    _rate_limit(rate_state, domain, SITE_RATE_LIMIT_SECONDS)
    url = f'https://{domain}'
    try:
        resp = requests.get(url, timeout=DEFAULT_TIMEOUT, allow_redirects=True)
    except Exception as e:
        return -1, '', f'request failed: {e}'
    if resp.status_code not in (200, 301, 302):
        return -1, '', f'bad status {resp.status_code}'
    final_domain = _extract_domain(resp.url)
    if _looks_like_directory(final_domain):
        return -1, '', 'redirected to directory'
    html = resp.text
    if _is_parked_domain(html):
        return -1, '', 'parked'
    s = score_domain(final_domain, html, resp.url, norm)
    return s, resp.url, 'scored'


def discover_official_domain(firm_name: str, refresh: bool, rate_state: dict, cache: dict, alias_cache: dict) -> dict:
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Step 0: KNOWN_DOMAINS lookup
    norm = normalize_firm_name(firm_name)
    known = _lookup_known_domain(norm)
    if known:
        domain = known if not known.startswith('http') else _extract_domain(known)
        # KNOWN_DOMAINS are authoritative — accept without HTTP verification
        # (many BigLaw sites return 403 to scrapers but are real)
        chosen_url = f'https://www.{domain}'
        cache[norm['clean_name']] = {
            'domain': domain,
            'chosen_url': chosen_url,
            'confidence': 1.0,
            'timestamp': datetime.utcnow().isoformat(),
            'evidence': 'known_domain table',
        }
        alias_cache[norm['clean_name']] = domain
        return {
            'firm': firm_name,
            'normalized_firm': norm['clean_name'],
            'official_website_url': chosen_url,
            'discovery_method': 'known_domain',
            'confidence_score': 1.0,
            'evidence': 'known_domain table',
            'notes': 'known domain (no verification needed)',
        }

    # Step 1: cache check
    cache_entry = cache.get(norm['clean_name'])
    if cache_entry and not refresh:
        try:
            ts = datetime.fromisoformat(cache_entry.get('timestamp'))
            if datetime.utcnow() - ts <= timedelta(days=30):
                return {
                    'firm': firm_name,
                    'normalized_firm': norm['clean_name'],
                    'official_website_url': cache_entry.get('chosen_url', ''),
                    'discovery_method': 'cache',
                    'confidence_score': cache_entry.get('confidence', 0.0),
                    'evidence': cache_entry.get('evidence', 'cache'),
                    'notes': 'cache hit',
                }
        except Exception:
            pass

    # Step 2: alias cache
    alias_hit = alias_cache.get(norm['clean_name'])
    if alias_hit:
        alias_url = alias_hit if alias_hit.startswith('http') else f'https://{alias_hit}'
        return {
            'firm': firm_name,
            'normalized_firm': norm['clean_name'],
            'official_website_url': alias_url,
            'discovery_method': 'alias_match',
            'confidence_score': 0.9,
            'evidence': 'alias cache',
            'notes': 'alias cache hit',
        }

    tokens = norm['full_tokens']
    aliases = generate_aliases(norm['clean_name'])
    expected_domains = _expected_domains(aliases)
    candidates = generate_domain_candidates(norm)

    scored: list[tuple[int, str, str, str]] = []

    with ThreadPoolExecutor(max_workers=8) as executor:
        # Submit search in parallel with candidate validation
        search_future = executor.submit(_run_search, firm_name, rate_state)
        cand_futures = {executor.submit(_validate_candidate, d, tokens, norm, rate_state): d for d in candidates}

        for fut in as_completed(cand_futures):
            domain = cand_futures[fut]
            try:
                s, chosen_url, note = fut.result()
                if s >= 0:
                    scored.append((s, chosen_url, domain, 'deterministic'))
            except Exception:
                pass

        # Collect search results
        try:
            search_domains = search_future.result(timeout=20)
        except Exception:
            search_domains = []

        new_search = [d for d in search_domains if d not in candidates]
        search_futures = {executor.submit(_validate_candidate, d, tokens, norm, rate_state): d for d in new_search[:8]}
        for fut in as_completed(search_futures):
            domain = search_futures[fut]
            try:
                s, chosen_url, note = fut.result()
                if s >= 0:
                    scored.append((s, chosen_url, domain, 'search'))
            except Exception:
                pass

    if scored:
        scored.sort(key=lambda x: -x[0])
        best_score, best_url, best_domain, method = scored[0]
        if best_score >= 4:
            cache[norm['clean_name']] = {
                'domain': best_domain,
                'chosen_url': best_url,
                'confidence': min(1.0, best_score / 10.0),
                'timestamp': datetime.utcnow().isoformat(),
                'evidence': method,
            }
            alias_cache[norm['clean_name']] = best_domain
            return {
                'firm': firm_name,
                'normalized_firm': norm['clean_name'],
                'official_website_url': best_url,
                'discovery_method': method,
                'confidence_score': min(1.0, best_score / 10.0),
                'evidence': method,
                'notes': f'score={best_score}',
            }

    # Fallback: alias guess
    for alias in aliases:
        token = alias.replace(' ', '')
        for guess in [f'{token}.com']:
            s, chosen_url, note = _validate_candidate(guess, tokens, norm, rate_state)
            if s >= 4:
                cache[norm['clean_name']] = {
                    'domain': guess,
                    'chosen_url': chosen_url,
                    'confidence': min(1.0, s / 10.0),
                    'timestamp': datetime.utcnow().isoformat(),
                    'evidence': 'alias_guess',
                }
                alias_cache[norm['clean_name']] = guess
                return {
                    'firm': firm_name,
                    'normalized_firm': norm['clean_name'],
                    'official_website_url': chosen_url,
                    'discovery_method': 'alias_guess',
                    'confidence_score': min(1.0, s / 10.0),
                    'evidence': 'alias_guess',
                    'notes': f'score={s}',
                }

    return {
        'firm': firm_name,
        'normalized_firm': norm['clean_name'],
        'official_website_url': '',
        'discovery_method': 'unresolved',
        'confidence_score': 0.0,
        'evidence': '',
        'notes': 'no verified domain',
        'candidates': candidates,
    }


def process_excel(input_path: str, refresh: bool, max_firms: int, start_index: int) -> None:
    if not os.path.exists(input_path):
        raise FileNotFoundError(input_path)

    wb = load_workbook(input_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if 'Firm' not in headers:
        raise ValueError("Input must have a 'Firm' column")

    firm_col = headers.index('Firm') + 1

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Firms'
    out_ws.append(['Firm', 'normalized_firm', 'official_website_url',
                   'discovery_method', 'confidence_score', 'evidence', 'notes'])

    failures = Workbook()
    fail_ws = failures.active
    fail_ws.title = 'Failures'
    fail_ws.append(['Firm', 'normalized_firm', 'candidates', 'failure_reason', 'timestamp'])

    cache = _load_json(CACHE_FILE)
    alias_cache = _load_json(ALIAS_FILE)
    rate_state: dict[str, float] = {}

    firms = []
    for row in ws.iter_rows(min_row=2):
        firm = row[firm_col - 1].value
        if firm:
            firms.append(str(firm).strip())

    if start_index > 0:
        firms = firms[start_index:]
    if max_firms > 0:
        firms = firms[:max_firms]

    success = 0
    for i, firm in enumerate(firms):
        _log(f'[{i+1}/{len(firms)}] {firm}')
        result = discover_official_domain(firm, refresh, rate_state, cache, alias_cache)
        out_ws.append([
            result['firm'],
            result['normalized_firm'],
            result['official_website_url'],
            result['discovery_method'],
            result['confidence_score'],
            result['evidence'],
            result['notes'],
        ])
        if result.get('official_website_url'):
            success += 1
        else:
            fail_ws.append([
                result['firm'],
                result['normalized_firm'],
                ', '.join(result.get('candidates', [])),
                result.get('notes', ''),
                datetime.utcnow().isoformat(),
            ])

    base = os.path.splitext(os.path.basename(input_path))[0]
    output_path = os.path.join(os.path.dirname(input_path) or '.', f'{base}_with_websites.xlsx')
    out_wb.save(output_path)
    failures.save(OUTPUT_FAILURES)
    _save_json(CACHE_FILE, cache)
    _save_json(ALIAS_FILE, alias_cache)

    total = len(firms)
    _log(f'Done: {success}/{total} resolved ({100*success//total if total else 0}%)')
    _log(f'Saved: {output_path}')
    _log(f'Failures: {OUTPUT_FAILURES}')


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Firm website discovery with caching')
    parser.add_argument('excel_path', help='Input Excel file with Firm column')
    parser.add_argument('--refresh', action='store_true', help='Ignore cache')
    parser.add_argument('--max-firms', type=int, default=0, help='Max firms to process')
    parser.add_argument('--start-index', type=int, default=0, help='Start index')
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    process_excel(args.excel_path, args.refresh, args.max_firms, args.start_index)


if __name__ == '__main__':
    main()
