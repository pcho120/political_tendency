#!/usr/bin/env python3
"""find_attorney.py - High-Coverage Attorney Scraper (Option B)

Strategy:
1. FAST PATH: Sitemap + robots.txt + requests (30s per firm)
2. FALLBACK: Stabilization-based Playwright discovery with:
   - Pagination, load-more, infinite scroll (stabilization)
   - Filter enumeration (practice/office)
   - API detection + full enumeration
   - Coverage auditing (expected vs actual)

Usage:
  python find_attorney.py "Company list_with_websites.xlsx"
  python find_attorney.py --debug-firm "Kirkland & Ellis" --headful true
"""

from __future__ import annotations

import argparse
import logging
import gzip
import json
import os
import re
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse, parse_qs, urlencode, urlunparse

import requests
from openpyxl import load_workbook

# Phase 1-3 Observation System
from observation_logger import ObservationLogger, FirmObservation
from pattern_aggregator import PatternAggregator
from discovery_rules import DiscoveryRulesEngine, DiscoveryRecommendation

# New Architecture Engines
from compliance_engine import (
    ComplianceEngine,
    CLASS_BLOCKED_BY_BOT,
    CLASS_AUTH_REQUIRED,
)
from rate_limit_manager import RateLimitManager, RateLimitBlockedError
from field_merger import FieldMerger, MergedAttorneyProfile
from coverage_engine import CoverageEngine, CoverageMetrics

# Coverage loop, field enrichment, external directories
from coverage_loop import (
    CoverageLoop, CoverageLoopResult, ExpectedTotalResolver, ExpectedTotalResult,
    filter_us_attorneys, PaginationEnumerator, AlphabetEnumerator,
    FirmSummaryWriter, FirmSummaryRow,
)
from field_enricher import FieldEnricher, EnrichmentLog
from external_directory_extractor import ExternalDirectoryExtractor, ExternalResult

# Configuration
DEFAULT_SHEET_NAME = "Attorneys"
MAX_WORKERS = 4

# TASK 2: Strict Timeout Policy
REQUEST_TIMEOUT = 5  # HTTP requests (was 10s)
SITEMAP_PROBE_TIMEOUT = 5  # Sitemap probing
PROFILE_FETCH_TIMEOUT = 10  # Individual profile fetch
PLAYWRIGHT_PAGE_TIMEOUT = 20000  # Playwright page load (20s in ms)
MAX_FIRM_TIME = 45  # TASK 5: Firm execution cap (seconds)

MIN_ATTORNEYS_THRESHOLD = 5
DEFAULT_STABILIZATION = 3
DEFAULT_MAX_SCROLL_SECONDS = 120
RATE_LIMIT_DELAY = 0.5  # seconds between requests per domain
COVERAGE_THRESHOLD = 0.98  # 98% coverage required for SUCCESS
MIN_PROFILE_LINKS_FOR_HTML_SITEMAP = 5  # Minimum profile links to validate HTML sitemap

# TASK 4: Batch enrichment
LARGE_FIRM_THRESHOLD = 1000  # URLs threshold for batch processing
BATCH_SIZE = 50  # Profiles per batch for large firms

# Known sitemap paths to probe (XML and HTML)
KNOWN_SITEMAP_PATHS = [
    "/sitemap.xml",
    "/sitemap_index.xml",
    "/sitemap",
    "/sitemap/lawyers",
    "/sitemap/people",
    "/sitemap/attorneys",
    "/sitemap/professionals",
    "/site-map",
    "/site-map/lawyers",
    "/site-map/people",
    "/site-map/attorneys",
]

# Firm Type Classifications
FIRM_TYPE_XML_SITEMAP = "XML_SITEMAP"
FIRM_TYPE_HTML_SITEMAP = "HTML_SITEMAP"
FIRM_TYPE_DIRECTORY_HTML = "DIRECTORY_HTML"
FIRM_TYPE_DIRECTORY_FILTERED = "DIRECTORY_FILTERED"
FIRM_TYPE_SPA_API = "SPA_API"
# REMOVED: FIRM_TYPE_HARD_CASE - replaced with observation-based system

# Coverage Status
STATUS_SUCCESS = "SUCCESS"
STATUS_PARTIAL = "PARTIAL"
STATUS_UNKNOWN_PATTERN = "UNKNOWN_PATTERN"  # Replaces HARD_CASE - no clear pattern observed
STATUS_LEGALLY_INCOMPLETE = "LEGALLY_INCOMPLETE"  # Compliance engine blocked access

# Discovery Status (for external directory fallback)
DISCOVERY_SUCCESS = "SUCCESS"
DISCOVERY_BLOCKED = "BLOCKED"  # 403 / Cloudflare / bot protection
DISCOVERY_INCOMPLETE = "DISCOVERY_INCOMPLETE"  # Replaces DISCOVERY_FAILED - insufficient observations

# Discovery Failure Reasons
FAILURE_SITEMAP_NOT_FOUND = "sitemap_not_found"
FAILURE_SITEMAP_EMPTY = "sitemap_found_but_empty"
FAILURE_ZERO_PROFILES = "directory_detected_but_zero_profiles"
FAILURE_EXTRACTION_ERROR = "extraction_error"
FAILURE_ALL_BLOCKED = "all_profiles_blocked"

# Multi-Source Types (from firm_finder_desktop.py discovery)
SOURCE_TYPE_ATTORNEY_LIST = "attorney_list"
SOURCE_TYPE_PROFILE_CORE = "profile_core"
SOURCE_TYPE_EDUCATION = "education"
SOURCE_TYPE_BAR_ADMISSION = "bar_admission"
SOURCE_TYPE_PRACTICE = "practice"
SOURCE_TYPE_MIXED = "mixed"

# Field precedence for multi-source aggregation (higher = more trusted)
FIELD_SOURCE_PRECEDENCE = {
    "profile_core": 100,     # Main profile page - highest trust
    "mixed": 90,             # Mixed source - very high trust
    "attorney_list": 80,     # Directory page - high trust
    "education": 70,         # Education-specific page
    "bar_admission": 70,     # Bar admission-specific page
    "practice": 60,          # Practice area page
    "external_directory": 30, # External directories - lower trust
}


# ---------------------------------------------------------------------------
# URL locale / language filter
# ---------------------------------------------------------------------------
_LOCALE_URL_REJECT_RE = re.compile(
    r'/(?:de|ar|es|fr|zh|ja|ko|pt|ru|it|nl|pl|sv|tr|da|fi|nb|cs|hu|ro|sk|uk)(?:/|$)',
    re.IGNORECASE,
)


def _is_locale_url(url: str) -> bool:
    """Return True if the URL belongs to a non-English locale (should be rejected)."""
    try:
        return bool(_LOCALE_URL_REJECT_RE.search(url))
    except Exception:
        return False


# ---------------------------------------------------------------------------
# US office whitelist (cities, states, DC)
# ---------------------------------------------------------------------------
US_OFFICE_WHITELIST: frozenset = frozenset({
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
    "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
    "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
    "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
    "New Hampshire", "New Jersey", "New Mexico", "New York",
    "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
    "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
    "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
    "West Virginia", "Wisconsin", "Wyoming",
    "District of Columbia", "D.C.", "DC", "Washington D.C.", "Washington DC",
    "New York", "Los Angeles", "Chicago", "Houston", "Phoenix", "Philadelphia",
    "San Antonio", "San Diego", "Dallas", "San Jose", "Austin", "Jacksonville",
    "Fort Worth", "Columbus", "Charlotte", "Indianapolis", "San Francisco",
    "Seattle", "Denver", "Nashville", "Oklahoma City", "El Paso", "Boston",
    "Portland", "Las Vegas", "Louisville", "Baltimore", "Milwaukee",
    "Albuquerque", "Tucson", "Fresno", "Sacramento", "Mesa", "Kansas City",
    "Atlanta", "Omaha", "Colorado Springs", "Raleigh", "Long Beach",
    "Virginia Beach", "Minneapolis", "Tampa", "New Orleans", "Arlington",
    "Wichita", "Bakersfield", "Aurora", "Anaheim", "Santa Ana", "Corpus Christi",
    "Riverside", "St. Louis", "Lexington", "Pittsburgh", "Stockton",
    "Anchorage", "Cincinnati", "St. Paul", "Greensboro", "Toledo", "Newark",
    "Plano", "Henderson", "Orlando", "Lincoln", "Jersey City", "Chandler",
    "St. Petersburg", "Laredo", "Norfolk", "Madison", "Durham", "Lubbock",
    "Winston-Salem", "Garland", "Glendale", "Hialeah", "Reno", "Baton Rouge",
    "Irvine", "Chesapeake", "Scottsdale", "North Las Vegas", "Fremont",
    "Gilbert", "San Bernardino", "Birmingham", "Rochester", "Richmond",
    "Spokane", "Des Moines", "Montgomery", "Modesto", "Fayetteville",
    "Tacoma", "Shreveport", "Fontana", "Moreno Valley",
    "Akron", "Yonkers", "Huntington Beach", "Little Rock",
    "Providence", "Grand Rapids", "Salt Lake City", "Tallahassee", "Huntsville",
    "Worcester", "Brownsville", "Knoxville", "Santa Clarita", "Augusta",
    "Peoria", "Garden Grove", "Oceanside", "Chattanooga", "Fort Lauderdale",
    "Rancho Cucamonga", "Santa Rosa", "Cape Coral", "Tempe",
    "Sioux Falls", "Jackson", "Overland Park", "Eugene",
    "Elk Grove", "Pembroke Pines", "Salem", "Corona",
    "Fort Collins", "Alexandria", "Hayward", "Lancaster", "Palmdale",
    "Salinas", "Sunnyvale", "Pomona", "Escondido",
    "Paterson", "Torrance", "Pasadena", "Orange", "Fullerton", "Clarksville",
    "Savannah", "Dayton", "Hampton", "Surprise", "Roseville", "Macon",
    "Thornton", "Columbia", "Hartford", "Bridgeport", "Hollywood", "Naperville",
    "McKinney", "Murfreesboro", "Syracuse", "Denton", "Bellevue",
    "Frisco", "Mesquite", "Rockford", "Gainesville", "Killeen", "Waco",
    "Springfield", "Warren", "Sterling Heights", "Cedar Rapids", "Thousand Oaks",
    "Visalia", "Olathe", "New Haven", "Simi Valley", "Concord",
    "Stamford", "Elizabeth", "Topeka", "Miramar", "Coral Springs",
    "Carrollton", "Midland", "West Valley City",
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA",
    "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD",
    "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
    "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC",
    "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY",
})


def is_us_attorney(profile) -> bool:
    """Return True if the attorney's office/bar_admissions indicate a US location."""
    from attorney_extractor import US_STATES as _US_STATES

    offices = getattr(profile, 'offices', []) or []
    for office in offices:
        office_lower = office.lower() if office else ''
        for term in US_OFFICE_WHITELIST:
            if term.lower() in office_lower:
                return True

    bars = getattr(profile, 'bar_admissions', []) or []
    for bar in bars:
        bar_lower = bar.lower() if bar else ''
        for state in _US_STATES:
            if state.lower() in bar_lower:
                return True

    return False

def _ext_dir_profile_valid(att) -> bool:
    """Hard gate for external directory profiles: require name + bar + US office + title."""
    from attorney_extractor import _VALID_NAME_RE as _VNR, _HEADER_TERMS as _HT
    name = getattr(att, 'full_name', None)
    if not name or not _VNR.match(name.strip()) or name.strip().lower() in _HT:
        return False
    if not getattr(att, 'bar_admissions', None):
        return False
    if not getattr(att, 'title', None):
        return False
    if not is_us_attorney(att):
        return False
    return True

@dataclass
class SourceFailure:
    """Tracks a failed source during multi-source extraction"""
    firm: str
    source_url: str
    source_type: str
    failure_type: str  # http_error, parse_error, blocked, timeout, empty_response
    http_status: int | None = None
    fields_expected: list[str] = field(default_factory=list)
    fields_extracted: list[str] = field(default_factory=list)
    error_message: str = ""
    timestamp: str = ""

    def to_row(self) -> list:
        """Convert to Excel row for source_failure_report.xlsx"""
        import datetime
        return [
            self.firm,
            self.source_url,
            self.source_type,
            self.failure_type,
            self.http_status or "",
            ", ".join(self.fields_expected) if self.fields_expected else "",
            ", ".join(self.fields_extracted) if self.fields_extracted else "",
            self.error_message,
            self.timestamp or datetime.datetime.now().isoformat()
        ]


def merge_attorney_fields(
    base_profile,
    supplemental_data: dict,
    source_type: str,
    precedence_map: dict[str, int] | None = None
) -> None:
    """Merge fields from supplemental source into base profile using precedence rules.

    Args:
        base_profile: AttorneyProfile to merge into
        supplemental_data: Dict with field values from supplemental source
        source_type: Type of source (profile_core, education, etc.) for precedence
        precedence_map: Custom precedence map (defaults to FIELD_SOURCE_PRECEDENCE)

    Mutates base_profile in place.
    Only overwrites fields if:
    1. Field is empty in base_profile, OR
    2. Source precedence is higher than current source
    """
    if precedence_map is None:
        precedence_map = FIELD_SOURCE_PRECEDENCE

    source_precedence = precedence_map.get(source_type, 50)

    # Track current source precedence per field
    if not hasattr(base_profile, '_field_sources'):
        base_profile._field_sources = {}

    field_mapping = {
        'full_name': 'full_name',
        'name': 'full_name',
        'title': 'title',
        'offices': 'offices',
        'department': 'department',
        'practice_areas': 'practice_areas',
        'industries': 'industries',
        'bar_admissions': 'bar_admissions',
        'education': 'education',
    }

    for src_field, dst_field in field_mapping.items():
        if src_field not in supplemental_data:
            continue

        new_value = supplemental_data[src_field]
        if not new_value:  # Skip empty values
            continue

        current_value = getattr(base_profile, dst_field, None)
        current_precedence = base_profile._field_sources.get(dst_field, 0)

        # Decide whether to overwrite
        should_overwrite = False

        if not current_value:
            # Field is empty - always take new value
            should_overwrite = True
        elif isinstance(current_value, list) and len(current_value) == 0:
            # Empty list - always take new value
            should_overwrite = True
        elif source_precedence > current_precedence:
            # Higher precedence source - overwrite
            should_overwrite = True
        elif source_precedence == current_precedence:
            # Same precedence - merge lists, keep scalar
            if isinstance(current_value, list) and isinstance(new_value, list):
                # Merge lists (unique values)
                existing_set = set(str(v) for v in current_value)
                for v in new_value:
                    if str(v) not in existing_set:
                        current_value.append(v)
                continue  # Already merged, skip overwrite

        if should_overwrite:
            setattr(base_profile, dst_field, new_value)
            base_profile._field_sources[dst_field] = source_precedence

            # Update diagnostics
            if hasattr(base_profile, 'diagnostics'):
                base_profile.diagnostics[f'{dst_field}_source'] = source_type

            # Remove from missing_fields if applicable
            if hasattr(base_profile, 'missing_fields') and dst_field in base_profile.missing_fields:
                base_profile.missing_fields.remove(dst_field)


@dataclass
class FirmProfile:
    """Detected firm type and characteristics"""
    firm: str
    base_url: str
    detected_types: list[str] = field(default_factory=list)
    signals: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "firm": self.firm,
            "base_url": self.base_url,
            "detected_types": self.detected_types,
            "signals": self.signals,
        }


@dataclass
class FirmSourceMap:
    """Multi-source configuration for a firm from discovery Excel"""
    firm: str
    official_website: str
    sources: list[dict] = field(default_factory=list)  # List of source configs
    # Each source: {source_url, source_type, fields_provided, validation_status, discovery_method}

    def get_attorney_list_sources(self) -> list[dict]:
        """Get sources that provide attorney lists (sitemaps, directories)

        ACCEPTS: VALID and UNVERIFIED sources (discovery phase is liberal)
        """
        return [s for s in self.sources
                if s.get('source_type') == SOURCE_TYPE_ATTORNEY_LIST
                and s.get('validation_status') in ['VALID', 'UNVERIFIED']]

    def get_profile_sources(self) -> list[dict]:
        """Get sources that provide profile data (for field enrichment)

        ACCEPTS: VALID and UNVERIFIED sources (discovery phase is liberal)
        """
        return [s for s in self.sources
                if s.get('source_type') in [SOURCE_TYPE_PROFILE_CORE, SOURCE_TYPE_MIXED, SOURCE_TYPE_PRACTICE]
                and s.get('validation_status') in ['VALID', 'UNVERIFIED']]

    def get_secondary_sources(self) -> list[dict]:
        """Get secondary sources (education, bar admissions tabs)

        ACCEPTS: VALID and UNVERIFIED sources (discovery phase is liberal)
        """
        return [s for s in self.sources
                if s.get('source_type') in [SOURCE_TYPE_EDUCATION, SOURCE_TYPE_BAR_ADMISSION]
                and s.get('validation_status') in ['VALID', 'UNVERIFIED']]

    def get_fields_by_source(self) -> dict[str, list[str]]:
        """Map source_type -> available fields"""
        result = {}
        for s in self.sources:
            src_type = s.get('source_type', 'unknown')
            fields = s.get('fields_provided', '')
            if isinstance(fields, str):
                fields = [f.strip() for f in fields.split(',') if f.strip()]
            if fields:
                result[src_type] = fields
        return result

    def to_dict(self) -> dict:
        return {
            "firm": self.firm,
            "official_website": self.official_website,
            "sources": self.sources,
            "attorney_list_count": len(self.get_attorney_list_sources()),
            "profile_source_count": len(self.get_profile_sources()),
            "secondary_source_count": len(self.get_secondary_sources()),
        }


@dataclass
class CoverageReport:
    """Coverage validation results"""
    firm: str
    expected_total: int | None = None
    discovered_urls: int = 0
    strategies_attempted: list[str] = field(default_factory=list)
    final_status: str = "UNKNOWN"  # SUCCESS | PARTIAL | HARD_CASE
    discovery_status: str = "UNKNOWN"  # SUCCESS | BLOCKED | DISCOVERY_INCOMPLETE
    failure_reason: str | None = None  # Specific failure reason if applicable
    notes: list[str] = field(default_factory=list)
    url_sources: dict[str, int] = field(default_factory=dict)  # strategy -> count

    def to_dict(self) -> dict:
        coverage_ratio = None
        if self.expected_total and self.expected_total > 0:
            coverage_ratio = self.discovered_urls / self.expected_total

        return {
            "firm": self.firm,
            "expected_total": self.expected_total,
            "discovered_urls": self.discovered_urls,
            "coverage_ratio": coverage_ratio,
            "strategies_attempted": self.strategies_attempted,
            "final_status": self.final_status,
            "discovery_status": self.discovery_status,
            "failure_reason": self.failure_reason,
            "notes": self.notes,
            "url_sources": self.url_sources,
        }


@dataclass
class DiscoveryMetrics:
    """Track discovery coverage metrics (legacy compatibility)"""
    directory_url: str = ""
    expected_total: int = 0
    discovered_unique: int = 0
    discovered_by_dom: int = 0
    discovered_by_pagination: int = 0
    discovered_by_loadmore: int = 0
    discovered_by_scroll: int = 0
    discovered_by_filters: int = 0
    discovered_by_api: int = 0
    failure_notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "directory_url": self.directory_url,
            "expected_total": self.expected_total,
            "discovered_unique": self.discovered_unique,
            "discovered_by_dom": self.discovered_by_dom,
            "discovered_by_pagination": self.discovered_by_pagination,
            "discovered_by_loadmore": self.discovered_by_loadmore,
            "discovered_by_scroll": self.discovered_by_scroll,
            "discovered_by_filters": self.discovered_by_filters,
            "discovered_by_api": self.discovered_by_api,
            "failure_notes": self.failure_notes,
            "coverage_ratio": (
                self.discovered_unique / self.expected_total
                if self.expected_total > 0
                else 0.0
            ),
        }


class AttorneyFinder:
    def __init__(
        self,
        *,
        limit: int,
        sheet_name: str,
        max_firms: int,
        workers: int,
        debug_firm: str = "",
        debug_domain: str = "",
        headful: bool = False,
        stabilization: int = DEFAULT_STABILIZATION,
        max_scroll_seconds: int = DEFAULT_MAX_SCROLL_SECONDS,
        sources_file: str = "",  # Multi-source discovery Excel from firm_finder_desktop.py
        output_dir: str = ".",  # Output directory for attorneys.xlsx, attorneys.jsonl, coverage_metrics.json
    ) -> None:
        self.limit = limit
        self.sheet_name = sheet_name
        self.max_firms = max_firms
        self.workers = workers
        self.debug_firm = debug_firm
        self.debug_domain = debug_domain
        self.headful = headful
        self.stabilization = stabilization
        self.max_scroll_seconds = max_scroll_seconds
        self.sources_file = sources_file

        # Multi-source map: firm_name -> FirmSourceMap
        self.firm_source_maps: dict[str, FirmSourceMap] = {}

        # Track source failures for reporting
        self.source_failures: list[SourceFailure] = []

        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }
        )

        # Create debug directory
        self.debug_dir = Path("debug_reports")
        self.debug_dir.mkdir(exist_ok=True)

        # Track last request time per domain for rate limiting
        self.domain_last_request: dict[str, float] = {}

        # Firm-level enrichment mode
        self.enrichment_mode: str = "REQUESTS"  # REQUESTS | PLAYWRIGHT | PLAYWRIGHT_ONLY

        # Phase 1-3 Observation System
        self.observation_logger = ObservationLogger("firm_observations.jsonl")
        self.discovery_engine = DiscoveryRulesEngine("firm_observations.jsonl", "rule_confidence.json")

        # New Architecture Engines
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.compliance_engine = ComplianceEngine(report_dir=self.debug_dir)
        self.rate_limit_manager = RateLimitManager(default_delay=RATE_LIMIT_DELAY)
        self.field_merger = FieldMerger()
        self.coverage_engine = CoverageEngine()
        self.all_coverage_metrics: list[CoverageMetrics] = []
        self.jsonl_path = self.output_dir / "attorneys.jsonl"
        self.jsonl_file = open(self.jsonl_path, "a", encoding="utf-8")

        # Coverage loop, field enrichment, firm summary
        self.expected_total_resolver = ExpectedTotalResolver()
        self.firm_summary_writer = FirmSummaryWriter()
        self.field_enricher = FieldEnricher()

    def log(self, msg: str) -> None:
        try:
            print(msg, flush=True)
        except UnicodeEncodeError:
            import sys
            sys.stdout.buffer.write((msg + "\n").encode("utf-8", errors="replace"))
            sys.stdout.buffer.flush()

    def _rate_limit(self, domain: str) -> None:
        """Enforce per-domain rate limiting"""
        last = self.domain_last_request.get(domain, 0)
        elapsed = time.time() - last
        if elapsed < RATE_LIMIT_DELAY:
            time.sleep(RATE_LIMIT_DELAY - elapsed)
        self.domain_last_request[domain] = time.time()

    # ========================================================================
    # MULTI-SOURCE LOADING
    # ========================================================================

    def load_source_maps(self) -> None:
        """Load multi-source discovery Excel from firm_finder_desktop.py

        Expected columns:
        - Firm, Official Website, Source URL, Source Type
        - Fields Provided, Validation Status, Discovery Method, Notes

        Builds self.firm_source_maps: {firm_name -> FirmSourceMap}
        """
        if not self.sources_file or not os.path.exists(self.sources_file):
            self.log("No sources file specified or file not found - using single-source mode")
            return

        self.log(f"\n=== Loading Multi-Source Discovery: {self.sources_file} ===")

        try:
            wb = load_workbook(self.sources_file, read_only=True)
            ws = wb.active

            # Find column indices
            headers = [cell.value for cell in ws[1]]
            col_map = {}
            expected_cols = ["Firm", "Official Website", "Source URL", "Source Type",
                           "Fields Provided", "Validation Status", "Discovery Method", "Notes"]

            for col_name in expected_cols:
                try:
                    col_map[col_name] = headers.index(col_name)
                except ValueError:
                    self.log(f"  Warning: Column '{col_name}' not found in sources file")

            if "Firm" not in col_map:
                self.log("  ERROR: 'Firm' column required in sources file")
                wb.close()
                return

            # Read all rows and group by firm
            firm_sources: dict[str, list[dict]] = {}
            firm_websites: dict[str, str] = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                firm_name = row[col_map["Firm"]] if col_map.get("Firm") is not None else None
                if not firm_name:
                    continue

                firm_name = str(firm_name).strip()

                # Get official website
                website = ""
                if col_map.get("Official Website") is not None:
                    website = str(row[col_map["Official Website"]] or "").strip()

                if firm_name not in firm_websites and website:
                    firm_websites[firm_name] = website

                # Build source config
                source = {
                    "source_url": str(row[col_map.get("Source URL", 2)] or "").strip() if col_map.get("Source URL") is not None else "",
                    "source_type": str(row[col_map.get("Source Type", 3)] or "").strip() if col_map.get("Source Type") is not None else "",
                    "fields_provided": str(row[col_map.get("Fields Provided", 4)] or "").strip() if col_map.get("Fields Provided") is not None else "",
                    "validation_status": str(row[col_map.get("Validation Status", 5)] or "").strip() if col_map.get("Validation Status") is not None else "",
                    "discovery_method": str(row[col_map.get("Discovery Method", 6)] or "").strip() if col_map.get("Discovery Method") is not None else "",
                    "notes": str(row[col_map.get("Notes", 7)] or "").strip() if col_map.get("Notes") is not None else "",
                }

                if firm_name not in firm_sources:
                    firm_sources[firm_name] = []
                firm_sources[firm_name].append(source)

            wb.close()

            # Build FirmSourceMap objects
            for firm_name, sources in firm_sources.items():
                self.firm_source_maps[firm_name] = FirmSourceMap(
                    firm=firm_name,
                    official_website=firm_websites.get(firm_name, ""),
                    sources=sources
                )

            # Log summary
            total_sources = sum(len(s.sources) for s in self.firm_source_maps.values())
            valid_sources = sum(
                len([src for src in s.sources if src.get('validation_status') == 'VALID'])
                for s in self.firm_source_maps.values()
            )

            self.log(f"  Loaded {len(self.firm_source_maps)} firms with {total_sources} total sources ({valid_sources} valid)")

            # Log sample
            for firm_name, source_map in list(self.firm_source_maps.items())[:3]:
                self.log(f"    {firm_name}: {len(source_map.sources)} sources")
                for src in source_map.sources[:2]:
                    self.log(f"      - {src['source_type']}: {src['fields_provided'][:50]}...")

        except Exception as e:
            self.log(f"  ERROR loading sources file: {e}")
            import traceback
            traceback.print_exc()

    def get_source_map(self, firm_name: str) -> FirmSourceMap | None:
        """Get source map for a firm (case-insensitive match)"""
        # Exact match first
        if firm_name in self.firm_source_maps:
            return self.firm_source_maps[firm_name]

        # Case-insensitive match
        firm_lower = firm_name.lower().strip()
        for key, value in self.firm_source_maps.items():
            if key.lower().strip() == firm_lower:
                return value

        return None

    def detect_firm_type(self, firm_name: str, base_url: str) -> FirmProfile:
        """REFACTORED: Observation-based firm type detection using Phase 1-3 system

        Flow:
        1. Get discovery recommendation from Phase 3 rules engine
        2. Phase 3 uses existing observation or creates new one (Phase 1)
        3. Convert Phase 3 strategy recommendations to FirmProfile format
        4. Return profile with detected types based on recommended strategies
        """
        self.log(f"\n=== Phase 1-3 Discovery: {firm_name} ===")

        # Get observation-based recommendation from Phase 3
        recommendation = self.discovery_engine.get_discovery_recommendation(firm_name, base_url)

        # Convert recommendation to FirmProfile format (for backward compatibility)
        profile = FirmProfile(firm=firm_name, base_url=base_url)

        # Map Phase 3 strategies to legacy firm type constants
        strategy_to_type_map = {
            "xml_sitemap_as_list": FIRM_TYPE_XML_SITEMAP,
            "xml_sitemap_navigation": FIRM_TYPE_XML_SITEMAP,
            "alphabet_enumeration": FIRM_TYPE_DIRECTORY_HTML,  # Alphabet nav uses directory strategy
            "directory_listing": FIRM_TYPE_DIRECTORY_HTML,
            "json_api": FIRM_TYPE_SPA_API,
        }

        # Extract detected types from recommended strategies
        detected_types_set = set()
        for strategy in recommendation.strategies:
            legacy_type = strategy_to_type_map.get(strategy.strategy_type)
            if legacy_type:
                detected_types_set.add(legacy_type)

        profile.detected_types = list(detected_types_set) if detected_types_set else ["UNKNOWN_PATTERN"]

        # Convert recommendation notes to signals
        profile.signals = recommendation.notes

        # Add classification as signal
        profile.signals.insert(0, f"Classification: {recommendation.classification}")

        # Log results
        self.log(f"  Classification: {recommendation.classification}")
        self.log(f"  Detected types: {', '.join(profile.detected_types)}")
        self.log(f"  Strategies ({len(recommendation.strategies)}):")
        for i, strat in enumerate(recommendation.strategies[:5], 1):
            self.log(f"    {i}. {strat.strategy_type} (confidence: {strat.confidence_score:.2f}, priority: {strat.priority})")

        return profile

    def _check_for_sitemap(self, base_url: str) -> list[str]:
        """Check robots.txt for sitemap directives"""
        domain = urlparse(base_url).netloc
        scheme = urlparse(base_url).scheme

        sitemap_urls = []
        try:
            robots_url = f"{scheme}://{domain}/robots.txt"
            try:
                self.rate_limit_manager.wait(domain)
            except RateLimitBlockedError:
                pass
            resp = self.session.get(robots_url, timeout=REQUEST_TIMEOUT)
            if resp.status_code == 200:
                for line in resp.text.split('\n'):
                    if line.lower().startswith('sitemap:'):
                        sitemap_url = line.split(':', 1)[1].strip()
                        sitemap_urls.append(sitemap_url)
                        self.log(f"  Found sitemap in robots.txt: {sitemap_url}")
        except Exception as e:
            self.log(f"  robots.txt check failed: {e}")

        # Add default locations if nothing found
        if not sitemap_urls:
            sitemap_urls = [
                f"{scheme}://{domain}/sitemap.xml",
                f"{scheme}://{domain}/sitemap_index.xml",
            ]

        return sitemap_urls

    def _classify_sitemap(self, sitemap_url: str, domain: str) -> str | None:
        """Classify sitemap as XML or HTML"""
        try:
            try:
                self.rate_limit_manager.wait(domain)
            except RateLimitBlockedError:
                return None
            resp = self.session.get(sitemap_url, timeout=REQUEST_TIMEOUT)
            if resp.status_code != 200:
                return None

            content = resp.content
            if sitemap_url.endswith(".gz"):
                content = gzip.decompress(content)

            # Try XML parsing
            try:
                ET.fromstring(content)
                return "XML"
            except ET.ParseError:
                # Check if it's HTML
                if b"<html" in content.lower() or b"<!doctype html" in content.lower():
                    return "HTML"
                return None
        except Exception:
            return None

    def _html_sitemap_has_profile_links(self, sitemap_url: str, base_url: str, domain: str) -> bool:
        """Validate HTML sitemap by checking if it contains profile links

        Returns: True if sitemap has >= MIN_PROFILE_LINKS_FOR_HTML_SITEMAP profile links
        """
        try:
            try:
                self.rate_limit_manager.wait(domain)
            except RateLimitBlockedError:
                return False
            resp = self.session.get(sitemap_url, timeout=REQUEST_TIMEOUT)

            if resp.status_code != 200:
                return False

            content = resp.text

            # Extract links (limit to first 300 for performance)
            import re
            links = re.findall(r'href=["\'](.*?)["\']', content)[:300]

            # Count profile-like URLs
            profile_link_count = sum(1 for link in links if self._is_profile_like_url(urljoin(base_url, link), domain))

            if profile_link_count >= MIN_PROFILE_LINKS_FOR_HTML_SITEMAP:
                self.log(f"  HTML sitemap validated at {sitemap_url} ({profile_link_count} profile links)")
                return True
            else:
                return False

        except Exception:
            return False

    def _probe_sitemap_paths(self, base_url: str) -> dict[str, tuple[str, int]]:
        """Probe known sitemap paths and classify them with validation

        Returns: {path: (type, profile_count)} where type is "XML" or "HTML"
        """
        domain = urlparse(base_url).netloc
        scheme = urlparse(base_url).scheme
        findings = {}

        for path in KNOWN_SITEMAP_PATHS:
            full_url = f"{scheme}://{domain}{path}"

            try:
                try:
                    self.rate_limit_manager.wait(domain)
                except RateLimitBlockedError:
                    continue
                resp = self.session.get(full_url, timeout=REQUEST_TIMEOUT)

                if resp.status_code != 200:
                    continue

                content = resp.content

                # Try XML parsing first
                try:
                    ET.fromstring(content)
                    findings[path] = ("XML", 0)
                    self.log(f"    Found XML sitemap at {path}")
                    continue
                except ET.ParseError:
                    pass

                # Check if it's HTML with profile links (VALIDATED)
                if b"<html" in content.lower() or b"<!doctype html" in content.lower():
                    # Extract links to verify it's a real sitemap (limit to 300)
                    import re
                    links = re.findall(r'href=["\'](.*?)["\']', content.decode('utf-8', errors='ignore'))[:300]

                    # Count profile-like URLs
                    profile_link_count = sum(1 for link in links if self._is_profile_like_url(urljoin(base_url, link), domain))

                    if profile_link_count >= MIN_PROFILE_LINKS_FOR_HTML_SITEMAP:
                        findings[path] = ("HTML", profile_link_count)
                        self.log(f"    Found HTML sitemap at {path} ({profile_link_count} profile links)")
                        continue
                    else:
                        self.log(f"    Skipping {path} (only {profile_link_count} profile links, likely error page)")

            except Exception as e:
                # Silent fail for each path - expected for most paths
                pass

        return findings

    def _probe_directory_type(self, base_url: str) -> dict:
        """Probe directory pages to detect type (requires Playwright)"""
        result = {"types": [], "signals": []}

        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            result["signals"].append("Playwright not available for directory probing")
            return result

        captured_api_data = []

        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=not self.headful)
                context = browser.new_context()

                # Block heavy resources
                def handle_route(route):
                    if route.request.resource_type in ["image", "font", "media", "stylesheet"]:
                        route.abort()
                    else:
                        route.continue_()

                context.route("**/*", handle_route)
                page = context.new_page()

                # Intercept API responses
                def handle_response(response):
                    try:
                        url_lower = response.url.lower()
                        if any(kw in url_lower for kw in ["api", "graphql", "search", "people", "attorneys", "lawyers", "professionals"]):
                            if response.ok and "json" in response.headers.get("content-type", "").lower():
                                try:
                                    data = response.json()
                                    records = self._extract_attorneys_from_json(data, base_url)
                                    if records:
                                        captured_api_data.append({"url": response.url, "records": records})
                                except Exception:
                                    pass
                    except Exception:
                        pass

                page.on("response", handle_response)

                # Probe common directory paths
                probe_paths = ["/people", "/professionals", "/attorneys", "/lawyers"]
                best_profile_count = 0
                has_filters = False

                for path in probe_paths[:2]:  # Quick probe, only first 2
                    try:
                        target = urljoin(base_url, path)
                        response = page.goto(target, timeout=15000, wait_until="domcontentloaded")
                        if not response or not response.ok:
                            continue

                        page.wait_for_timeout(2000)

                        # Check for profile links
                        links = page.evaluate("""
                            () => {
                                const anchors = Array.from(document.querySelectorAll('a'));
                                return anchors.map(a => a.href).filter(Boolean);
                            }
                        """)

                        profile_count = sum(1 for href in links if self._is_profile_like_url(href, urlparse(base_url).netloc))
                        if profile_count > best_profile_count:
                            best_profile_count = profile_count

                        # Check for filters
                        filter_count = page.evaluate("""
                            () => {
                                const selects = document.querySelectorAll('select');
                                const filterDivs = document.querySelectorAll('[class*="filter"], [class*="facet"]');
                                return selects.length + filterDivs.length;
                            }
                        """)
                        if filter_count > 0:
                            has_filters = True

                    except Exception:
                        pass

                page.close()
                context.close()
                browser.close()

                # Classify based on findings
                if captured_api_data:
                    result["types"].append(FIRM_TYPE_SPA_API)
                    result["signals"].append(f"API detected: {len(captured_api_data)} endpoint(s)")

                if best_profile_count > 0:
                    if has_filters:
                        result["types"].append(FIRM_TYPE_DIRECTORY_FILTERED)
                        result["signals"].append(f"Directory with filters: {best_profile_count} profile links")
                    else:
                        result["types"].append(FIRM_TYPE_DIRECTORY_HTML)
                        result["signals"].append(f"Standard HTML directory: {best_profile_count} profile links")
                else:
                    result["signals"].append("No profile links found in directory pages")

        except Exception as e:
            result["signals"].append(f"Directory probing error: {e}")

        return result

    # ========================================================================
    # STRATEGY ROUTER
    # ========================================================================

    def select_strategies(self, profile: FirmProfile, firm_name: str = "") -> list[str]:
        """Select discovery strategies based on firm profile

        PRIORITY ORDER (highest to lowest):
        1. XML_SITEMAP - most reliable, structured data
        2. HTML_SITEMAP - validated profile links
        3. DIRECTORY strategies - fallback enumeration
        4. HARD_CASE - last resort

        XML sitemaps ALWAYS override all other methods
        """
        strategies = []

        # Fix 3: Kirkland override — xml_sitemap returns 0 URLs; force directory strategies
        if "kirkland" in firm_name.lower():
            self.log("  [KIRKLAND_OVERRIDE] Disabling xml_sitemap; using Playwright DOM scroll strategy")
            strategies.extend(["kirkland_scroll", "directory_listing", "alphabet_enumeration", "dom_exhaustion"])
            self.log(f"  Selected strategies: {', '.join(strategies)}")
            return strategies

        # PRIORITY 1: XML sitemap (ALWAYS use if available)
        if FIRM_TYPE_XML_SITEMAP in profile.detected_types:
            strategies.append("xml_sitemap")
            # XML found - skip all other strategies (most reliable source)
            self.log(f"  Selected strategies: {', '.join(strategies)}")
            return strategies

        # PRIORITY 2: HTML sitemap (only if no XML)
        if FIRM_TYPE_HTML_SITEMAP in profile.detected_types:
            strategies.append("html_sitemap")

        # PRIORITY 3: SPA/API (if detected)
        if FIRM_TYPE_SPA_API in profile.detected_types:
            strategies.append("api_enumeration")

        # PRIORITY 4: Directory strategies (only if no sitemaps)
        if FIRM_TYPE_DIRECTORY_FILTERED in profile.detected_types:
            strategies.append("filter_enumeration")

        if FIRM_TYPE_DIRECTORY_HTML in profile.detected_types:
            strategies.append("dom_exhaustion")  # pagination + load-more + scroll

        # FALLBACK: Unknown pattern (only if nothing else worked)
        if "UNKNOWN_PATTERN" in profile.detected_types or not strategies:
            strategies.append("hard_case_fallback")  # Keep same strategy name for now

        self.log(f"  Selected strategies: {', '.join(strategies)}")
        return strategies


    def run(self, excel_path: str) -> int:
        if not os.path.exists(excel_path):
            raise FileNotFoundError(excel_path)
        if not excel_path.lower().endswith(".xlsx"):
            raise ValueError("Input must be an .xlsx file")

        # Load multi-source discovery data if available
        self.load_source_maps()

        wb = load_workbook(excel_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if "Official Website" not in headers or "Firm" not in headers:
            for sheet in wb.worksheets:
                sheet_headers = [cell.value for cell in sheet[1]]
                if "Official Website" in sheet_headers and "Firm" in sheet_headers:
                    ws = sheet
                    headers = sheet_headers
                    break

        try:
            url_col_idx = headers.index("official_website_url") + 1
            firm_col_idx = headers.index("Firm") + 1
        except ValueError as e:
            raise ValueError(
                "File must have 'Firm' and 'Official Website' columns."
            ) from e

        if self.sheet_name in wb.sheetnames:
            del wb[self.sheet_name]
        out_ws = wb.create_sheet(self.sheet_name)
        # Updated header with new fields
        out_ws.append(
            [
                "Firm",
                "Attorney Name",
                "Title",
                "Offices",
                "Departments",
                "Practice Areas",
                "Industries",
                "Bar Admissions",
                "Education",
                "Extraction Status",
                "Missing Fields",
                "Profile URL",
                "Data Source"
            ]
        )

        firms = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            firm_name = row[firm_col_idx - 1].value
            base_url = row[url_col_idx - 1].value
            if not firm_name or not base_url:
                continue

            # Debug mode filters
            if self.debug_firm and self.debug_firm.lower() not in firm_name.lower():
                continue
            if self.debug_domain:
                domain = urlparse(str(base_url)).netloc
                if self.debug_domain.lower() not in domain.lower():
                    continue

            firms.append((firm_name, str(base_url)))
            if self.max_firms > 0 and len(firms) >= self.max_firms:
                break

        if not firms:
            self.log("No firms to process (check debug filters)")
            return 0

        total_processed = 0

        # Process sequentially to manage Playwright browser lifecycle
        for firm, url in firms:
            try:
                attorneys, metrics = self.process_firm(firm, url)
                for att in attorneys:
                    # Handle new AttorneyProfile dataclass
                    if hasattr(att, 'full_name'):  # New AttorneyProfile format
                        # Format list fields as comma-separated strings
                        offices_str = ", ".join(att.offices) if att.offices else ""
                        departments_str = att.department if att.department else ""
                        practices_str = ", ".join(att.practice_areas) if att.practice_areas else ""
                        industries_str = ", ".join(att.industries) if att.industries else ""
                        bars_str = ", ".join(att.bar_admissions) if att.bar_admissions else ""

                        # Education as JSON-stringified list of dicts (Step 2f)
                        education_str = json.dumps(
                            [e.to_dict() for e in att.education], ensure_ascii=False
                        ) if att.education else "[]"

                        missing_fields_str = ", ".join(att.missing_fields) if att.missing_fields else ""

                        # Extract data_source from diagnostics
                        data_source = att.diagnostics.get('data_source', 'firm_website')

                        out_ws.append([
                            firm,
                            att.full_name or "",
                            att.title or "",
                            offices_str,
                            departments_str,
                            practices_str,
                            industries_str,
                            bars_str,
                            education_str,
                            att.extraction_status,
                            missing_fields_str,
                            att.profile_url,
                            data_source
                        ])
                    else:  # Legacy dict format (for backward compatibility)
                        out_ws.append([
                            firm,
                            att.get("name", ""),
                            att.get("title", ""),
                            att.get("office", ""),  # Single office field
                            "",  # Departments (not in legacy)
                            att.get("practice", ""),  # Single practice field
                            "",  # Industries (not in legacy)
                            "",  # Bar admissions (not in legacy)
                            "",  # Education (not in legacy)
                            att.get("enrichment_status", "UNKNOWN"),
                            "",  # Missing fields (not in legacy)
                            att.get("url", ""),
                            "firm_website"  # Data source (legacy format always from firm website)
                        ])
                    total_processed += 1
                    # Stream to JSONL
                    if hasattr(att, 'to_dict'):
                        self.jsonl_file.write(
                            json.dumps(att.to_dict(), ensure_ascii=False, default=str) + "\n"
                        )
                        self.jsonl_file.flush()

                # Save metrics and attorneys JSON
                self._save_metrics(firm, metrics)
                self._save_attorneys_json(firm, attorneys)

                self.log(f"[OK] {firm}: {len(attorneys)} attorneys")
                wb.save(excel_path)
            except Exception as e:
                self.log(f"[ERROR] {firm}: {e}")
                import traceback
                traceback.print_exc()

        self.log(f"\nTotal: {total_processed} attorneys")

        # Generate source failure report if any failures occurred
        if self.source_failures:
            self._save_source_failure_report(excel_path)

        # Write firm-level summary CSV
        _summary_path = self.output_dir / "firm_level_summary.csv"
        self.firm_summary_writer.write(str(_summary_path))
        self.log(f"  Firm summary saved: {_summary_path}")

        # Save aggregated coverage metrics and close JSONL
        self.coverage_engine.save_run_metrics(
            self.all_coverage_metrics,
            output_path=self.output_dir / "coverage_metrics.json"
        )
        if hasattr(self, 'jsonl_file') and self.jsonl_file:
            self.jsonl_file.close()

        return total_processed

    def _save_metrics(self, firm_name: str, metrics: DiscoveryMetrics) -> None:
        """Save discovery metrics to JSON"""
        safe_name = re.sub(r'[^\w\s-]', '', firm_name).strip().replace(' ', '_')
        metrics_path = self.debug_dir / f"{safe_name}_metrics.json"

        with open(metrics_path, 'w', encoding='utf-8') as f:
            json.dump(metrics.to_dict(), f, indent=2)

        self.log(f"  Metrics saved: {metrics_path}")

    def _save_attorneys_json(self, firm_name: str, attorneys: list) -> None:
        """Save attorney profiles to JSON with full fidelity

        Args:
            firm_name: Name of the firm
            attorneys: List of AttorneyProfile objects or dicts
        """
        safe_name = re.sub(r'[^\w\s-]', '', firm_name).strip().replace(' ', '_')
        json_path = self.debug_dir / f"{safe_name}_attorneys.json"

        # Convert AttorneyProfile dataclasses to dicts
        attorneys_data = []
        for att in attorneys:
            if hasattr(att, 'full_name'):  # AttorneyProfile dataclass
                from attorney_extractor import EducationRecord

                att_dict = {
                    "firm": att.firm,
                    "profile_url": att.profile_url,
                    "full_name": att.full_name,
                    "title": att.title,
                    "offices": att.offices,
                    "department": att.department,
                    "practice_areas": att.practice_areas,
                    "industries": att.industries,
                    "bar_admissions": att.bar_admissions,
                    "education": [
                        {
                            "degree": edu.degree,
                            "school": edu.school,
                            "year": edu.year
                        }
                        for edu in att.education
                    ],
                    "extraction_status": att.extraction_status,
                    "missing_fields": att.missing_fields,
                    "diagnostics": att.diagnostics
                }
                attorneys_data.append(att_dict)
            else:  # Legacy dict format
                attorneys_data.append(att)

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(attorneys_data, f, indent=2, ensure_ascii=False)

        self.log(f"  Attorney profiles saved: {json_path}")

    def _save_source_failure_report(self, excel_path: str) -> None:
        """Generate source_failure_report.xlsx with all source-level failures.

        Output columns:
        - Firm, Source URL, Source Type, Failure Type, HTTP Status
        - Fields Expected, Fields Extracted, Error Message, Timestamp
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        if not self.source_failures:
            return

        # Create report file path next to input file
        base_name = os.path.splitext(os.path.basename(excel_path))[0]
        report_path = os.path.join(os.path.dirname(excel_path), f"{base_name}_source_failure_report.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Source Failures"

        # Header row with styling
        headers = [
            "Firm", "Source URL", "Source Type", "Failure Type", "HTTP Status",
            "Fields Expected", "Fields Extracted", "Error Message", "Timestamp"
        ]
        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        # Add failure rows
        for failure in self.source_failures:
            ws.append(failure.to_row())

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(report_path)
        self.log(f"\n=== Source Failure Report: {report_path} ({len(self.source_failures)} failures) ===")

    def process_firm(self, firm_name: str, base_url: str) -> tuple[list[dict], DiscoveryMetrics]:
        """REFACTORED: Self-adapting attorney discovery engine

        Flow:
        1. Check for multi-source data (from firm_finder_desktop.py discovery)
        2. If multi-source available, use source map for targeted extraction
        3. Otherwise, fall back to firm type detection and strategy execution
        4. Collect URLs (discovery only)
        5. Enrich URLs with metadata
        6. Aggregate data across sources with field precedence
        """
        self.log(f"\nProcessing: {firm_name}...")
        start = time.time()
        firm_start_time = start  # alias for firm-level timeout tracking
        firm_timeout_exceeded = False

        # Compliance gate: check robots.txt + bot-wall before any crawl
        compliance_result = self.compliance_engine.check(
            firm=firm_name,
            base_url=base_url,
            paths_to_check=["/people", "/attorneys", "/professionals", "/sitemap.xml"],
        )
        self.rate_limit_manager.apply_compliance(compliance_result)
        if compliance_result.accessibility in (CLASS_BLOCKED_BY_BOT, CLASS_AUTH_REQUIRED):
            _cr = CoverageReport(firm=firm_name)
            _cr.final_status = STATUS_LEGALLY_INCOMPLETE
            _cr.failure_reason = compliance_result.accessibility
            _cr.notes.append(f"Legally inaccessible: {compliance_result.accessibility}")
            _cm = self.coverage_engine.compute(
                firm=firm_name,
                profiles=[],
                legally_incomplete=True,
                legally_incomplete_reason=compliance_result.accessibility,
            )
            self.all_coverage_metrics.append(_cm)
            self.log(f"  [COMPLIANCE BLOCKED] {compliance_result.accessibility} - skipping firm")
            # FirmSummaryWriter: record legally-blocked firm
            _blocked_loop = CoverageLoopResult(
                firm=firm_name,
                expected_total=None,
                expected_total_source="unknown",
                discovered_urls=0,
                extracted_count=0,
                coverage_ratio=None,
                status="LEGALLY_INCOMPLETE",
                legally_incomplete_reason=compliance_result.accessibility,
                sources_tried=[],
                gaps_remaining=0,
                notes=[f"Legally inaccessible: {compliance_result.accessibility}"],
            )
            _blocked_row = FirmSummaryWriter.build_row(
                firm=firm_name,
                loop_result=_blocked_loop,
                attorneys=[],
                us_attorneys=[],
            )
            self.firm_summary_writer.add(_blocked_row)
            _lm = DiscoveryMetrics()
            _lm.failure_notes = _cr.notes
            return [], _lm

        # Reset firm-level enrichment mode per firm
        self.enrichment_mode = "REQUESTS"

        # Check for multi-source configuration
        source_map = self.get_source_map(firm_name)

        if source_map and source_map.sources:
            # MULTI-SOURCE MODE: Use pre-discovered sources
            return self._process_firm_multi_source(firm_name, base_url, source_map, start)

        # SINGLE-SOURCE MODE: Original behavior
        # Phase 1: Detect Firm Type
        profile = self.detect_firm_type(firm_name, base_url)
        self._save_firm_profile(profile)

        # If bot protection detected during validation/observation, switch to Playwright-only
        if any("bot protection" in s.lower() or "cloudflare" in s.lower() or "recaptcha" in s.lower() for s in profile.signals):
            self.enrichment_mode = "PLAYWRIGHT_ONLY"
            self.log("  Enrichment mode set to PLAYWRIGHT_ONLY (bot protection detected)")

        # Phase 2: Select Strategies
        strategies = self.select_strategies(profile, firm_name)

        # Phase 3: Execute Strategies via CoverageLoop
        coverage_report = CoverageReport(firm=firm_name)
        all_profile_urls = set()
        loop_result: CoverageLoopResult | None = None

        # Build strategy_fns: each strategy -> zero-arg callable returning set[str]
        # _execute_strategy returns (urls, expected_total); cache for coverage_report back-fill
        _strategy_cache: dict[str, tuple[set[str], int | None]] = {}

        def _make_strategy_fn(strat: str):
            def _fn() -> set[str]:
                s_urls, s_exp = self._execute_strategy(strat, base_url, profile)
                _strategy_cache[strat] = (s_urls, s_exp)
                return s_urls
            return _fn

        strategy_fns = {s: _make_strategy_fn(s) for s in strategies}
        # Use extended timeout for Playwright-based strategies (they legitimately take longer)
        _PLAYWRIGHT_STRATEGIES = {"kirkland_scroll"}
        _uses_playwright = bool(set(strategies) & _PLAYWRIGHT_STRATEGIES)
        firm_hard_timeout = 7200 if _uses_playwright else MAX_FIRM_TIME

        # Resolve expected_total before loop (use directory page text if available)
        page_text = getattr(profile, 'directory_page_text', None)
        expected_total_result = self.expected_total_resolver.resolve(
            page_text=page_text,
        )

        cov_loop = CoverageLoop(
            firm=firm_name,
            expected_total_result=expected_total_result,
            strategy_fns=strategy_fns,
            strategy_order=strategies,
            log_fn=self.log,
            hard_timeout=firm_hard_timeout,
            limit=self.limit,
        )
        loop_result = cov_loop.run()
        # TASK 5: Detect firm timeout from elapsed wall-clock time
        if time.time() - firm_start_time > firm_hard_timeout:
            firm_timeout_exceeded = True
        all_profile_urls = cov_loop.all_urls

        # Back-fill coverage_report from loop execution
        for strat, (s_urls, _) in _strategy_cache.items():
            if s_urls:
                coverage_report.url_sources[strat] = len(s_urls)
                coverage_report.strategies_attempted.append(strat)
        coverage_report.expected_total = loop_result.expected_total
        coverage_report.discovered_urls = loop_result.discovered_urls

        # Fix 1: Apply locale filter BEFORE limit so sampling is not wasted on non-English URLs
        _pre_locale_single = len(all_profile_urls)
        all_profile_urls = {u for u in all_profile_urls if not _is_locale_url(u)}
        if len(all_profile_urls) < _pre_locale_single:
            self.log(f"  [PRE_FILTER] Removed {_pre_locale_single - len(all_profile_urls)} non-English locale URLs before limit")

        # Fix 2 (Latham debug): ignore --limit for Latham to verify enrichment works
        _latham_debug = "latham" in firm_name.lower()
        if _latham_debug:
            self.log(f"[DEBUG] Latham debug mode: ignoring --limit, processing all discovered URLs")

        # Apply limit AFTER locale filter
        if not _latham_debug and self.limit > 0 and len(all_profile_urls) > self.limit:
            self.log(f"\n--- Applying limit: {self.limit} URLs (from {len(all_profile_urls)} discovered) ---")
            all_profile_urls = set(sorted(all_profile_urls)[:self.limit])
        elif _latham_debug and len(all_profile_urls) < 200:
            self.log(f"[DEBUG] Latham: {len(all_profile_urls)} URLs after locale filter (target >=200, continuing)")
        coverage_report.discovered_urls = len(all_profile_urls)

        # Phase 3.5: URL QUALITY GATE (filter non-profiles)
        if len(all_profile_urls) > 100 and self.enrichment_mode != "PLAYWRIGHT_ONLY":
            self.log(f"\n--- Running URL Quality Gate ({len(all_profile_urls)} candidates) ---")
            from profile_quality_gate import URLQualityGate

            quality_gate = URLQualityGate(session=self.session, timeout=REQUEST_TIMEOUT)
            gate_result = quality_gate.filter_candidates(
                candidate_urls=all_profile_urls,
                base_url=base_url,
                sample_size=min(200, len(all_profile_urls)),
                min_confidence=60.0
            )

            # Log quality gate results
            self.log(f"  Sampled: {gate_result.sampled_count} URLs")
            self.log(f"  Positive samples: {gate_result.positive_samples}")
            self.log(f"  Pattern confidence: {gate_result.inferred_pattern.confidence:.2f}")
            self.log(f"  Filtered URLs: {len(gate_result.filtered_profile_urls)} (from {gate_result.total_candidates})")
            self.log(f"  Rejected: {len(gate_result.rejected_urls)}")

            if gate_result.rejection_reasons:
                self.log(f"  Top rejection reasons:")
                for reason, count in sorted(gate_result.rejection_reasons.items(), key=lambda x: x[1], reverse=True)[:5]:
                    self.log(f"    {reason}: {count}")

            # Save quality gate report
            safe_name = re.sub(r'[^\w\s-]', '', firm_name).strip().replace(' ', '_')
            gate_report_path = self.debug_dir / f"{safe_name}_quality_gate.json"
            with open(gate_report_path, 'w', encoding='utf-8') as f:
                json.dump(gate_result.to_dict(), f, indent=2)
            self.log(f"  Quality gate report saved: {gate_report_path}")

            # Use filtered URLs
            all_profile_urls = gate_result.filtered_profile_urls
            coverage_report.discovered_urls = len(all_profile_urls)  # Update count

        # Phase 4: Determine Final Status
        coverage_report.final_status = self._determine_final_status(coverage_report)

        # TASK 5: Override status if firm timeout occurred
        if firm_timeout_exceeded:
            coverage_report.final_status = STATUS_PARTIAL
            coverage_report.notes.append(f"Firm timeout exceeded ({MAX_FIRM_TIME}s) - partial results")

        self._save_coverage_report(coverage_report)

        # Phase 5: Enrich URLs with Metadata (separate phase)
        # Sort for deterministic results
        urls_to_enrich = sorted(all_profile_urls)
        # --- LOCALE FILTER (Step 2c): reject non-English locale URLs ---
        _pre_locale = len(urls_to_enrich)
        urls_to_enrich = [u for u in urls_to_enrich if not _is_locale_url(u)]
        if len(urls_to_enrich) < _pre_locale:
            self.log(f"  [LOCALE_FILTER] Removed {_pre_locale - len(urls_to_enrich)} non-English locale URLs")


        self.log(f"\n--- Enriching {len(urls_to_enrich)} profile URLs ---")
        # Kirkland (and other JS-SPA firms) require Playwright rendering for profile pages.
        # requests.get() returns an empty shell HTML for these sites.
        _js_spa_firms = ['kirkland']
        if any(tok in firm_name.lower() for tok in _js_spa_firms):
            if self.enrichment_mode != 'PLAYWRIGHT_ONLY':
                self.enrichment_mode = 'PLAYWRIGHT_ONLY'
                self.log(f"  [ENRICHMENT] Force PLAYWRIGHT_ONLY for JS-SPA firm: {firm_name}")
        attorneys = self._enrich_profile_urls(urls_to_enrich, base_url, firm_name)

        # Phase 5.5: Determine Discovery Status and External Directory Fallback
        discovery_status, failure_reason = self._determine_discovery_status(coverage_report, attorneys)
        if discovery_status == DISCOVERY_BLOCKED and self.enrichment_mode != "PLAYWRIGHT_ONLY":
            self.log("  All profiles blocked under requests; retrying with Playwright batch mode")
            self.enrichment_mode = "PLAYWRIGHT_ONLY"
            attorneys = self._enrich_profile_urls(urls_to_enrich, base_url, firm_name)
            discovery_status, failure_reason = self._determine_discovery_status(coverage_report, attorneys)
        coverage_report.discovery_status = discovery_status
        coverage_report.failure_reason = failure_reason

        # Save updated coverage report with discovery status
        self._save_coverage_report(coverage_report)

        # Trigger external directory fallback for BLOCKED or DISCOVERY_INCOMPLETE
        should_fallback = discovery_status in [DISCOVERY_BLOCKED, DISCOVERY_INCOMPLETE]

        if should_fallback:
            self.log(f"\n--- Discovery Status: {discovery_status} ({failure_reason}) ---")
            self.log(f"--- Activating External Directory Fallback ---")
            external_attorneys = self._fetch_external_directory_data(firm_name, attorneys)

            if external_attorneys:
                # Merge external directory data with firm website data
                attorneys = self._merge_external_data(attorneys, external_attorneys)
                self.log(f"  Merged {len(external_attorneys)} external directory profiles")
                for att in attorneys:
                    if hasattr(att, 'diagnostics'):
                        att.diagnostics['source_origin'] = 'external'
            else:
                self.log(f"  No external directory data found")

        # Phase 5b: Filter to US-only attorneys
        us_attorneys = filter_us_attorneys(attorneys, log_fn=self.log)
        self.log(f"  US attorneys: {len(us_attorneys)}/{len(attorneys)}")
        elapsed = time.time() - start
        self.log(f"\n[DONE] {len(attorneys)} attorneys ({elapsed:.1f}s)")
        self.log(f"  Status: {coverage_report.final_status}")
        self.log(f"  Discovery: {coverage_report.discovery_status}" + (f" ({coverage_report.failure_reason})" if coverage_report.failure_reason else ""))
        if coverage_report.expected_total:
            ratio = coverage_report.discovered_urls / coverage_report.expected_total
            self.log(f"  Coverage: {coverage_report.discovered_urls}/{coverage_report.expected_total} ({ratio*100:.1f}%)")

        # Legacy DiscoveryMetrics for backward compatibility
        metrics = DiscoveryMetrics()
        metrics.discovered_unique = len(all_profile_urls)
        metrics.expected_total = coverage_report.expected_total or 0
        metrics.failure_notes = coverage_report.notes

        # Record CoverageMetrics for this firm
        _cov = self.coverage_engine.compute(
            firm=firm_name,
            profiles=attorneys,
            discovered_urls=len(all_profile_urls),
            blocked_count=0,
            expected_total=coverage_report.expected_total or None,
        )
        self.all_coverage_metrics.append(_cov)

        # FirmSummaryWriter: record per-firm row (normal path)
        if loop_result is not None:
            _summary_row = FirmSummaryWriter.build_row(
                firm=firm_name,
                loop_result=loop_result,
                attorneys=attorneys,
                us_attorneys=us_attorneys,
            )
            self.firm_summary_writer.add(_summary_row)

        return attorneys, metrics

    # ========================================================================
    # MULTI-SOURCE PROCESSING
    # ========================================================================

    def _process_firm_multi_source(
        self,
        firm_name: str,
        base_url: str,
        source_map: FirmSourceMap,
        start: float
    ) -> tuple[list[dict], DiscoveryMetrics]:
        """Process firm using pre-discovered multi-source configuration

        Flow:
        1. Get attorney URLs from attorney_list sources
        2. Enrich profiles using profile_core sources
        3. Supplement with secondary sources (education, bar_admission)
        4. Aggregate data with field precedence
        """
        self.log(f"\n=== Multi-Source Mode: {len(source_map.sources)} sources available ===")

        # Log source breakdown
        attorney_lists = source_map.get_attorney_list_sources()
        profile_sources = source_map.get_profile_sources()
        secondary_sources = source_map.get_secondary_sources()

        self.log(f"  Attorney list sources: {len(attorney_lists)}")
        self.log(f"  Profile sources: {len(profile_sources)}")
        self.log(f"  Secondary sources: {len(secondary_sources)}")

        # Phase 1: Collect attorney URLs from attorney_list sources
        all_profile_urls = set()
        coverage_report = CoverageReport(firm=firm_name)

        for source in attorney_lists:
            source_url = source.get('source_url', '')
            source_type = source.get('source_type', '')
            method = source.get('discovery_method', '')
            fields_str = source.get('fields_provided', '')
            expected_fields = [f.strip() for f in fields_str.split(',') if f.strip()] if fields_str else []

            self.log(f"\n--- Fetching URLs from: {source_url} ({method}) ---")

            try:
                urls, failure = self._fetch_urls_from_source_with_tracking(
                    source, base_url, firm_name, expected_fields
                )
                if urls:
                    all_profile_urls.update(urls)
                    coverage_report.url_sources[f"{method}:{source_url[:50]}"] = len(urls)
                    self.log(f"  Collected {len(urls)} URLs")
                elif failure:
                    # Source failed but don't fail the whole firm
                    self.source_failures.append(failure)
                    self.log(f"  Source failed: {failure.failure_type} - continuing with other sources")
            except Exception as e:
                # Track unexpected exceptions as source failures
                import datetime
                failure = SourceFailure(
                    firm=firm_name,
                    source_url=source_url,
                    source_type=source_type,
                    failure_type="exception",
                    fields_expected=expected_fields,
                    error_message=str(e),
                    timestamp=datetime.datetime.now().isoformat()
                )
                self.source_failures.append(failure)
                self.log(f"  Error fetching from source: {e} - continuing with other sources")

        coverage_report.discovered_urls = len(all_profile_urls)
        coverage_report.strategies_attempted.append("multi_source")

        self.log(f"\n  Total unique URLs collected: {len(all_profile_urls)}")

        # Fix 1+2: Apply locale filter BEFORE limit (multi-source path)
        urls_to_enrich = sorted(all_profile_urls)
        # --- LOCALE FILTER: reject non-English locale URLs BEFORE limit ---
        _pre_locale_ms = len(urls_to_enrich)
        urls_to_enrich = [u for u in urls_to_enrich if not _is_locale_url(u)]
        if len(urls_to_enrich) < _pre_locale_ms:
            self.log(f"  [LOCALE_FILTER] Removed {_pre_locale_ms - len(urls_to_enrich)} non-English locale URLs")
        # Fix 2 (Latham debug): ignore limit for Latham
        _latham_debug_ms = "latham" in firm_name.lower()
        if _latham_debug_ms:
            self.log(f"[DEBUG] Processing {len(urls_to_enrich)} Latham URLs after locale filter")
        elif self.limit > 0 and len(urls_to_enrich) > self.limit:
            urls_to_enrich = urls_to_enrich[:self.limit]
            self.log(f"  Limiting to {self.limit} URLs for enrichment")

        # Phase 3: Enrich profiles (standard enrichment)
        self.log(f"\n--- Enriching {len(urls_to_enrich)} profile URLs ---")
        attorneys = self._enrich_profile_urls(urls_to_enrich, base_url, firm_name)

        # Phase 4: Supplement with secondary sources (education, bar admissions)
        if secondary_sources:
            self.log(f"\n--- Supplementing with {len(secondary_sources)} secondary sources ---")
            attorneys = self._supplement_from_secondary_sources(
                attorneys, secondary_sources, base_url, firm_name
            )

        # Phase 5: Determine status and log
        coverage_report.final_status = self._determine_final_status(coverage_report)
        discovery_status, failure_reason = self._determine_discovery_status(coverage_report, attorneys)
        coverage_report.discovery_status = discovery_status
        coverage_report.failure_reason = failure_reason

        self._save_coverage_report(coverage_report)

        elapsed = time.time() - start
        self.log(f"\n[DONE] {len(attorneys)} attorneys ({elapsed:.1f}s) [Multi-Source Mode]")
        self.log(f"  Status: {coverage_report.final_status}")
        self.log(f"  Discovery: {coverage_report.discovery_status}")

        # Legacy metrics
        metrics = DiscoveryMetrics()
        metrics.discovered_unique = len(all_profile_urls)
        metrics.expected_total = coverage_report.expected_total or 0

        # Record CoverageMetrics for this firm
        _cov = self.coverage_engine.compute(
            firm=firm_name,
            profiles=attorneys,
            discovered_urls=len(all_profile_urls),
            blocked_count=0,
            expected_total=coverage_report.expected_total or None,
        )
        self.all_coverage_metrics.append(_cov)
        # Phase 5b: Filter to US-only attorneys (multi-source path)
        us_attorneys_ms = filter_us_attorneys(attorneys, log_fn=self.log)
        self.log(f"  US attorneys: {len(us_attorneys_ms)}/{len(attorneys)}")

        # FirmSummaryWriter: record per-firm row (multi-source path)
        _ms_loop = CoverageLoopResult(
            firm=firm_name,
            expected_total=coverage_report.expected_total,
            expected_total_source="multi_source_discovery",
            discovered_urls=coverage_report.discovered_urls,
            extracted_count=len(attorneys),
            coverage_ratio=(
                coverage_report.discovered_urls / coverage_report.expected_total
                if coverage_report.expected_total else None
            ),
            status=coverage_report.final_status,
            legally_incomplete_reason=(
                coverage_report.failure_reason
                if coverage_report.final_status == STATUS_LEGALLY_INCOMPLETE else None
            ),
            sources_tried=[s.get('source_url', '') for s in source_map.sources],
            gaps_remaining=max(0, (coverage_report.expected_total or 0) - coverage_report.discovered_urls),
            notes=coverage_report.notes,
        )
        _ms_row = FirmSummaryWriter.build_row(
            firm=firm_name,
            loop_result=_ms_loop,
            attorneys=attorneys,
            us_attorneys=us_attorneys_ms,
        )
        self.firm_summary_writer.add(_ms_row)

        return attorneys, metrics

    def _fetch_urls_from_source(self, source: dict, base_url: str) -> set[str]:
        """Fetch attorney profile URLs from a discovered source

        Handles different discovery methods:
        - xml_sitemap: Parse XML sitemap
        - html_directory: Parse HTML directory page
        - known_pattern: JSON API endpoint
        """
        urls = set()
        source_url = source.get('source_url', '')
        method = source.get('discovery_method', '')
        domain = urlparse(base_url).netloc

        if not source_url:
            return urls

        try:
            try:
                self.rate_limit_manager.wait(domain)
            except RateLimitBlockedError:
                return urls
            response = self.session.get(source_url, timeout=REQUEST_TIMEOUT)

            if response.status_code != 200:
                self.log(f"    Source returned {response.status_code}")
                return urls

            if method == "xml_sitemap":
                # Parse XML sitemap
                try:
                    root = ET.fromstring(response.content)
                    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}

                    url_elements = root.findall(".//sm:url/sm:loc", ns)
                    if not url_elements:
                        url_elements = root.findall(".//url/loc")

                    for elem in url_elements:
                        if elem.text and self._is_profile_like_url(elem.text, domain):
                            urls.add(elem.text)
                except ET.ParseError:
                    self.log(f"    Failed to parse XML sitemap")

            elif method == "html_directory":
                # Parse HTML directory
                import re
                links = re.findall(r'href=["\'](.*?)["\']', response.text)
                for link in links:
                    full_url = urljoin(base_url, link)
                    if self._is_profile_like_url(full_url, domain):
                        urls.add(full_url)

            elif method == "known_pattern":
                # JSON API response
                try:
                    data = response.json()
                    if isinstance(data, list):
                        for item in data:
                            if isinstance(item, dict):
                                url = item.get('url') or item.get('profileUrl') or item.get('link')
                                if url:
                                    full_url = urljoin(base_url, url)
                                    urls.add(full_url)
                except json.JSONDecodeError:
                    self.log(f"    Failed to parse JSON response")
            else:
                # Default: try HTML parsing
                import re
                links = re.findall(r'href=["\'](.*?)["\']', response.text)
                for link in links:
                    full_url = urljoin(base_url, link)
                    if self._is_profile_like_url(full_url, domain):
                        urls.add(full_url)

        except Exception as e:
            self.log(f"    Error fetching source: {e}")

        return urls

    def _fetch_urls_from_source_with_tracking(
        self,
        source: dict,
        base_url: str,
        firm_name: str,
        expected_fields: list[str]
    ) -> tuple[set[str], SourceFailure | None]:
        """Fetch attorney profile URLs from a discovered source with failure tracking.

        Returns:
            (urls, failure): urls if successful, failure object if failed
        """
        import datetime

        urls = set()
        source_url = source.get('source_url', '')
        source_type = source.get('source_type', '')
        method = source.get('discovery_method', '')
        domain = urlparse(base_url).netloc

        if not source_url:
            return urls, None

        try:
            try:
                self.rate_limit_manager.wait(domain)
            except RateLimitBlockedError:
                failure = SourceFailure(
                    firm=firm_name,
                    source_url=source_url,
                    source_type=source_type,
                    failure_type="blocked",
                    error_message="RateLimitBlocked",
                    timestamp=datetime.datetime.now().isoformat()
                )
                return set(), failure
            response = self.session.get(source_url, timeout=REQUEST_TIMEOUT)

            if response.status_code != 200:
                self.log(f"    Source returned {response.status_code}")
                failure = SourceFailure(
                    firm=firm_name,
                    source_url=source_url,
                    source_type=source_type,
                    failure_type="http_error",
                    http_status=response.status_code,
                    fields_expected=expected_fields,
                    timestamp=datetime.datetime.now().isoformat()
                )
                # Check for blocking (403, 429)
                if response.status_code in [403, 429, 503]:
                    failure.failure_type = "blocked"
                return urls, failure

            # Check for very small response (likely blocked/redirected)
            if len(response.content) < 500:
                failure = SourceFailure(
                    firm=firm_name,
                    source_url=source_url,
                    source_type=source_type,
                    failure_type="empty_response",
                    http_status=response.status_code,
                    fields_expected=expected_fields,
                    error_message=f"Response too small ({len(response.content)} bytes)",
                    timestamp=datetime.datetime.now().isoformat()
                )
                return urls, failure

            # Parse based on method
            if method == "xml_sitemap":
                try:
                    root = ET.fromstring(response.content)
                    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}

                    url_elements = root.findall(".//sm:url/sm:loc", ns)
                    if not url_elements:
                        url_elements = root.findall(".//url/loc")

                    for elem in url_elements:
                        if elem.text and self._is_profile_like_url(elem.text, domain):
                            urls.add(elem.text)
                except ET.ParseError as e:
                    failure = SourceFailure(
                        firm=firm_name,
                        source_url=source_url,
                        source_type=source_type,
                        failure_type="parse_error",
                        fields_expected=expected_fields,
                        error_message=f"XML parse error: {e}",
                        timestamp=datetime.datetime.now().isoformat()
                    )
                    return urls, failure

            elif method == "html_directory":
                import re
                links = re.findall(r'href=["\'](.*?)["\']', response.text)
                for link in links:
                    full_url = urljoin(base_url, link)
                    if self._is_profile_like_url(full_url, domain):
                        urls.add(full_url)

            elif method == "known_pattern":
                try:
                    data = response.json()
                    if isinstance(data, list):
                        for item in data:
                            if isinstance(item, dict):
                                url = item.get('url') or item.get('profileUrl') or item.get('link')
                                if url:
                                    full_url = urljoin(base_url, url)
                                    urls.add(full_url)
                except json.JSONDecodeError as e:
                    failure = SourceFailure(
                        firm=firm_name,
                        source_url=source_url,
                        source_type=source_type,
                        failure_type="parse_error",
                        fields_expected=expected_fields,
                        error_message=f"JSON parse error: {e}",
                        timestamp=datetime.datetime.now().isoformat()
                    )
                    return urls, failure
            else:
                # Default: try HTML parsing
                import re
                links = re.findall(r'href=["\'](.*?)["\']', response.text)
                for link in links:
                    full_url = urljoin(base_url, link)
                    if self._is_profile_like_url(full_url, domain):
                        urls.add(full_url)

            # Check if we got any URLs (source works but produced nothing)
            if not urls:
                failure = SourceFailure(
                    firm=firm_name,
                    source_url=source_url,
                    source_type=source_type,
                    failure_type="empty_result",
                    http_status=response.status_code,
                    fields_expected=expected_fields,
                    error_message="Source responded but no profile URLs found",
                    timestamp=datetime.datetime.now().isoformat()
                )
                return urls, failure

        except requests.exceptions.Timeout:
            failure = SourceFailure(
                firm=firm_name,
                source_url=source_url,
                source_type=source_type,
                failure_type="timeout",
                fields_expected=expected_fields,
                error_message=f"Request timeout ({REQUEST_TIMEOUT}s)",
                timestamp=datetime.datetime.now().isoformat()
            )
            return urls, failure
        except requests.exceptions.RequestException as e:
            failure = SourceFailure(
                firm=firm_name,
                source_url=source_url,
                source_type=source_type,
                failure_type="network_error",
                fields_expected=expected_fields,
                error_message=str(e),
                timestamp=datetime.datetime.now().isoformat()
            )
            return urls, failure
        except Exception as e:
            failure = SourceFailure(
                firm=firm_name,
                source_url=source_url,
                source_type=source_type,
                failure_type="exception",
                fields_expected=expected_fields,
                error_message=str(e),
                timestamp=datetime.datetime.now().isoformat()
            )
            return urls, failure

        return urls, None

    def _supplement_from_secondary_sources(
        self,
        attorneys: list,
        secondary_sources: list[dict],
        base_url: str,
        firm_name: str
    ) -> list:
        """Supplement attorney profiles with data from secondary sources

        Secondary sources are page patterns like:
        - [profile_url]/education
        - [profile_url]?tab=credentials

        This method attempts to extract additional fields from these sources.
        """
        from attorney_extractor import AttorneyProfile, EducationRecord

        if not secondary_sources or not attorneys:
            return attorneys

        supplemented_count = 0

        for source in secondary_sources:
            source_url_pattern = source.get('source_url', '')
            source_type = source.get('source_type', '')
            fields_str = source.get('fields_provided', '')

            # Parse fields this source can provide
            expected_fields = [f.strip() for f in fields_str.split(',') if f.strip()]

            if not expected_fields:
                continue

            self.log(f"  Secondary source: {source_type} -> {expected_fields}")

            # Determine URL pattern (e.g., "/education", "?tab=credentials")
            pattern = ""
            if "[profile_url]" in source_url_pattern:
                pattern = source_url_pattern.replace("[profile_url]", "")
            else:
                # Try to infer pattern from source URL
                for suffix in ["/education", "/credentials", "/admissions", "/background",
                              "?tab=education", "?tab=credentials"]:
                    if suffix in source_url_pattern:
                        pattern = suffix
                        break

            if not pattern:
                continue

            # For each attorney, try to fetch supplementary data
            for att in attorneys:
                if not hasattr(att, 'profile_url') or not att.profile_url:
                    continue

                # Check if this attorney needs these fields
                needs_education = "education" in expected_fields and not att.education
                needs_bar = "bar_admissions" in expected_fields and not att.bar_admissions

                if not needs_education and not needs_bar:
                    continue

                # Build secondary URL
                if "?" in pattern:
                    secondary_url = att.profile_url + pattern
                else:
                    secondary_url = att.profile_url.rstrip("/") + pattern

                try:
                    domain = urlparse(base_url).netloc
                    try:
                        self.rate_limit_manager.wait(domain)
                    except RateLimitBlockedError:
                        continue

                    resp = self.session.get(secondary_url, timeout=REQUEST_TIMEOUT)
                    if resp.status_code != 200:
                        continue

                    # Quick extraction from secondary page
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(resp.text, "html.parser")
                    text = soup.get_text(separator=" ", strip=True).lower()

                    # Try to extract education
                    if needs_education and "education" in expected_fields:
                        edu_records = self._quick_extract_education(soup, text)
                        if edu_records:
                            att.education = edu_records
                            att.diagnostics['education_source'] = source_type
                            if 'education' in att.missing_fields:
                                att.missing_fields.remove('education')
                            supplemented_count += 1

                    # Try to extract bar admissions
                    if needs_bar and "bar_admissions" in expected_fields:
                        bars = self._quick_extract_bar_admissions(soup, text)
                        if bars:
                            att.bar_admissions = bars
                            att.diagnostics['bar_admissions_source'] = source_type
                            if 'bar_admissions' in att.missing_fields:
                                att.missing_fields.remove('bar_admissions')
                            supplemented_count += 1

                    # Recalculate status after supplement
                    att.calculate_status()

                except Exception:
                    continue

        self.log(f"  Supplemented {supplemented_count} fields from secondary sources")
        return attorneys

    def _quick_extract_education(self, soup, text: str) -> list:
        """Quick extraction of education from secondary page"""
        from attorney_extractor import EducationRecord

        education = []

        # Look for common education patterns
        edu_patterns = [
            r"(J\.D\.|LL\.M\.|B\.A\.|B\.S\.|M\.A\.|M\.B\.A\.|Ph\.D\.)[,\s]+([^,\n]+?)(?:,\s*(\d{4}))?",
            r"([A-Z][a-z]+ University|[A-Z][a-z]+ College|[A-Z][a-z]+ School of Law)",
        ]

        import re
        for pattern in edu_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches[:5]:  # Limit to 5 education entries
                if isinstance(match, tuple):
                    degree = match[0] if len(match) > 0 else ""
                    school = match[1] if len(match) > 1 else ""
                    year = match[2] if len(match) > 2 else ""
                else:
                    degree = ""
                    school = match
                    year = ""

                if school and len(school) > 3:
                    education.append(EducationRecord(
                        degree=degree.strip() if degree else None,
                        school=school.strip(),
                        year=year.strip() if year else None
                    ))

        return education

    def _quick_extract_bar_admissions(self, soup, text: str) -> list[str]:
        """Quick extraction of bar admissions from secondary page"""
        bars = []

        # US state names for bar admissions
        states = [
            "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
            "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
            "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
            "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
            "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
            "New Hampshire", "New Jersey", "New Mexico", "New York",
            "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
            "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
            "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
            "West Virginia", "Wisconsin", "Wyoming", "District of Columbia"
        ]

        for state in states:
            if state.lower() in text:
                bars.append(state)

        return bars

    # ========================================================================
    # STRATEGY EXECUTION
    # ========================================================================

    def _execute_strategy(self, strategy: str, base_url: str, profile: FirmProfile) -> tuple[set[str], int | None]:
        """Execute a single discovery strategy

        Returns: (profile_urls, expected_total)
        """
        if strategy == "xml_sitemap":
            return self._strategy_xml_sitemap(base_url)
        elif strategy == "html_sitemap":
            return self._strategy_html_sitemap(base_url)
        elif strategy == "api_enumeration":
            return self._strategy_api_enumeration(base_url)
        elif strategy == "filter_enumeration":
            return self._strategy_filter_enumeration(base_url)
        elif strategy == "dom_exhaustion":
            return self._strategy_dom_exhaustion(base_url)
        elif strategy == "hard_case_fallback":
            return self._strategy_hard_case_fallback(base_url)
        elif strategy == "directory_listing":
            return self._strategy_directory_listing(base_url)
        elif strategy == "alphabet_enumeration":
            return self._strategy_alphabet_enumeration(base_url)
        elif strategy == "xml_sitemap_navigation":
            return self._strategy_xml_sitemap_navigation(base_url)
        elif strategy == "kirkland_scroll":
            return self._strategy_kirkland_scroll(base_url)
        else:
            print(f"[ERROR] Unknown strategy: {strategy}")
            return set(), None

    def _strategy_xml_sitemap(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Extract URLs from XML sitemap with source validation"""
        urls = set()
        try:
            profile_urls = self._extract_profile_urls_from_sitemap(base_url)
            self.log(f"  XML Sitemap: {len(profile_urls)} profile URLs found")

            # CRITICAL: Validate source before accepting all URLs
            if len(profile_urls) > 0:
                from source_validator import validate_source

                validation_result = validate_source(
                    source_url=base_url + "/sitemap.xml",
                    source_type="xml_sitemap",
                    candidate_urls=list(profile_urls),
                    session=self.session,
                    sample_size=min(3, len(profile_urls)),
                    timeout=5  # Quick validation
                )

                if not validation_result.is_valid:
                    self.log(f"  Sitemap validation FAILED: {validation_result.failure_reason}")
                    self.log(f"  Field validation: {validation_result.field_validation}")

                    # Log source failure using global SourceFailure dataclass
                    self.source_failures.append(SourceFailure(
                        firm=firm_name if hasattr(self, 'current_firm') else "unknown",
                        source_url=base_url + "/sitemap.xml",
                        source_type="xml_sitemap",
                        failure_type="validation_failed",
                        error_message=validation_result.failure_reason or "validation_failed"
                    ))

                    # Reject source - return empty
                    return set(), None
                else:
                    self.log(f"  Sitemap validation PASSED")
                    self.log(f"  Accepting {len(profile_urls)} URLs from validated source")
                    urls.update(profile_urls)

        except Exception as e:
            self.log(f"  XML Sitemap error: {e}")

        return urls, None  # Sitemaps don't provide expected totals

    def _strategy_html_sitemap(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Extract URLs from HTML sitemap"""
        urls = set()
        domain = urlparse(base_url).netloc

        try:
            # Strategy 1: Check robots.txt sitemaps
            sitemap_urls = self._check_for_sitemap(base_url)

            for sitemap_url in sitemap_urls:
                if self._classify_sitemap(sitemap_url, domain) == "HTML":
                    # Parse as HTML
                    try:
                        self.rate_limit_manager.wait(domain)
                    except RateLimitBlockedError:
                        continue
                    resp = self.session.get(sitemap_url, timeout=REQUEST_TIMEOUT)
                    if resp.status_code == 200:
                        # Extract all links from HTML
                        import re
                        links = re.findall(r'href=["\'](.*?)["\']', resp.text)
                        for link in links:
                            full_url = urljoin(base_url, link)
                            if self._is_profile_like_url(full_url, domain):
                                urls.add(full_url)

            # Strategy 2: Probe known HTML sitemap paths (comprehensive)
            if len(urls) == 0:
                self.log("  No HTML sitemaps from robots.txt, probing known paths...")
                html_sitemap_findings = self._probe_sitemap_paths(base_url)

                # FIX: _probe_sitemap_paths returns {path: (type, profile_count)}
                for path, (sitemap_type, profile_count) in html_sitemap_findings.items():
                    if sitemap_type == "HTML":
                        scheme = urlparse(base_url).scheme
                        full_url = f"{scheme}://{domain}{path}"

                        try:
                            self.rate_limit_manager.wait(domain)
                        except RateLimitBlockedError:
                            continue
                        resp = self.session.get(full_url, timeout=REQUEST_TIMEOUT)
                        if resp.status_code == 200:
                            import re
                            links = re.findall(r'href=["\'](.*?)["\']', resp.text)
                            for link in links:
                                full_link = urljoin(base_url, link)
                                if self._is_profile_like_url(full_link, domain):
                                    urls.add(full_link)

            # Strategy 3: Fallback to known Kirkland-style sitemap paths
            if len(urls) == 0:
                fallback_paths = ["/sitemap/lawyers", "/sitemap/professionals"]
                scheme = urlparse(base_url).scheme
                for path in fallback_paths:
                    try:
                        full_url = f"{scheme}://{domain}{path}"
                        try:
                            self.rate_limit_manager.wait(domain)
                        except RateLimitBlockedError:
                            continue
                        resp = self.session.get(full_url, timeout=REQUEST_TIMEOUT)
                        if resp.status_code == 200:
                            import re
                            links = re.findall(r'href=["\'](.*?)["\']', resp.text)
                            for link in links:
                                full_link = urljoin(base_url, link)
                                if self._is_profile_like_url(full_link, domain):
                                    urls.add(full_link)
                    except Exception:
                        pass

            self.log(f"  HTML Sitemap: {len(urls)} profile URLs found")
        except Exception as e:
            self.log(f"  HTML Sitemap error: {e}")

        return urls, None

    def _strategy_api_enumeration(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Use Playwright to capture and enumerate API"""
        # This reuses the existing Playwright path but focuses on API
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            self.log("  Playwright not available")
            return set(), None

        urls = set()
        expected_total = None
        captured_api_data = []

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=not self.headful)
            context = browser.new_context()

            def handle_route(route):
                if route.request.resource_type in ["image", "font", "media", "stylesheet"]:
                    route.abort()
                else:
                    route.continue_()

            context.route("**/*", handle_route)
            page = context.new_page()

            def handle_response(response):
                try:
                    url_lower = response.url.lower()
                    if any(kw in url_lower for kw in ["api", "graphql", "search", "people", "attorneys", "lawyers", "professionals"]):
                        if response.ok and "json" in response.headers.get("content-type", "").lower():
                            try:
                                data = response.json()
                                records = self._extract_attorneys_from_json(data, base_url)
                                if records:
                                    captured_api_data.append({"url": response.url, "records": records, "json": data})
                            except Exception:
                                pass
                except Exception:
                    pass

            page.on("response", handle_response)

            # Navigate to likely directory page
            probe_paths = ["/people", "/professionals", "/attorneys", "/lawyers"]
            for path in probe_paths:
                try:
                    target = urljoin(base_url, path)
                    page.goto(target, timeout=15000, wait_until="networkidle")
                    page.wait_for_timeout(2000)
                    if captured_api_data:
                        break
                except Exception:
                    pass

            # Enumerate API if captured
            if captured_api_data:
                for api_data in captured_api_data:
                    # Add initial records
                    for rec in api_data["records"]:
                        if rec.get("url"):
                            urls.add(rec["url"])
                            rec['_json_api_used'] = True  # JSON-priority flag (Step 2g)

                    # Try pagination
                    paginated_urls = self._paginate_api_smart(api_data["url"], api_data.get("json"), base_url)
                    urls.update(paginated_urls)

                    # Extract expected total
                    if not expected_total and api_data.get("json"):
                        expected_total = self._extract_total_from_json(api_data["json"])

            page.close()
            context.close()
            browser.close()

        self.log(f"  API Enumeration: {len(urls)} URLs, expected: {expected_total or 'unknown'}")
        return urls, expected_total

    def _strategy_filter_enumeration(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Enumerate through filters

        FALLBACK: If not implemented, automatically fall back to dom_exhaustion
        """
        self.log("  Filter enumeration strategy not yet implemented")
        self.log("  Falling back to dom_exhaustion...")

        # Automatic fallback to DOM exhaustion
        return self._strategy_dom_exhaustion(base_url)

    def _strategy_dom_exhaustion(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Exhaust pagination/load-more/scroll"""
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            return set(), None

        urls = set()
        expected_total = None

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=not self.headful)
            context = browser.new_context()

            def handle_route(route):
                if route.request.resource_type in ["image", "font", "media", "stylesheet"]:
                    route.abort()
                else:
                    route.continue_()

            context.route("**/*", handle_route)
            page = context.new_page()

            # Find directory page
            probe_paths = ["/people", "/professionals", "/attorneys", "/lawyers"]
            directory_url = None

            for path in probe_paths:
                try:
                    target = urljoin(base_url, path)
                    response = page.goto(target, timeout=15000, wait_until="domcontentloaded")
                    if response and response.ok:
                        page.wait_for_timeout(1000)
                        links = page.evaluate("() => Array.from(document.querySelectorAll('a')).map(a => a.href)")
                        profile_count = sum(1 for href in links if self._is_profile_like_url(href, urlparse(base_url).netloc))
                        if profile_count > 0:
                            directory_url = target
                            expected_total = self._extract_expected_total(page)
                            break
                except Exception:
                    pass

            if directory_url:
                page.goto(directory_url, timeout=15000, wait_until="domcontentloaded")
                page.wait_for_timeout(2000)

                # Alphabet navigation handling (A-Z)
                try:
                    letter_links = page.evaluate(
                        """
                        () => {
                            const links = Array.from(document.querySelectorAll('a'));
                            return links
                                .filter(a => (a.textContent || '').trim().length === 1 && /[A-Za-z]/.test((a.textContent || '').trim()))
                                .map(a => a.href)
                                .filter(Boolean);
                        }
                        """
                    )
                    if letter_links and len(letter_links) >= 5:
                        self.log(f"  Alphabet navigation detected ({len(letter_links)} letters)")
                        for letter_url in letter_links[:26]:
                            try:
                                page.goto(letter_url, timeout=15000, wait_until="domcontentloaded")
                                page.wait_for_timeout(1500)
                                urls.update(self._extract_profile_urls_from_page(page, base_url))
                            except Exception:
                                pass
                except Exception:
                    pass

                # Exhaust pagination
                metrics = DiscoveryMetrics()  # Temporary for legacy methods
                pagination_urls = self._handle_pagination_stabilized(page, base_url, metrics)
                urls.update(pagination_urls)

                # Exhaust load-more
                loadmore_urls = self._handle_loadmore_stabilized(page, base_url, metrics)
                urls.update(loadmore_urls)

                # Exhaust scroll
                scroll_urls = self._handle_scroll_stabilized(page, base_url, metrics)
                urls.update(scroll_urls)

            page.close()
            context.close()
            browser.close()

        self.log(f"  DOM Exhaustion: {len(urls)} URLs, expected: {expected_total or 'unknown'}")
        return urls, expected_total

    def _strategy_hard_case_fallback(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Last resort for HARD_CASE firms"""
        self.log("  HARD_CASE detected - no standard discovery method available")
        return set(), None

    def _strategy_directory_listing(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Discover attorney profiles via directory listing pages.
        Stub — delegates to dom_exhaustion as a best-effort fallback.
        """
        self.log("  Directory Listing: delegating to DOM exhaustion")
        return self._strategy_dom_exhaustion(base_url)

    def _strategy_alphabet_enumeration(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Enumerate attorney profiles by alphabet (A-Z) nav links.
        Stub — delegates to dom_exhaustion as a best-effort fallback.
        """
        self.log("  Alphabet Enumeration: delegating to DOM exhaustion")
        return self._strategy_dom_exhaustion(base_url)

    def _strategy_xml_sitemap_navigation(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Navigate XML sitemap index to discover attorney profile URLs.
        Delegates to the existing xml_sitemap strategy.
        """
        self.log("  XML Sitemap Navigation: delegating to xml_sitemap")
        return self._strategy_xml_sitemap(base_url)

    def _strategy_kirkland_scroll(self, base_url: str) -> tuple[set[str], int | None]:
        """Strategy: Kirkland & Ellis attorney discovery via per-letter Playwright DOM scroll.

        For each letter A-Z:
          1. Navigate to https://www.kirkland.com/lawyers?letter={letter}
          2. Wait for initial lawyer cards to load
          3. Scroll to bottom repeatedly until no new anchor tags appear for 3 consecutive scrolls
          4. Extract all href attributes containing /lawyers/{letter.lower()}/
          5. Normalize to absolute URLs and deduplicate globally
        """
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            self.log("  [KIRKLAND_SCROLL] Playwright not available — skipping")
            return set(), None

        urls: set[str] = set()
        domain = "www.kirkland.com"

        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=not self.headful)
                self.log(f"  [KIRKLAND_SCROLL] Playwright launch | headful={self.headful} headless={not self.headful}")
                context = browser.new_context(
                    user_agent=(
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/121.0.0.0 Safari/537.36"
                    )
                )
                # Do NOT block stylesheets — Sitecore SPA needs them to render cards
                def _handle_route(route) -> None:
                    if route.request.resource_type in ("image", "font", "media"):
                        route.abort()
                    else:
                        route.continue_()
                context.route("**/*", _handle_route)
                pg = context.new_page()
                letter_counts: dict[str, dict] = {}  # TASK E: per-letter debug artifact

                for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                    letter_url = f"https://www.kirkland.com/lawyers?letter={letter}"
                    letter_lower = letter.lower()
                    path_prefix = f"/lawyers/{letter_lower}/"

                    try:
                        self.rate_limit_manager.wait(domain)
                    except Exception:
                        pass

                    try:
                        pg.goto(letter_url, timeout=30000, wait_until="domcontentloaded")
                        # Wait for initial person-result cards to appear (Vue SPA rendering)
                        try:
                            pg.wait_for_selector(".person-result", timeout=12000)
                        except Exception:
                            pg.wait_for_timeout(3000)  # fallback
                    except Exception as e:
                        self.log(
                            f"  [KIRKLAND_SCROLL] Letter={letter} navigation error: {type(e).__name__}: {e}"
                        )
                        continue

                    # Scroll + click 'See More' until exhausted.
                    # After each action wait for new .person-result cards to appear.
                    consecutive_no_new = 0
                    prev_card_count = 0
                    max_scrolls = 200  # safety cap per letter
                    scroll_count = 0
                    while consecutive_no_new < 3 and scroll_count < max_scrolls:
                        # Click 'See More' button if visible (preferred over scroll)
                        try:
                            see_more = pg.locator(".search-results__load-more")
                            if see_more.is_visible(timeout=500):
                                see_more.click()
                                pg.wait_for_timeout(2000)  # wait for new cards to load
                        except Exception:
                            pass  # button not present — fall through to scroll
                        pg.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                        pg.wait_for_timeout(2000)  # give Vue time to render new cards
                        card_count = pg.eval_on_selector_all(
                            ".person-result",
                            "els => els.length",
                        )
                        if card_count > prev_card_count:
                            consecutive_no_new = 0
                            prev_card_count = card_count
                        else:
                            consecutive_no_new += 1
                        scroll_count += 1
                    # Collect all matching hrefs after scroll is complete
                    # Use precise selector: person-result name links only
                    final_hrefs: list[str] = pg.eval_on_selector_all(
                        ".person-result__name a",
                        "els => els.map(el => el.getAttribute('href'))",
                    )
                    # Fallback: also collect all /lawyers/ links in case selector misses some
                    all_hrefs: list[str] = pg.eval_on_selector_all(
                        "a[href*='/lawyers/']",
                        "els => els.map(el => el.getAttribute('href'))",
                    )
                    combined = set(final_hrefs + all_hrefs)
                    letter_urls: set[str] = set()
                    for href in combined:
                        if href and path_prefix in href:
                            full_url = (
                                href if href.startswith("http")
                                else f"https://www.kirkland.com{href}"
                            )
                            letter_urls.add(full_url)
                    before_letter_count = len(urls)
                    urls.update(letter_urls)
                    added_this_letter = len(urls) - before_letter_count
                    letter_counts[letter] = {
                        'profiles_added': added_this_letter,
                        'sample_urls': sorted(letter_urls)[:5],
                    }
                    self.log(
                        f"  [KIRKLAND_SCROLL] Letter={letter} profiles_added={added_this_letter} total={len(urls)}"
                    )
                pg.close()
                context.close()
                browser.close()
                # TASK E: write per-letter debug artifact
                try:
                    debug_dir = Path('debug_reports')
                    debug_dir.mkdir(parents=True, exist_ok=True)
                    firm_slug = re.sub(r'[^\w]+', '_', self.firm or 'unknown').strip('_')
                    artifact_path = debug_dir / f'{firm_slug}_letter_counts.json'
                    with artifact_path.open('w', encoding='utf-8') as _fh:
                        json.dump(letter_counts, _fh, indent=2)
                    self.log(f"  [KIRKLAND_SCROLL] Per-letter debug artifact: {artifact_path}")
                except Exception as _write_err:
                    self.log(f"  [KIRKLAND_SCROLL] Could not write letter_counts artifact: {_write_err}")
                context.close()
                browser.close()
        except Exception as e:
            self.log(f"  [KIRKLAND_SCROLL] Fatal error: {type(e).__name__}: {e}")
            if urls:
                self.log(f"  [KIRKLAND_SCROLL] Returning {len(urls)} URLs collected before error")
                return urls, None
            return set(), None

        if not urls:
            self.log("  [KIRKLAND_SCROLL] No attorney URLs discovered")
            return set(), None

        self.log(f"  [KIRKLAND_SCROLL] Total profile URLs discovered: {len(urls)}")
        return urls, None

    # ========================================================================
    # COVERAGE VALIDATION
    # ========================================================================

    def _evaluate_coverage(self, report: CoverageReport, current_urls: int) -> bool:
        """Check if coverage is sufficient to stop strategy execution

        Returns: True if coverage is sufficient, False if need more strategies
        """
        if not report.expected_total:
            # No expected total, use minimum threshold
            return current_urls >= MIN_ATTORNEYS_THRESHOLD

        ratio = current_urls / report.expected_total
        return ratio >= COVERAGE_THRESHOLD

    def _determine_final_status(self, report: CoverageReport) -> str:
        """Determine final coverage status"""
        if report.discovered_urls == 0:
            report.notes.append("No attorneys discovered by any strategy")
            return STATUS_UNKNOWN_PATTERN

        if not report.expected_total:
            if report.discovered_urls >= MIN_ATTORNEYS_THRESHOLD:
                report.notes.append(f"Discovered {report.discovered_urls} attorneys (no expected total available)")
                return STATUS_SUCCESS
            else:
                report.notes.append(f"Only {report.discovered_urls} attorneys found (below minimum threshold)")
                return STATUS_PARTIAL

        ratio = report.discovered_urls / report.expected_total
        if ratio >= COVERAGE_THRESHOLD:
            report.notes.append(f"Coverage: {ratio*100:.1f}% ({report.discovered_urls}/{report.expected_total})")
            return STATUS_SUCCESS
        elif ratio >= 0.5:
            report.notes.append(f"Partial coverage: {ratio*100:.1f}% ({report.discovered_urls}/{report.expected_total})")
            return STATUS_PARTIAL
        else:
            report.notes.append(f"Low coverage: {ratio*100:.1f}% ({report.discovered_urls}/{report.expected_total})")
            return STATUS_UNKNOWN_PATTERN

    def _determine_discovery_status(self, report: CoverageReport, attorneys: list) -> tuple[str, str | None]:
        """Determine discovery status for external directory fallback

        Returns:
            (discovery_status, failure_reason)
            - discovery_status: SUCCESS | BLOCKED | DISCOVERY_INCOMPLETE
            - failure_reason: Specific failure reason if applicable
        """
        # Case 1: BLOCKED - Majority of attorneys have bot protection
        if attorneys:
            blocked_count = 0
            for att in attorneys:
                if hasattr(att, 'diagnostics'):
                    diag = att.diagnostics
                    if diag.get('http_403') or diag.get('bot_protection') or diag.get('blocked_403'):
                        blocked_count += 1

            blocked_ratio = blocked_count / len(attorneys)
            if blocked_ratio > 0.7:  # >70% blocked
                return DISCOVERY_BLOCKED, FAILURE_ALL_BLOCKED

        # Case 2: DISCOVERY_INCOMPLETE - Zero attorneys discovered
        if report.discovered_urls == 0:
            # Determine specific failure reason from strategies attempted
            if not report.strategies_attempted:
                return DISCOVERY_INCOMPLETE, FAILURE_EXTRACTION_ERROR

            # Check if sitemap strategies were tried
            sitemap_strategies = [s for s in report.strategies_attempted if 'sitemap' in s.lower()]
            if sitemap_strategies:
                # Sitemap found but yielded no profiles
                return DISCOVERY_INCOMPLETE, FAILURE_SITEMAP_EMPTY
            else:
                # No sitemap found at all
                return DISCOVERY_INCOMPLETE, FAILURE_SITEMAP_NOT_FOUND

        # Case 3: DISCOVERY_INCOMPLETE - Expected attorneys but got very few
        if report.expected_total and report.expected_total > 0:
            ratio = report.discovered_urls / report.expected_total
            if ratio < 0.1:  # Less than 10% coverage
                return DISCOVERY_INCOMPLETE, FAILURE_ZERO_PROFILES

        # Case 4: SUCCESS - Sufficient attorneys discovered
        return DISCOVERY_SUCCESS, None

    # ========================================================================
    # PERSISTENCE
    # ========================================================================

    def _save_firm_profile(self, profile: FirmProfile) -> None:
        """Save FirmProfile to JSON"""
        safe_name = re.sub(r'[^\w\s-]', '', profile.firm).strip().replace(' ', '_')
        path = self.debug_dir / f"{safe_name}_profile.json"

        with open(path, 'w', encoding='utf-8') as f:
            json.dump(profile.to_dict(), f, indent=2)

        self.log(f"  Profile saved: {path}")

    def _save_coverage_report(self, report: CoverageReport) -> None:
        """Save CoverageReport to JSON"""
        safe_name = re.sub(r'[^\w\s-]', '', report.firm).strip().replace(' ', '_')
        path = self.debug_dir / f"{safe_name}_coverage.json"

        with open(path, 'w', encoding='utf-8') as f:
            json.dump(report.to_dict(), f, indent=2)

        self.log(f"  Coverage report saved: {path}")

    def _try_fast_path(self, base_url: str, metrics: DiscoveryMetrics) -> list[dict]:
        """Fast path: Sitemap + parallel requests"""
        attorneys = []

        try:
            # Try sitemap (with robots.txt check)
            profile_urls = self._extract_profile_urls_from_sitemap(base_url)
            if profile_urls:
                self.log(f"  Sitemap: found {len(profile_urls)} URLs")

                # Parallel metadata fetch
                with ThreadPoolExecutor(max_workers=8) as executor:
                    futures = {
                        executor.submit(self._fetch_profile_metadata, url): url
                        for url in profile_urls[
                            : self.limit * 3 if self.limit > 0 else len(profile_urls)
                        ]
                    }
                    for future in as_completed(futures):
                        try:
                            att = future.result()
                            if att and att.get("name"):
                                attorneys.append(att)
                        except Exception:
                            pass
        except Exception as e:
            self.log(f"  Fast path error: {e}")
            metrics.failure_notes.append(f"Fast path error: {e}")

        return attorneys

    def _try_playwright_path(
        self, base_url: str, metrics: DiscoveryMetrics
    ) -> tuple[list[dict], DiscoveryMetrics]:
        """Playwright fallback with stabilization-based discovery"""
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            self.log("  Playwright not available, skipping")
            metrics.failure_notes.append("Playwright not installed")
            return [], metrics

        attorneys = []
        all_profile_urls = set()
        captured_api_data = []

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=not self.headful)

            context = browser.new_context()

            # Block heavy resources
            def handle_route(route):
                if route.request.resource_type in [
                    "image",
                    "font",
                    "media",
                    "stylesheet",
                ]:
                    route.abort()
                else:
                    route.continue_()

            context.route("**/*", handle_route)

            page = context.new_page()

            # Intercept API responses
            def handle_response(response):
                try:
                    url_lower = response.url.lower()

                    if any(
                        kw in url_lower
                        for kw in [
                            "api",
                            "graphql",
                            "search",
                            "people",
                            "attorneys",
                            "lawyers",
                            "professionals",
                        ]
                    ):
                        if response.ok and "json" in response.headers.get(
                            "content-type", ""
                        ).lower():
                            try:
                                data = response.json()
                                records = self._extract_attorneys_from_json(
                                    data, base_url
                                )
                                if records:  # Only log if records found
                                    self.log(f"  API detected: {response.url}, {len(records)} records")
                                captured_api_data.append(
                                    {
                                        "url": response.url,
                                        "records": records,
                                        "json": data,  # Store full JSON for structure analysis
                                    }
                                )
                            except Exception:
                                pass
                except Exception:
                    pass

            page.on("response", handle_response)

            # Find best directory URL
            dir_url, expected_total, has_api = self._find_best_directory_url(page, base_url, captured_api_data)
            if not dir_url:
                self.log("  Could not find directory page")
                metrics.failure_notes.append("No directory page found")
                page.close()
                context.close()
                browser.close()
                return [], metrics

            metrics.directory_url = dir_url
            metrics.expected_total = expected_total
            self.log(f"  Directory: {dir_url}")
            if expected_total > 0:
                self.log(f"  Expected total: {expected_total}")
            if has_api:
                self.log("  Detected API-based directory (no DOM links required)")

            # Navigate to directory
            page.goto(dir_url, timeout=15000, wait_until="domcontentloaded")
            page.wait_for_timeout(2000)

            # DISCOVERY PHASE 1: Try API enumeration first (fastest + highest coverage)
            api_urls = self._enumerate_api_if_available(
                captured_api_data, base_url, metrics
            )
            all_profile_urls.update(api_urls)

            # DISCOVERY PHASE 2: Exhaust pagination/load-more/scroll (baseline, fast)
            if metrics.discovered_by_api == 0:  # Only if API didn't work
                baseline_urls = self._exhaust_directory_content(page, base_url, metrics)
                all_profile_urls.update(baseline_urls)

            # DISCOVERY PHASE 3: Enumerate filters (slower, only if needed)
            if (
                metrics.discovered_by_api == 0
                and metrics.expected_total > 0
                and len(all_profile_urls) < metrics.expected_total * 0.5
            ):  # Only if we're missing >50% of expected attorneys
                self.log(f"  Low coverage ({len(all_profile_urls)}/{metrics.expected_total}), trying filter enumeration...")
                filter_urls = self._enumerate_filters(page, base_url, metrics)
                all_profile_urls.update(filter_urls)

            # Save debug artifacts if low coverage
            if metrics.expected_total > 0:
                ratio = len(all_profile_urls) / metrics.expected_total
                if ratio < 0.98 or len(all_profile_urls) < MIN_ATTORNEYS_THRESHOLD:
                    self._save_debug_artifacts(page, base_url, metrics)

            page.close()
            context.close()
            browser.close()

        # Enrich URLs with metadata
        self.log(f"  Enriching {len(all_profile_urls)} profile URLs...")
        # Legacy path still using old approach - extract firm name from base_url for now
        domain = urlparse(base_url).netloc
        attorneys = self._enrich_profile_urls(list(all_profile_urls), base_url, domain)

        # Record CoverageMetrics for this legacy Playwright path
        _cov = self.coverage_engine.compute(
            firm=base_url,
            profiles=attorneys,
            discovered_urls=len(all_profile_urls),
            blocked_count=0,
            expected_total=metrics.expected_total or None,
        )
        self.all_coverage_metrics.append(_cov)

        return attorneys, metrics

    def _find_best_directory_url(self, page, base_url: str, captured_api_data: list[dict]) -> tuple[str, int, bool]:
        """Find directory page with most profile links OR API data

        Returns: (best_url, expected_total, has_api)
        """
        probe_paths = [
            "/people",
            "/professionals",
            "/attorneys",
            "/lawyers",
            "/our-people",
            "/our-team",
            "/team",
            "/find-a-professional",
            "/find-a-lawyer",
            "/attorney-search",
        ]

        best_url = ""
        best_score = 0
        expected_total = 0
        has_api = False

        for path in probe_paths:
            # Clear API data for this probe
            captured_api_data.clear()

            try:
                target = urljoin(base_url, path)
                self.log(f"  Probing: {target}")
                response = page.goto(
                    target, timeout=20000, wait_until="networkidle"
                )
                if not response or not response.ok:
                    self.log(f"    Status: {response.status if response else 'No response'}")
                    continue

                page.wait_for_timeout(2000)  # Extra wait for SPAs

                # Count profile-like links
                links = page.evaluate(
                    """
                    () => {
                        const anchors = Array.from(document.querySelectorAll('a'));
                        return anchors.map(a => a.href).filter(Boolean);
                    }
                    """
                )

                profile_count = sum(
                    1
                    for href in links
                    if self._is_profile_like_url(href, urlparse(base_url).netloc)
                )

                self.log(f"    Found {len(links)} total links, {profile_count} profile-like")

                # Debug: Show sample URLs
                if len(links) > 0 and profile_count == 0:
                    self.log(f"    Sample URLs:")
                    for href in links[:5]:
                        self.log(f"      {href}")

                # Check for total count indicators
                total = self._extract_expected_total(page)

                # Check if API has records for this page
                api_record_count = 0
                for api_data in captured_api_data:
                    if api_data.get("records"):
                        api_record_count += len(api_data["records"])

                # Calculate score: prioritize DOM links, but accept API-only
                current_score = profile_count
                current_has_api = False

                if api_record_count > 0:
                    self.log(f"    API detected: {api_record_count} records")
                    current_has_api = True
                    # If no DOM links but API has data, use API count as score
                    if profile_count == 0:
                        current_score = api_record_count
                        self.log(f"    Using API-based directory (DOM links = 0)")

                # Extract expected total from API if available
                if current_has_api and total == 0:
                    for api_data in captured_api_data:
                        json_data = api_data.get("json")
                        if json_data:
                            api_total = self._extract_total_from_json(json_data)
                            if api_total > 0:
                                total = api_total
                                break

                if current_score > best_score:
                    best_score = current_score
                    best_url = target
                    expected_total = total
                    has_api = current_has_api
                    self.log(f"    [BEST] New best: {current_score} {'API records' if current_has_api and profile_count == 0 else 'profiles'}")

            except Exception as e:
                self.log(f"    Error: {e}")

        if best_url:
            self.log(f"  Best directory: {best_url} ({best_score} {'API records' if has_api and best_score > 0 else 'profiles'})")
        return best_url, expected_total, has_api

    def _extract_expected_total(self, page) -> int:
        """Extract expected attorney count from directory page"""
        try:
            text = page.inner_text("body")

            # Patterns: "of 1234", "Showing 1-50 of 1234", "1234 Professionals", etc.
            patterns = [
                r"of\s+(\d{2,5})\s+(?:results|professionals|attorneys|lawyers|people)",
                r"showing\s+\d+\s*[-–]\s*\d+\s+of\s+(\d{2,5})",
                r"(\d{2,5})\s+(?:professionals|attorneys|lawyers|people|results)",
                r"total[:\s]+(\d{2,5})",
                r"(\d{3,5})\s+results",
            ]

            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    count = int(match.group(1))
                    if 10 <= count <= 10000:  # Sanity check
                        self.log(f"  Detected total count: {count}")
                        return count
        except Exception:
            pass

        return 0

    def _enumerate_api_if_available(
        self, captured_api_data: list[dict], base_url: str, metrics: DiscoveryMetrics
    ) -> set[str]:
        """If API detected, enumerate all pages via JSON structure analysis"""
        if not captured_api_data:
            return set()

        self.log("  Detected API, attempting full enumeration...")
        all_urls = set()
        expected_totals = []

        for api_data in captured_api_data:
            api_url = api_data["url"]
            initial_records = api_data["records"]
            initial_json = api_data.get("json")  # Store full JSON response

            if not initial_records:
                continue

            # Add initial records
            for rec in initial_records:
                if rec.get("url"):
                    all_urls.add(rec["url"])

            # Extract expected total from API response
            if initial_json:
                total = self._extract_total_from_json(initial_json)
                if total > 0:
                    expected_totals.append(total)
                    self.log(f"  API reported total: {total}")

            # Try to paginate API (JSON structure-based)
            paginated_urls = self._paginate_api_smart(api_url, initial_json, base_url)
            all_urls.update(paginated_urls)

        metrics.discovered_by_api = len(all_urls)
        if metrics.discovered_by_api > 0:
            self.log(f"  API enumeration: {metrics.discovered_by_api} URLs")

        # Update expected_total from API if available
        if expected_totals and metrics.expected_total == 0:
            metrics.expected_total = max(expected_totals)

        return all_urls

    def _paginate_api_endpoint(self, api_url: str, base_url: str) -> set[str]:
        """Attempt to paginate API endpoint (legacy URL-based method)"""
        all_urls = set()

        parsed = urlparse(api_url)
        query_params = parse_qs(parsed.query)

        # Common pagination params
        page_param = None
        for param in ["page", "pageNumber", "pageNum", "p"]:
            if param in query_params:
                page_param = param
                break

        offset_param = None
        for param in ["offset", "start", "skip", "from"]:
            if param in query_params:
                offset_param = param
                break

        limit_param = None
        for param in ["limit", "size", "pageSize", "count", "take"]:
            if param in query_params:
                limit_param = param
                break

        if not page_param and not offset_param:
            # Can't paginate
            return all_urls

        # Try paginating
        page = 1
        offset = 0
        limit = int(query_params.get(limit_param, ["50"])[0]) if limit_param else 50
        stabilization_counter = 0

        domain = urlparse(base_url).netloc

        for attempt in range(200):  # Safety cap
            if stabilization_counter >= self.stabilization:
                break

            # Build paginated URL
            new_params = query_params.copy()
            if page_param:
                new_params[page_param] = [str(page)]
                page += 1
            elif offset_param:
                new_params[offset_param] = [str(offset)]
                offset += limit

            new_query = urlencode(new_params, doseq=True)
            paginated_url = urlunparse(
                (parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment)
            )

            try:
                try:
                    self.rate_limit_manager.wait(domain)
                except RateLimitBlockedError:
                    break
                resp = self.session.get(paginated_url, timeout=REQUEST_TIMEOUT)
                if resp.status_code != 200:
                    break

                data = resp.json()
                records = self._extract_attorneys_from_json(data, base_url)

                if not records:
                    stabilization_counter += 1
                else:
                    stabilization_counter = 0
                    for rec in records:
                        if rec.get("url"):
                            all_urls.add(rec["url"])

            except Exception:
                break

        return all_urls

    def _paginate_api_smart(self, api_url: str, initial_json: dict | None, base_url: str) -> set[str]:
        """Smart API pagination based on JSON structure analysis"""
        all_urls = set()

        if not initial_json:
            return all_urls

        domain = urlparse(base_url).netloc

        # Strategy 1: Detect pagination metadata in JSON
        pagination_info = self._detect_pagination_structure(initial_json)

        if pagination_info:
            self.log(f"  Detected pagination: {pagination_info}")

            current_page = pagination_info.get("current_page", 1)
            total_pages = pagination_info.get("total_pages", 0)
            page_size = pagination_info.get("page_size", 50)
            total_items = pagination_info.get("total_items", 0)

            # Calculate total pages if not provided
            if total_pages == 0 and total_items > 0 and page_size > 0:
                total_pages = (total_items + page_size - 1) // page_size

            # Cap at reasonable limit
            total_pages = min(total_pages, 100)

            # Try URL-based pagination first (fast)
            parsed = urlparse(api_url)
            query_params = parse_qs(parsed.query)

            # Detect page parameter
            page_param = None
            for param in ["page", "pageNumber", "pageNum", "p", "offset", "skip"]:
                if param in query_params:
                    page_param = param
                    break

            if page_param and total_pages > 1:
                self.log(f"  Paginating {total_pages} pages via URL parameter '{page_param}'...")

                for page in range(2, total_pages + 1):  # Start from page 2 (already have page 1)
                    try:
                        new_params = query_params.copy()

                        # Handle offset vs page
                        if page_param in ["offset", "skip"]:
                            new_params[page_param] = [str((page - 1) * page_size)]
                        else:
                            new_params[page_param] = [str(page)]

                        new_query = urlencode(new_params, doseq=True)
                        paginated_url = urlunparse(
                            (parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment)
                        )

                        try:
                            self.rate_limit_manager.wait(domain)
                        except RateLimitBlockedError:
                            break
                        resp = self.session.get(paginated_url, timeout=REQUEST_TIMEOUT)

                        if resp.status_code != 200:
                            break

                        data = resp.json()
                        records = self._extract_attorneys_from_json(data, base_url)

                        for rec in records:
                            if rec.get("url"):
                                all_urls.add(rec["url"])

                        if not records:
                            break

                    except Exception as e:
                        self.log(f"  Page {page} failed: {e}")
                        break

        # Strategy 2: POST/GraphQL pagination (if GET failed)
        if not all_urls and self._looks_like_graphql(api_url):
            self.log("  Attempting GraphQL pagination...")
            graphql_urls = self._paginate_graphql(api_url, initial_json, base_url)
            all_urls.update(graphql_urls)

        return all_urls

    def _detect_pagination_structure(self, json_data: dict) -> dict | None:
        """Detect pagination metadata in JSON response"""
        if not isinstance(json_data, dict):
            return None

        pagination_keys = {
            "current": ["page", "currentPage", "current_page", "pageNumber"],
            "total_pages": ["totalPages", "total_pages", "pageCount", "pages"],
            "total_items": ["total", "totalCount", "total_count", "totalItems", "count"],
            "page_size": ["pageSize", "page_size", "limit", "size", "perPage"],
        }

        result = {}

        def search_recursive(obj, depth=0):
            if depth > 3 or not isinstance(obj, dict):
                return

            for key, value in obj.items():
                key_lower = key.lower()

                # Check for pagination keys
                if not result.get("current_page"):
                    for candidate in pagination_keys["current"]:
                        if candidate.lower() in key_lower and isinstance(value, (int, str)):
                            try:
                                result["current_page"] = int(value)
                            except:
                                pass

                if not result.get("total_pages"):
                    for candidate in pagination_keys["total_pages"]:
                        if candidate.lower() in key_lower and isinstance(value, (int, str)):
                            try:
                                result["total_pages"] = int(value)
                            except:
                                pass

                if not result.get("total_items"):
                    for candidate in pagination_keys["total_items"]:
                        if candidate.lower() == key_lower and isinstance(value, (int, str)):
                            try:
                                val = int(value)
                                # Sanity check: should be reasonable attorney count
                                if 10 <= val <= 50000:
                                    result["total_items"] = val
                            except:
                                pass

                if not result.get("page_size"):
                    for candidate in pagination_keys["page_size"]:
                        if candidate.lower() in key_lower and isinstance(value, (int, str)):
                            try:
                                result["page_size"] = int(value)
                            except:
                                pass

                # Recurse into nested objects
                if isinstance(value, dict):
                    search_recursive(value, depth + 1)

        search_recursive(json_data)

        return result if result else None

    def _extract_total_from_json(self, json_data: dict) -> int:
        """Extract total count from JSON response"""
        pagination_info = self._detect_pagination_structure(json_data)
        if pagination_info:
            total = pagination_info.get("total_items", 0)
            if total > 0:
                return total
        return 0

    def _looks_like_graphql(self, url: str) -> bool:
        """Check if URL looks like GraphQL endpoint"""
        return "graphql" in url.lower()

    def _paginate_graphql(self, api_url: str, initial_json: dict, base_url: str) -> set[str]:
        """Attempt GraphQL pagination (stub for future implementation)"""
        # TODO: Implement GraphQL pagination
        # This would require:
        # 1. Detecting GraphQL query structure
        # 2. Extracting variables
        # 3. Modifying pagination variables
        # 4. POST requests with modified query
        return set()


    def _enumerate_filters(
        self, page, base_url: str, metrics: DiscoveryMetrics
    ) -> set[str]:
        """Enumerate single-facet filters (Practice, Office, etc.)"""
        self.log("  Attempting filter enumeration...")
        all_urls = set()

        try:
            # Detect filter UI
            filter_groups = page.evaluate(
                """
                () => {
                    // Look for common filter patterns
                    const filters = [];

                    // Pattern 1: Select dropdowns
                    const selects = Array.from(document.querySelectorAll('select'));
                    for (const select of selects) {
                        const label = select.previousElementSibling?.innerText || select.id || '';
                        if (/practice|office|location|industry/i.test(label + select.name + select.id)) {
                            const options = Array.from(select.options)
                                .filter(opt => opt.value && opt.value !== '')
                                .map(opt => ({value: opt.value, text: opt.text}));
                            if (options.length > 0) {
                                filters.push({
                                    type: 'select',
                                    element: select,
                                    label: label,
                                    options: options
                                });
                            }
                        }
                    }

                    // Pattern 2: Clickable filter buttons/links
                    const filterContainers = Array.from(document.querySelectorAll('[class*="filter"], [class*="facet"]'));
                    for (const container of filterContainers) {
                        const label = container.querySelector('label, h3, h4, .label')?.innerText || '';
                        if (/practice|office|location/i.test(label)) {
                            const clickables = Array.from(container.querySelectorAll('a, button, [role="button"]'))
                                .filter(el => el.innerText.trim())
                                .map(el => ({text: el.innerText.trim(), selector: el.tagName}));
                            if (clickables.length > 1) {
                                filters.push({
                                    type: 'clickable',
                                    label: label,
                                    items: clickables
                                });
                            }
                        }
                    }

                    return filters;
                }
                """
            )

            if not filter_groups:
                self.log("  No filters detected")
                return all_urls

            self.log(f"  Found {len(filter_groups)} filter groups")

            # Enumerate each filter group
            for group_idx, group in enumerate(filter_groups[:1]):  # Limit to 1 group for speed
                filter_type = group.get("type")

                if filter_type == "select":
                    # Handle select dropdowns
                    options = group.get("options", [])
                    for idx, opt in enumerate(options[:10]):  # Cap at 10 values per filter
                        try:
                            self.log(f"  Filter {group_idx+1}/{len(filter_groups[:1])}, option {idx+1}/{min(10, len(options))}: {opt['text'][:40]}")

                            # Select option
                            page.select_option(
                                f"select:has-text('{group['label']}')", opt["value"]
                            )
                            page.wait_for_timeout(800)

                            # Exhaust content with quick extraction (no deep exhaustion)
                            urls = self._extract_profile_urls_from_page(page, base_url)
                            all_urls.update(urls)

                            # Reset (reload page)
                            page.reload()
                            page.wait_for_timeout(800)
                        except Exception as e:
                            self.log(f"  Filter option error: {e}")

                elif filter_type == "clickable":
                    # Handle clickable filters
                    items = group.get("items", [])
                    for idx, item in enumerate(items[:10]):  # Cap at 10
                        try:
                            self.log(f"  Filter {group_idx+1}/{len(filter_groups[:1])}, item {idx+1}/{min(10, len(items))}: {item['text'][:40]}")

                            # Click filter
                            page.click(f"text={item['text']}")
                            page.wait_for_timeout(800)

                            # Quick extraction (no deep exhaustion)
                            urls = self._extract_profile_urls_from_page(page, base_url)
                            all_urls.update(urls)

                            # Reset
                            page.reload()
                            page.wait_for_timeout(800)
                        except Exception as e:
                            self.log(f"  Filter item error: {e}")

            metrics.discovered_by_filters = len(all_urls)
            if metrics.discovered_by_filters > 0:
                self.log(f"  Filter enumeration: {metrics.discovered_by_filters} URLs")

        except Exception as e:
            self.log(f"  Filter enumeration error: {e}")
            metrics.failure_notes.append(f"Filter enumeration error: {e}")

        return all_urls

    def _exhaust_directory_content(
        self, page, base_url: str, metrics: DiscoveryMetrics
    ) -> set[str]:
        """Exhaust pagination, load-more, and infinite scroll with stabilization"""
        all_urls = set()
        start_time = time.time()

        # Phase 1: Pagination (highest priority)
        pagination_urls = self._handle_pagination_stabilized(page, base_url, metrics)
        all_urls.update(pagination_urls)

        # Phase 2: Load-more buttons
        if time.time() - start_time < self.max_scroll_seconds:
            loadmore_urls = self._handle_loadmore_stabilized(page, base_url, metrics)
            all_urls.update(loadmore_urls)

        # Phase 3: Infinite scroll
        if time.time() - start_time < self.max_scroll_seconds:
            scroll_urls = self._handle_scroll_stabilized(page, base_url, metrics)
            all_urls.update(scroll_urls)

        # Phase 4: Extract all visible links
        dom_urls = self._extract_profile_urls_from_page(page, base_url)
        all_urls.update(dom_urls)
        metrics.discovered_by_dom = len(dom_urls)

        return all_urls

    def _handle_pagination_stabilized(
        self, page, base_url: str, metrics: DiscoveryMetrics
    ) -> set[str]:
        """Handle pagination with stabilization"""
        all_urls = set()
        stabilization_counter = 0
        page_num = 1
        max_pages = 50  # Safety cap

        while stabilization_counter < self.stabilization and page_num <= max_pages:
            try:
                prev_count = len(all_urls)

                # Extract URLs from current page
                urls = self._extract_profile_urls_from_page(page, base_url)
                all_urls.update(urls)

                # Check for next button
                next_button = None
                for selector in [
                    'a[rel="next"]',
                    'a:has-text("Next")',
                    'button:has-text("Next")',
                    '.pagination a:has-text("›")',
                    '.pagination a:has-text("»")',
                    f'a:has-text("{page_num + 1}")',
                ]:
                    try:
                        if page.locator(selector).count() > 0:
                            next_button = selector
                            break
                    except Exception:
                        pass

                if not next_button:
                    break

                # Click next
                page.click(next_button, timeout=5000)
                page.wait_for_timeout(1500)
                page_num += 1

                # Check stabilization
                if len(all_urls) == prev_count:
                    stabilization_counter += 1
                else:
                    stabilization_counter = 0

            except Exception:
                break

        metrics.discovered_by_pagination = len(all_urls)
        if metrics.discovered_by_pagination > 0:
            self.log(f"  Pagination: {metrics.discovered_by_pagination} URLs ({page_num} pages)")

        return all_urls

    def _handle_loadmore_stabilized(
        self, page, base_url: str, metrics: DiscoveryMetrics
    ) -> set[str]:
        """Handle load-more buttons with stabilization"""
        all_urls = set()
        stabilization_counter = 0
        clicks = 0
        max_clicks = 30  # Safety cap

        while stabilization_counter < self.stabilization and clicks < max_clicks:
            try:
                prev_count = len(all_urls)

                # Extract URLs
                urls = self._extract_profile_urls_from_page(page, base_url)
                all_urls.update(urls)

                # Find load-more button
                load_more = None
                for selector in [
                    'button:has-text("Load More")',
                    'button:has-text("Show More")',
                    'a:has-text("Load More")',
                    'a:has-text("Show More")',
                    '[class*="load-more"]',
                    '[class*="show-more"]',
                ]:
                    try:
                        btn = page.locator(selector)
                        if btn.count() > 0 and btn.first.is_visible():
                            load_more = selector
                            break
                    except Exception:
                        pass

                if not load_more:
                    break

                # Click
                page.click(load_more, timeout=5000)
                page.wait_for_timeout(1500)
                clicks += 1

                # Check stabilization
                if len(all_urls) == prev_count:
                    stabilization_counter += 1
                else:
                    stabilization_counter = 0

            except Exception:
                break

        metrics.discovered_by_loadmore = len(all_urls)
        if metrics.discovered_by_loadmore > 0:
            self.log(
                f"  Load-more: {metrics.discovered_by_loadmore} URLs ({clicks} clicks)"
            )

        return all_urls

    def _handle_scroll_stabilized(
        self, page, base_url: str, metrics: DiscoveryMetrics
    ) -> set[str]:
        """Handle infinite scroll with stabilization"""
        all_urls = set()
        stabilization_counter = 0
        scrolls = 0
        max_scrolls = 20  # Safety cap

        while stabilization_counter < self.stabilization and scrolls < max_scrolls:
            try:
                prev_count = len(all_urls)

                # Extract URLs
                urls = self._extract_profile_urls_from_page(page, base_url)
                all_urls.update(urls)

                # Scroll to bottom
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                page.wait_for_timeout(1000)
                scrolls += 1

                # Check stabilization
                if len(all_urls) == prev_count:
                    stabilization_counter += 1
                else:
                    stabilization_counter = 0

            except Exception:
                break

        metrics.discovered_by_scroll = len(all_urls)
        if metrics.discovered_by_scroll > 0:
            self.log(f"  Scroll: {metrics.discovered_by_scroll} URLs ({scrolls} scrolls)")

        return all_urls

    def _extract_profile_urls_from_page(self, page, base_url: str) -> set[str]:
        """Extract all profile-like URLs from current page"""
        try:
            links = page.evaluate(
                """
                () => {
                    const anchors = Array.from(document.querySelectorAll('a'));
                    return anchors.map(a => a.href).filter(Boolean);
                }
                """
            )

            domain = urlparse(base_url).netloc
            profile_urls = set()
            all_candidate_count = 0

            for href in links:
                all_candidate_count += 1
                if self._is_profile_like_url(href, domain):
                    profile_urls.add(href)

            if all_candidate_count > 0 and len(profile_urls) == 0:
                # Log first few URLs to help debug
                self.log(f"  DEBUG: Found {all_candidate_count} total links, 0 profile URLs. Sample links:")
                for href in list(links)[:5]:
                    self.log(f"    {href}")

            return profile_urls
        except Exception as e:
            self.log(f"  Error extracting URLs: {e}")
            return set()

    def _is_profile_like_url(self, url: str, expected_domain: str) -> bool:
        """Check if URL looks like attorney profile"""
        try:
            parsed = urlparse(url)

            # Must be same domain
            if expected_domain not in parsed.netloc:
                return False

            # Must not be mailto/tel
            if parsed.scheme in ["mailto", "tel"]:
                return False

            url_lower = url.lower()

            # Must contain profile keywords or match /professionals/[a-z]/[slug]
            keywords = [
                "/lawyer",
                "/attorney",
                "/people/",
                "/professional",
                "/bio/",
                "/profile/",
                "/team/",
                "/our-people/",
            ]
            if not any(kw in url_lower for kw in keywords):
                if not re.search(r"/professionals/[a-z]/[a-z0-9-]+", url_lower):
                    return False

            # Filter out obvious junk
            junk_keywords = [
                "terms",
                "privacy",
                "advertising",
                "disclaimer",
                "cookie",
                "sitemap",
                "/search",
                "login",
                "subscribe",
                "career",
                "alumni",
            ]
            if any(junk in url_lower for junk in junk_keywords):
                return False

            # Must have reasonable path depth (at least something after the directory)
            path_parts = [p for p in parsed.path.split("/") if p]
            if len(path_parts) < 2:
                return False

            return True
        except Exception:
            return False

    def _enrich_profile_urls(self, urls: list[str], base_url: str, firm_name: str) -> list[dict]:
        """Fetch metadata for all profile URLs using Multi-Mode Extraction

        CRITICAL: NEVER drop attorneys due to enrichment failure.
        Discovery success is defined by URL count, NOT parsing success.

        Uses 3-mode extraction cascade:
        - Mode 1: requests HTML (fast)
        - Mode 2: Playwright rendered DOM (JS sites)
        - Mode 3: API interception (JSON sources)
        """
        from multi_mode_extractor import MultiModeExtractor

        def _run_batch(batch_urls: list[str], extractor: MultiModeExtractor, force_playwright: bool) -> list:
            results = []
            with ThreadPoolExecutor(max_workers=self.workers if not force_playwright else min(self.workers, 4)) as executor:
                futures = {
                    executor.submit(
                        extractor.extract_profile,
                        firm_name,
                        url,
                        lambda d: self.rate_limit_manager.wait(d),
                        force_playwright
                    ): url
                    for url in batch_urls
                }
                for future in as_completed(futures):
                    try:
                        results.append(future.result())
                    except Exception as e:
                        from attorney_extractor import AttorneyProfile
                        from profile_quality_gate import ReasonCode
                        url = futures[future]
                        profile = AttorneyProfile(firm=firm_name, profile_url=url)
                        profile.extraction_status = "FAILED"
                        profile.diagnostics["exception"] = str(type(e).__name__)
                        profile.diagnostics["full_name_reason"] = ReasonCode.EXCEPTION
                        profile.missing_fields = ["full_name", "title", "offices", "department", "practice_areas", "industries", "bar_admissions", "education"]
                        results.append(profile)
            return results

        extractor_requests = MultiModeExtractor(
            session=self.session,
            timeout=REQUEST_TIMEOUT,
            debug_dir=self.debug_dir,
            enable_playwright=False
        )
        extractor_playwright = MultiModeExtractor(
            session=self.session,
            timeout=REQUEST_TIMEOUT,
            debug_dir=self.debug_dir,
            enable_playwright=True
        )

        attorneys = []
        enriched_success = 0
        enriched_partial = 0
        enriched_failed = 0

        # Stage 1: Requests (or Playwright-only if enforced)
        if self.enrichment_mode == "PLAYWRIGHT_ONLY":
            self.log("  Enrichment mode: PLAYWRIGHT_ONLY")
            attorneys = _run_batch(urls, extractor_playwright, True)
        else:
            self.log("  Enrichment mode: REQUESTS")
            attorneys = _run_batch(urls, extractor_requests, False)

            # If >50% blocked in requests stage, switch firm to PLAYWRIGHT_ONLY
            blocked_count = 0
            for att in attorneys:
                diag = getattr(att, "diagnostics", {})
                if diag.get("http_403") or diag.get("bot_protection") or diag.get("blocked_403"):
                    blocked_count += 1
            if attorneys:
                blocked_ratio = blocked_count / len(attorneys)
                if blocked_ratio > 0.5:
                    self.enrichment_mode = "PLAYWRIGHT_ONLY"
                    self.log(f"  Switching enrichment mode to PLAYWRIGHT_ONLY (blocked ratio {blocked_ratio:.2f})")
                    attorneys = _run_batch(urls, extractor_playwright, True)

        # Stage 2: Playwright fallback for failed/partial results (if not forced already)
        if self.enrichment_mode != "PLAYWRIGHT_ONLY":
            retry_urls = [att.profile_url for att in attorneys if getattr(att, "extraction_status", "FAILED") in ["FAILED", "PARTIAL"]]
            if retry_urls:
                self.log(f"  Playwright fallback for {len(retry_urls)} profiles")
                retry_results = _run_batch(retry_urls, extractor_playwright, True)

                # Merge retry results by URL
                retry_map = {att.profile_url: att for att in retry_results}
                merged = []
                for att in attorneys:
                    merged.append(retry_map.get(att.profile_url, att))
                attorneys = merged

        # Track extraction status
        for profile in attorneys:
            if profile.extraction_status == "SUCCESS":
                enriched_success += 1
            elif profile.extraction_status == "PARTIAL":
                enriched_partial += 1
            else:
                enriched_failed += 1

        self.log(f"  Enrichment: {enriched_success} SUCCESS, {enriched_partial} PARTIAL, {enriched_failed} FAILED")

        # Log aggregated diagnostics
        if enriched_failed > 0 or enriched_partial > 0:
            # Aggregate reason codes
            reason_counts = {}
            for att in attorneys:
                if att.extraction_status != "SUCCESS":
                    for key, value in att.diagnostics.items():
                        if "_reason" in key:
                            reason_counts[value] = reason_counts.get(value, 0) + 1

            if reason_counts:
                self.log(f"  Top missing field reasons:")
                for reason, count in sorted(reason_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
                    self.log(f"    {reason}: {count}")

        # FieldEnricher pass: enrich PARTIAL/FAILED profiles using stored raw_html
        for att in attorneys:
            raw_html = att.diagnostics.get('raw_html', '')
            if raw_html and getattr(att, 'extraction_status', '') in ('PARTIAL', 'FAILED'):
                try:
                    self.field_enricher.enrich(
                        att,
                        raw_html,
                        profile_url=att.profile_url,
                        source_type="official_profile_html",
                    )
                except Exception as _fe_err:
                    att.diagnostics['field_enricher_error'] = str(_fe_err)

        return attorneys

    def _fetch_external_directory_data(self, firm_name: str, existing_attorneys: list) -> list:
        """Fetch attorney data from external directories for BLOCKED firms

        Args:
            firm_name: Name of the firm
            existing_attorneys: List of attorneys from firm website (may be empty/failed)

        Returns:
            List of AttorneyProfile objects from external directories
        """
        try:
            from external_directory_extractor import ExternalDirectoryExtractor

            extractor = ExternalDirectoryExtractor(
                session=self.session,
                timeout=REQUEST_TIMEOUT
            )

            # Fetch from external directories (limit to reasonable number)
            max_external = max(len(existing_attorneys), 50)
            external_attorneys, _ext_summaries = extractor.extract_by_firm(firm_name, max_results=max_external)

            self.log(f"  External directory extracted {len(external_attorneys)} attorneys")
            return external_attorneys

        except Exception as e:
            self.log(f"  External directory extraction failed: {e}")
            return []

    def _merge_external_data(self, firm_attorneys: list, external_attorneys: list) -> list:
        """Merge external directory data with firm website data

        MERGE RULES:
        - Firm website data takes precedence (never overwrite)
        - External directory only fills missing fields
        - All profiles include data_source indicator

        Args:
            firm_attorneys: List of AttorneyProfile from firm website
            external_attorneys: List of AttorneyProfile from external directories

        Returns:
            Merged list of AttorneyProfile objects
        """
        from attorney_extractor import AttorneyProfile, EducationRecord

        # Mark all firm website data with data_source
        for att in firm_attorneys:
            if hasattr(att, 'diagnostics'):
                # Only mark as firm_website if not already marked as external
                if 'data_source' not in att.diagnostics:
                    att.diagnostics['data_source'] = 'firm_website'

        # Create name index for matching
        firm_names = {}
        for att in firm_attorneys:
            if hasattr(att, 'full_name') and att.full_name:
                # Normalize name for matching
                norm_name = att.full_name.lower().strip()
                firm_names[norm_name] = att

        # Merge external data
        merged = list(firm_attorneys)  # Start with firm website data
        new_profiles = []

        for ext_att in external_attorneys:
            if not hasattr(ext_att, 'full_name') or not ext_att.full_name:
                continue

            norm_name = ext_att.full_name.lower().strip()

            if norm_name in firm_names:
                # Match found - fill missing fields only
                firm_att = firm_names[norm_name]

                # Fill missing fields from external directory
                if not firm_att.title and ext_att.title:
                    firm_att.title = ext_att.title
                    firm_att.diagnostics['title_source'] = 'external_directory'

                if not firm_att.offices and ext_att.offices:
                    firm_att.offices = ext_att.offices
                    firm_att.diagnostics['offices_source'] = 'external_directory'

                if not firm_att.practice_areas and ext_att.practice_areas:
                    firm_att.practice_areas = ext_att.practice_areas
                    firm_att.diagnostics['practice_areas_source'] = 'external_directory'

                if not firm_att.bar_admissions and ext_att.bar_admissions:
                    firm_att.bar_admissions = ext_att.bar_admissions
                    firm_att.diagnostics['bar_admissions_source'] = 'external_directory'

                if not firm_att.education and ext_att.education:
                    firm_att.education = ext_att.education
                    firm_att.diagnostics['education_source'] = 'external_directory'

                # Recalculate status after merge
                firm_att.calculate_status()
            else:
                # No match - apply hard gate before adding as new profile
                ext_att.diagnostics['data_source'] = 'external_directory'
                if _ext_dir_profile_valid(ext_att):
                    new_profiles.append(ext_att)
                else:
                    self.log(f"  [EXT_DIR_REJECT] {getattr(ext_att, 'full_name', 'unknown')} \u2014 failed hard gate")

        # Add new external profiles
        merged.extend(new_profiles)

        self.log(f"  Filled missing fields in {len(firm_names)} existing profiles")
        self.log(f"  Added {len(new_profiles)} new profiles from external directory")

        return merged

    def _save_debug_artifacts(self, page, base_url: str, metrics: DiscoveryMetrics) -> None:
        """Save screenshot and HTML for debugging"""
        try:
            safe_name = re.sub(r'[^\w\s-]', '', urlparse(base_url).netloc).strip().replace('.', '_')

            screenshot_path = self.debug_dir / f"{safe_name}_screenshot.png"
            html_path = self.debug_dir / f"{safe_name}_page.html"

            page.screenshot(path=str(screenshot_path))
            html = page.content()
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html)

            self.log(f"  Debug artifacts saved: {screenshot_path}, {html_path}")
        except Exception as e:
            self.log(f"  Could not save debug artifacts: {e}")

    def _extract_profile_urls_from_sitemap(self, base_url: str) -> list[str]:
        """Extract attorney profile URLs from sitemap (with robots.txt check)"""
        domain = urlparse(base_url).netloc
        scheme = urlparse(base_url).scheme

        # Check robots.txt for sitemap directives
        sitemap_urls = []
        try:
            robots_url = f"{scheme}://{domain}/robots.txt"
            self.rate_limit_manager.wait(domain)
            resp = self.session.get(robots_url, timeout=REQUEST_TIMEOUT)
            if resp.status_code == 200:
                for line in resp.text.split('\n'):
                    if line.lower().startswith('sitemap:'):
                        sitemap_url = line.split(':', 1)[1].strip()
                        sitemap_urls.append(sitemap_url)
        except Exception:
            pass

        # Add default locations
        if not sitemap_urls:
            sitemap_urls = [
                f"{scheme}://{domain}/sitemap.xml",
                f"{scheme}://{domain}/sitemap_index.xml",
            ]

        all_profile_urls = set()
        for sitemap_url in sitemap_urls:
            try:
                try:
                    self.rate_limit_manager.wait(domain)
                except RateLimitBlockedError:
                    continue
                resp = self.session.get(sitemap_url, timeout=REQUEST_TIMEOUT)
                if resp.status_code != 200:
                    continue

                content = resp.content
                if sitemap_url.endswith(".gz"):
                    content = gzip.decompress(content)

                root = ET.fromstring(content)
                self._parse_sitemap_recursive(root, all_profile_urls, base_url)

                if all_profile_urls:
                    break
            except Exception:
                continue

        return list(all_profile_urls)

    def _parse_sitemap_recursive(
        self, element: ET.Element, urls: set, base_url: str
    ) -> None:
        """Recursively parse sitemap XML"""
        domain = urlparse(base_url).netloc

        for child in element:
            tag = child.tag.lower()
            # Handle namespaces
            if "}" in tag:
                tag = tag.split("}", 1)[1]

            if "loc" in tag and child.text:
                loc = child.text.strip()
                if loc.endswith(".xml") or loc.endswith(".xml.gz"):
                    try:
                        try:
                            self.rate_limit_manager.wait(domain)
                        except RateLimitBlockedError:
                            pass
                        resp = self.session.get(loc, timeout=REQUEST_TIMEOUT)
                        if resp.status_code == 200:
                            content = resp.content
                            if loc.endswith(".gz"):
                                content = gzip.decompress(content)
                            sub_root = ET.fromstring(content)
                            self._parse_sitemap_recursive(sub_root, urls, base_url)
                    except Exception:
                        pass
                else:
                    if self._is_attorney_profile_url(loc):
                        # PROBLEM 2 FIX: reject URLs whose domain doesn't match the
                        # official firm domain (e.g. CDN subdomains like www-cm-prod.lw.com)
                        def _strip_www(h: str) -> str:
                            return h[4:] if h.startswith('www.') else h
                        loc_netloc = urlparse(loc).netloc
                        official_root = _strip_www(domain)
                        loc_root = _strip_www(loc_netloc)
                        if loc_root == official_root or loc_root.endswith('.' + official_root):
                            urls.add(loc)

            # Recurse into children
            self._parse_sitemap_recursive(child, urls, base_url)

    def _is_attorney_profile_url(self, url: str) -> bool:
        """Check if URL is attorney profile"""
        url_lower = url.lower()
        keywords = ["/lawyer", "/attorney", "/people/", "/professional", "/bio/", "/profile/"]
        if not any(kw in url_lower for kw in keywords):
            return False

        # Filter junk
        junk_keywords = [
            "terms",
            "privacy",
            "advertising",
            "disclaimer",
            "cookie",
            "sitemap",
        ]
        if any(junk in url_lower for junk in junk_keywords):
            return False

        return True

    def _fetch_profile_metadata(self, url: str, diagnostics: dict | None = None) -> dict | None:
        """Fetch metadata from profile page with robust extraction and diagnostics"""
        try:
            domain = urlparse(url).netloc
            try:
                self.rate_limit_manager.wait(domain)
            except RateLimitBlockedError:
                return None

            resp = self.session.get(url, timeout=REQUEST_TIMEOUT)

            # Diagnostic: Track HTTP errors
            if resp.status_code == 403:
                if diagnostics is not None:
                    diagnostics["HTTP 403 Forbidden"] = diagnostics.get("HTTP 403 Forbidden", 0) + 1
                return None
            elif resp.status_code == 429:
                if diagnostics is not None:
                    diagnostics["HTTP 429 Rate Limited"] = diagnostics.get("HTTP 429 Rate Limited", 0) + 1
                return None
            elif resp.status_code != 200:
                if diagnostics is not None:
                    diagnostics[f"HTTP {resp.status_code}"] = diagnostics.get(f"HTTP {resp.status_code}", 0) + 1
                return None

            html = resp.text

            # Diagnostic: Check for likely blocking/redirect (very small content)
            if len(html) < 2000:
                if diagnostics is not None:
                    diagnostics["Small content (<2KB, likely blocked/redirect)"] = diagnostics.get("Small content (<2KB, likely blocked/redirect)", 0) + 1
                return None

            # Enhanced extraction order
            name = self._extract_name_from_html(html, url)
            title = self._extract_title_from_html(html)
            practice = self._extract_practice_from_html(html)
            office = self._extract_office_from_html(html)

            # Name is required, but return partial data if available
            if not name:
                if diagnostics is not None:
                    diagnostics["No name extracted"] = diagnostics.get("No name extracted", 0) + 1
                return None

            return {
                "name": name,
                "title": title,
                "practice": practice,
                "office": office,
                "url": url,
            }
        except Exception as e:
            if diagnostics is not None:
                diagnostics[f"Exception: {type(e).__name__}"] = diagnostics.get(f"Exception: {type(e).__name__}", 0) + 1
            return None

    def _extract_name_from_html(self, html: str, url: str) -> str:
        """Extract name from HTML with enhanced extraction order

        Extraction priority:
        1. JSON-LD structured data
        2. OpenGraph meta tags
        3. H1 or elements with name-like class
        4. Title tag
        5. URL path fallback
        """

        # 1. JSON-LD extraction (highest priority for structured data)
        try:
            json_ld_blocks = re.findall(
                r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
                html,
                re.IGNORECASE | re.DOTALL
            )

            for block in json_ld_blocks:
                try:
                    data = json.loads(block)

                    # Handle both single objects and arrays
                    items = data if isinstance(data, list) else [data]

                    for item in items:
                        if isinstance(item, dict):
                            # Look for Person type
                            if item.get("@type") in ["Person", "http://schema.org/Person"]:
                                name = item.get("name", "")
                                if name and self._looks_like_person_name(name):
                                    return name.strip()
                except json.JSONDecodeError:
                    pass
        except Exception:
            pass

        # 2. OpenGraph fallback
        og_title_match = re.search(
            r'<meta[^>]*property=["\']og:title["\'][^>]*content=["\']([^"\']+)["\']',
            html,
            re.IGNORECASE
        )
        if og_title_match:
            og_title = og_title_match.group(1).strip()
            # Split on common separators
            parts = re.split(r'[|\-–—]', og_title)
            if parts:
                name = parts[0].strip()
                if name and len(name) < 100 and self._looks_like_person_name(name):
                    return name

        # 3. More flexible H1 and name-like elements
        # Look for h1/h2/div/span with class containing "name"
        name_element_match = re.search(
            r'<(?:h1|h2|div|span)[^>]*class=["\'][^"\']*name[^"\']*["\'][^>]*>(.*?)</(?:h1|h2|div|span)>',
            html,
            re.IGNORECASE | re.DOTALL
        )
        if name_element_match:
            name = re.sub(r"<[^>]+>", "", name_element_match.group(1)).strip()
            if name and len(name) < 100 and self._looks_like_person_name(name):
                return name

        # Original H1 extraction
        h1_match = re.search(r"<h1[^>]*>(.*?)</h1>", html, re.IGNORECASE | re.DOTALL)
        if h1_match:
            name = re.sub(r"<[^>]+>", "", h1_match.group(1)).strip()
            if name and len(name) < 100 and self._looks_like_person_name(name):
                return name

        # 4. Title tag extraction
        title_match = re.search(
            r"<title[^>]*>(.*?)</title>", html, re.IGNORECASE | re.DOTALL
        )
        if title_match:
            title_text = re.sub(r"<[^>]+>", "", title_match.group(1)).strip()
            parts = re.split(r"[|\-–—]", title_text)
            if parts:
                name = parts[0].strip()
                if name and len(name) < 100 and self._looks_like_person_name(name):
                    return name

        # 5. URL path fallback
        path = urlparse(url).path
        segments = [s for s in path.split("/") if s]
        if segments:
            last = segments[-1].replace("-", " ").replace("_", " ").title()
            if self._looks_like_person_name(last):
                return last

        return ""

    def _extract_title_from_html(self, html: str) -> str:
        """Extract title/position from HTML with enhanced extraction

        Extraction priority:
        1. JSON-LD jobTitle field
        2. Elements with job/role/position/designation classes
        3. Common title keywords (Partner, Associate, etc)
        """

        # 1. JSON-LD jobTitle
        try:
            json_ld_blocks = re.findall(
                r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
                html,
                re.IGNORECASE | re.DOTALL
            )

            for block in json_ld_blocks:
                try:
                    data = json.loads(block)
                    items = data if isinstance(data, list) else [data]

                    for item in items:
                        if isinstance(item, dict):
                            if item.get("@type") in ["Person", "http://schema.org/Person"]:
                                job_title = item.get("jobTitle", "")
                                if job_title and len(job_title) < 200:
                                    return job_title.strip()
                except json.JSONDecodeError:
                    pass
        except Exception:
            pass

        # 2. Enhanced class-based extraction
        patterns = [
            r'<[^>]*class="[^"]*(?:title|position|role|job|designation)[^"]*"[^>]*>(.*?)</[^>]+>',
            r'<span[^>]*>\s*(?:Partner|Associate|Counsel|Of Counsel|Senior Associate|Managing Partner|Senior Partner|Member|Shareholder|Principal)\s*</span>',
        ]
        for pat in patterns:
            match = re.search(pat, html, re.IGNORECASE | re.DOTALL)
            if match:
                text = re.sub(r"<[^>]+>", "", match.group(1)).strip()
                if text and len(text) < 200:
                    return text

        return ""

    def _extract_practice_from_html(self, html: str) -> str:
        """Extract practice area from HTML with JSON-LD support"""

        # 1. JSON-LD fields (knowsAbout, hasOccupation)
        try:
            json_ld_blocks = re.findall(
                r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
                html,
                re.IGNORECASE | re.DOTALL
            )

            for block in json_ld_blocks:
                try:
                    data = json.loads(block)
                    items = data if isinstance(data, list) else [data]

                    for item in items:
                        if isinstance(item, dict) and item.get("@type") in ["Person", "http://schema.org/Person"]:
                            # Check knowsAbout field
                            knows_about = item.get("knowsAbout", [])
                            if isinstance(knows_about, list) and knows_about:
                                practices = [str(k).strip() for k in knows_about if k]
                                if practices:
                                    return ", ".join(practices[:5])
                            elif isinstance(knows_about, str) and knows_about:
                                return knows_about.strip()
                except json.JSONDecodeError:
                    pass
        except Exception:
            pass

        # 2. Original regex fallback
        practice_section = re.search(
            r"<[^>]*>Practice[s]?\s*(?:Area[s]?)?</[^>]*>(.*?)</(?:div|section|ul)",
            html,
            re.IGNORECASE | re.DOTALL,
        )
        if practice_section:
            content = practice_section.group(1)
            links = re.findall(r"<a[^>]*>(.*?)</a>", content, re.IGNORECASE)
            practices = [re.sub(r"<[^>]+>", "", link).strip() for link in links]
            practices = [p for p in practices if p and len(p) < 100]
            if practices:
                return ", ".join(practices[:5])

        return ""

    def _extract_office_from_html(self, html: str) -> str:
        """Extract office/location from HTML with JSON-LD support"""

        # 1. JSON-LD address/workLocation fields
        try:
            json_ld_blocks = re.findall(
                r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
                html,
                re.IGNORECASE | re.DOTALL
            )

            for block in json_ld_blocks:
                try:
                    data = json.loads(block)
                    items = data if isinstance(data, list) else [data]

                    for item in items:
                        if isinstance(item, dict) and item.get("@type") in ["Person", "http://schema.org/Person"]:
                            # Check workLocation
                            work_location = item.get("workLocation", {})
                            if isinstance(work_location, dict):
                                location_name = work_location.get("name", "")
                                if location_name:
                                    return location_name.strip()

                            # Check address
                            address = item.get("address", {})
                            if isinstance(address, dict):
                                city = address.get("addressLocality", "")
                                if city:
                                    return city.strip()
                            elif isinstance(address, str) and address:
                                return address.strip()
                except json.JSONDecodeError:
                    pass
        except Exception:
            pass

        # 2. Original regex fallback
        patterns = [
            r'<[^>]*class="[^"]*office[^"]*"[^>]*>(.*?)</[^>]+>',
            r'<[^>]*class="[^"]*location[^"]*"[^>]*>(.*?)</[^>]+>',
        ]
        for pat in patterns:
            match = re.search(pat, html, re.IGNORECASE | re.DOTALL)
            if match:
                text = re.sub(r"<[^>]+>", "", match.group(1)).strip()
                if text and len(text) < 200:
                    return text

        return ""

    def _extract_attorneys_from_json(self, data: Any, base_url: str) -> list[dict]:
        """Extract attorney records from JSON"""
        records = []

        def visit(obj):
            if isinstance(obj, dict):
                has_name = any(
                    k in obj
                    for k in ["name", "fullName", "displayName", "firstName", "lastName"]
                )
                if has_name:
                    records.append(obj)
                for v in obj.values():
                    visit(v)
            elif isinstance(obj, list):
                for item in obj:
                    visit(item)

        visit(data)

        attorneys = []
        for rec in records:
            att = self._normalize_json_record(rec, base_url)
            if att and att.get("name") and self._looks_like_person_name(att["name"]):
                attorneys.append(att)

        return attorneys

    def _normalize_json_record(self, rec: dict, base_url: str) -> dict | None:
        """Normalize JSON record"""

        def pick(*keys):
            for k in keys:
                if k in rec and rec[k]:
                    return rec[k]
            return None

        def to_text(val):
            if val is None:
                return ""
            if isinstance(val, str):
                return val.strip()
            if isinstance(val, list):
                parts = [to_text(v) for v in val if v]
                return ", ".join(parts)
            return str(val).strip()

        first = pick("firstName", "first_name", "first")
        last = pick("lastName", "last_name", "last")
        name = pick("name", "fullName", "displayName") or " ".join(
            filter(None, [to_text(first), to_text(last)])
        )

        title = pick("title", "position", "role", "jobTitle")
        practice = pick("practice", "practiceAreas", "practices", "services")
        office = pick("office", "location", "officeName")
        url = pick("url", "profileUrl", "link", "path")

        url_text = to_text(url)
        if url_text and not url_text.startswith("http"):
            url_text = urljoin(base_url, url_text)

        name_text = to_text(name)
        if not name_text or len(name_text) < 4:
            return None

        return {
            "name": name_text,
            "title": to_text(title),
            "practice": to_text(practice),
            "office": to_text(office),
            "url": url_text,
        }

    def _looks_like_person_name(self, text: str) -> bool:
        """Check if text looks like person name"""
        if not text or len(text) < 4:
            return False
        if any(ch in text for ch in ["_", "#", "{"]):
            return False
        parts = [p for p in text.split() if p and p[0].isupper()]
        return len(parts) >= 2

    def _extract_single_profile(self, url: str, firm_name: str, extractor):
        """Extract single profile using AttorneyExtractor

        Args:
            url: Profile URL
            firm_name: Firm name
            extractor: AttorneyExtractor instance

        Returns:
            AttorneyProfile with extraction results
        """
        from attorney_extractor import AttorneyProfile

        domain = urlparse(url).netloc
        try:
            self.rate_limit_manager.wait(domain)
        except RateLimitBlockedError:
            profile = AttorneyProfile(firm=firm_name, profile_url=url)
            profile.extraction_status = "FAILED"
            profile.diagnostics["rate_limit_blocked"] = True
            profile.missing_fields = [
                "full_name", "title", "offices", "department",
                "practice_areas", "industries", "bar_admissions", "education"
            ]
            return profile

        try:
            resp = self.session.get(url, timeout=REQUEST_TIMEOUT)

            if resp.status_code != 200:
                # Return failed profile with HTTP error
                profile = AttorneyProfile(firm=firm_name, profile_url=url)
                profile.extraction_status = "FAILED"
                profile.diagnostics[f"HTTP_{resp.status_code}"] = True
                profile.missing_fields = ["full_name", "title", "offices", "department", "practice_areas", "industries", "bar_admissions", "education"]
                return profile

            html = resp.text

            # Check for blocking/redirect (very small content)
            if len(html) < 2000:
                profile = AttorneyProfile(firm=firm_name, profile_url=url)
                profile.extraction_status = "FAILED"
                profile.diagnostics["small_content"] = True
                profile.missing_fields = ["full_name", "title", "offices", "department", "practice_areas", "industries", "bar_admissions", "education"]
                return profile

            # Use AttorneyExtractor to extract all fields
            profile = extractor.extract_profile(firm_name, url, html)
            return profile

        except Exception as e:
            # Return failed profile with exception
            profile = AttorneyProfile(firm=firm_name, profile_url=url)
            profile.extraction_status = "FAILED"
            profile.diagnostics["exception"] = str(type(e).__name__)
            profile.missing_fields = ["full_name", "title", "offices", "department", "practice_areas", "industries", "bar_admissions", "education"]
            return profile


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="High-coverage attorney scraper with stabilization-based discovery"
    )
    parser.add_argument("excel_path", nargs="?", help="Excel file path")
    parser.add_argument(
        "--limit", type=int, default=0, help="Max attorneys per firm (0=all)"
    )
    parser.add_argument(
        "--max-firms", type=int, default=0, help="Max firms to process (0=all)"
    )
    parser.add_argument(
        "--workers", type=int, default=MAX_WORKERS, help="Parallel workers"
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Output sheet (default: {DEFAULT_SHEET_NAME})",
    )
    parser.add_argument(
        "--debug-firm", default="", help="Process only firms matching this name"
    )
    parser.add_argument(
        "--debug-domain", default="", help="Process only domains matching this string"
    )
    parser.add_argument(
        "--headful", type=bool, default=False, help="Run browser in headful mode"
    )
    parser.add_argument(
        "--stabilization",
        type=int,
        default=DEFAULT_STABILIZATION,
        help=f"Stabilization threshold (default: {DEFAULT_STABILIZATION})",
    )
    parser.add_argument(
        "--max-scroll-seconds",
        type=int,
        default=DEFAULT_MAX_SCROLL_SECONDS,
        help=f"Max seconds per directory mode (default: {DEFAULT_MAX_SCROLL_SECONDS})",
    )
    parser.add_argument(
        "--sources-file",
        default="",
        help="Multi-source discovery Excel from firm_finder_desktop.py (enables multi-source aggregation)",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory for output files (attorneys.jsonl, coverage_metrics.json)",
    )
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    log_level = logging.INFO
    if args.debug_firm or args.debug_domain:
        log_level = logging.DEBUG
    logging.basicConfig(level=log_level, format="%(levelname)s:%(name)s:%(message)s")
    excel_path = args.excel_path
    if not excel_path:
        excel_path = input("Excel file path: ").strip().strip('"')
    if not excel_path:
        raise SystemExit(2)

    finder = AttorneyFinder(
        limit=args.limit,
        sheet_name=args.sheet,
        max_firms=args.max_firms,
        workers=args.workers,
        debug_firm=args.debug_firm,
        debug_domain=args.debug_domain,
        headful=args.headful,
        stabilization=args.stabilization,
        max_scroll_seconds=args.max_scroll_seconds,
        sources_file=args.sources_file,
        output_dir=args.output_dir,
    )
    start = time.time()
    total = finder.run(excel_path)
    elapsed = time.time() - start
    finder.log(f"\nTotal runtime: {elapsed:.1f}s")
    return 0 if total >= 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
    #from discovery import discover_attorneys
    #urls = discover_attorneys("https://www.kirkland.com")
    #print(len(urls))
