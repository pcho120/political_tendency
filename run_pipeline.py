#!/usr/bin/env python3
"""run_pipeline.py - AmLaw200 Attorney Extraction Pipeline Runner
  1. Load firm -> URL mapping from cache/firm_domain_cache.json
  2. For each firm: discover attorney profile URLs  (discovery.py)
  3. For each profile URL: enrich the profile       (enrichment.py)
  4. Write results to:
       outputs/attorneys_<timestamp>.xlsx
       outputs/attorneys_<timestamp>.jsonl
  5. Per-firm debug logs written to debug_reports/<firm>/
    # Run all 200 firms
    python run_pipeline.py
    python run_pipeline.py --firms "kirkland" "latham"
    python run_pipeline.py --firms "gibson dunn" --max-profiles 10
    python run_pipeline.py --skip-discovery --resume outputs/attorneys_2026-02-23.jsonl
    python run_pipeline.py --discover-only --firms "kirkland"
    python run_pipeline.py --workers 4
    PIPELINE_RATE_DELAY    seconds between requests per firm (default 0.5)
    PIPELINE_TIMEOUT       HTTP timeout in seconds (default 15)
    PIPELINE_NO_PLAYWRIGHT set to "1" to disable Playwright entirely
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import signal
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from debug_logger import DebugLogger
from discovery import discover_attorneys, lookup_structure
from enrichment import ProfileEnricher
from attorney_extractor import AttorneyProfile, EducationRecord

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

CACHE_FILE = Path("cache") / "firm_domain_cache.json"
OUTPUT_DIR = Path("outputs")
DEBUG_DIR = Path("debug_reports")
_STOP_FILE = Path("STOP")

OUTPUT_DIR.mkdir(exist_ok=True)
DEBUG_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

log = logging.getLogger("pipeline")


# ---------------------------------------------------------------------------
# Excel output helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_OUTPUT_COLUMNS = [
    "Firm",
    "Attorney Name",
    "Title",
    "Offices",
    "Departments",
    "Practice Areas",
    "Industries",
    "Bar Admissions",
    "Education",
    "Extraction Status",
    "Missing Fields",
    "Profile URL",
    "Data Source",
]


def _profile_to_row(profile: AttorneyProfile) -> list[Any]:
    """Convert an AttorneyProfile to a flat list matching _OUTPUT_COLUMNS."""

    def _join(lst: list) -> str:
        return " | ".join(str(x) for x in lst) if lst else ""

    def _edu_str(records: list[EducationRecord]) -> str:
        parts = []
        for r in records:
            tokens = []
            if r.degree:
                tokens.append(r.degree)
            if r.school:
                tokens.append(r.school)
            if r.year:
                tokens.append(str(r.year))
            parts.append(", ".join(tokens))
        return " | ".join(parts)

    return [
        profile.firm,
        profile.full_name or "",
        profile.title or "",
        _join(profile.offices),
        _join(profile.department),
        _join(profile.practice_areas),
        _join(profile.industries),
        _join(profile.bar_admissions),
        _edu_str(profile.education),
        profile.extraction_status,
        _join(profile.missing_fields),
        profile.profile_url,
        "firm_website",
    ]


def _write_excel(profiles: list[AttorneyProfile], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attorneys"

    # Header row
    ws.append(_OUTPUT_COLUMNS)
    for col_idx, _ in enumerate(_OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for profile in profiles:
        ws.append(_profile_to_row(profile))

    # Auto-size columns (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(path)
    log.info(f"Excel saved → {path} ({len(profiles)} rows)")


def _write_jsonl(profiles: list[AttorneyProfile], path: Path) -> None:
    with path.open("w", encoding="utf-8") as fh:
        for p in profiles:
            fh.write(json.dumps(asdict(p)) + "\n")
    log.info(f"JSONL saved → {path} ({len(profiles)} records)")


# ---------------------------------------------------------------------------
# Firm loading
# ---------------------------------------------------------------------------

@dataclass
class FirmEntry:
    name: str          # human-readable name (from cache key)
    url: str           # chosen_url from cache
    confidence: float = 1.0


def load_firms(filter_names: list[str] | None = None) -> list[FirmEntry]:
    """Load firm → URL mapping from the domain cache."""
    if not CACHE_FILE.exists():
        log.error(f"Firm domain cache not found: {CACHE_FILE}")
        sys.exit(1)

    raw: dict[str, dict] = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
    firms: list[FirmEntry] = []

    for key, data in raw.items():
        url = data.get("chosen_url") or data.get("domain")
        if not url:
            continue
        if not url.startswith("http"):
            url = f"https://{url}"

        entry = FirmEntry(
            name=key,
            url=url,
            confidence=float(data.get("confidence", 1.0)),
        )
        firms.append(entry)

    if filter_names:
        # Case-insensitive partial match
        lowered = [f.lower() for f in filter_names]
        firms = [
            f for f in firms
            if any(token in f.name.lower() for token in lowered)
        ]

    log.info(f"Loaded {len(firms)} firms from {CACHE_FILE}")
    return firms


# ---------------------------------------------------------------------------
# Per-firm pipeline
# ---------------------------------------------------------------------------

@dataclass
class FirmResult:
    firm: FirmEntry
    discovered: int = 0
    enriched: int = 0
    failed: int = 0
    profiles: list[AttorneyProfile] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    elapsed_seconds: float = 0.0


def run_firm(
    firm: FirmEntry,
    *,
    max_profiles: int | None = None,
    discover_only: bool = False,
    rate_delay: float = 0.5,
    timeout: int = 15,
    enable_playwright: bool = True,
) -> FirmResult:
    """Run discovery + enrichment for a single firm."""
    result = FirmResult(firm=firm)
    t0 = time.monotonic()

    firm_debug_dir = DEBUG_DIR / _safe_name(firm.name)
    firm_debug_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers["User-Agent"] = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )

    logger = DebugLogger(firm=firm.name, output_dir=firm_debug_dir, verbose=True)

    # ----------------------------------------------------------------
    # Stage 0 — Look up structure type from site_structures.json
    # ----------------------------------------------------------------
    structure_info = lookup_structure(firm.url)
    structure_type = (structure_info or {}).get("structure_type", "UNKNOWN")
    log.info(f"[{firm.name}] Structure type: {structure_type}")

    # Skip bot-protected / auth-required firms immediately
    if structure_type in ("BOT_PROTECTED", "AUTH_REQUIRED"):
        log.info(f"[{firm.name}] Skipping — {structure_type}")
        result.errors.append(f"Skipped: {structure_type}")
        logger.flush()
        result.elapsed_seconds = time.monotonic() - t0
        return result

    # ----------------------------------------------------------------
    # Stage 1 — Discovery
    # ----------------------------------------------------------------
    log.info(f"[{firm.name}] Discovery starting → {firm.url}")
    try:
        profile_urls = discover_attorneys(
            firm.url,
            session=session,
            logger=logger,
            timeout=timeout,
            rate_delay=rate_delay,
            structure_info=structure_info,
        )
    except Exception as exc:
        msg = f"Discovery failed: {exc}"
        log.error(f"[{firm.name}] {msg}", exc_info=True)
        result.errors.append(msg)
        logger.flush()
        result.elapsed_seconds = time.monotonic() - t0
        return result

    result.discovered = len(profile_urls)
    log.info(f"[{firm.name}] Discovered {result.discovered} profile URLs")

    if discover_only or not profile_urls:
        logger.flush()
        result.elapsed_seconds = time.monotonic() - t0
        return result

    # Apply limit
    if max_profiles and len(profile_urls) > max_profiles:
        log.info(f"[{firm.name}] Limiting to {max_profiles} profiles (--max-profiles)")
        profile_urls = profile_urls[:max_profiles]

    # ----------------------------------------------------------------
    # Stage 2 — Enrichment
    # ----------------------------------------------------------------
    enricher = ProfileEnricher(
        session=session,
        logger=logger,
        enable_playwright=enable_playwright,
        timeout=timeout,
    )

    for i, url in enumerate(profile_urls, start=1):
        t_profile = time.monotonic()
        try:
            profile = enricher.enrich(url=url, html="", firm=firm.name)
            result.profiles.append(profile)

            status = profile.extraction_status
            elapsed_ms = int((time.monotonic() - t_profile) * 1000)
            logger.log_profile_result(
                url=url,
                status=status,
                missing_fields=profile.missing_fields,
                elapsed_ms=elapsed_ms,
            )

            if status in ("SUCCESS", "PARTIAL"):
                result.enriched += 1
            else:
                result.failed += 1

            if i % 25 == 0:
                log.info(
                    f"[{firm.name}] Progress: {i}/{len(profile_urls)} "
                    f"enriched={result.enriched} failed={result.failed}"
                )

        except Exception as exc:
            msg = f"Enrich failed ({url}): {exc}"
            log.error(f"[{firm.name}] {msg}", exc_info=True)
            result.errors.append(msg)
            result.failed += 1
            logger.log_profile_result(url=url, status="FAILED", notes=[str(exc)])

        # Polite rate limiting between profiles
        if rate_delay > 0:
            time.sleep(rate_delay)

    logger.flush()
    result.elapsed_seconds = time.monotonic() - t0
    log.info(
        f"[{firm.name}] Done | enriched={result.enriched} failed={result.failed} "
        f"elapsed={result.elapsed_seconds:.1f}s"
    )
    return result


def _safe_name(name: str) -> str:
    """Filesystem-safe firm name."""
    return re.sub(r"[^a-z0-9_-]", "_", name.lower().strip())


# ---------------------------------------------------------------------------
# Graceful stop utilities
# ---------------------------------------------------------------------------

def _parse_duration(s: str) -> int:
    """Parse '30m', '2h', '1h30m', '90s' → seconds. Raises ValueError on invalid."""
    s = s.strip().lower()
    m = re.fullmatch(r'(?:(\d+)h)?(?:(\d+)m)?(?:(\d+)s)?', s)
    if not m or not any(m.groups()):
        raise ValueError(f"Invalid duration '{s}'. Use formats like: 30m, 2h, 1h30m, 90s")
    h = int(m.group(1) or 0)
    mn = int(m.group(2) or 0)
    sc = int(m.group(3) or 0)
    total = h * 3600 + mn * 60 + sc
    if total <= 0:
        raise ValueError(f"Duration must be positive, got '{s}'")
    return total


# ---------------------------------------------------------------------------
# Resume from existing JSONL
# ---------------------------------------------------------------------------

def load_existing_jsonl(path: Path) -> list[AttorneyProfile]:
    """Load profiles already extracted from a previous run's JSONL output."""
    if not path.exists():
        log.error(f"Resume file not found: {path}")
        sys.exit(1)

    profiles = []
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            data = json.loads(line)
            edu = [
                EducationRecord(
                    degree=e.get("degree"),
                    school=e.get("school"),
                    year=e.get("year"),
                )
                for e in data.pop("education", [])
            ]
            profile = AttorneyProfile(**{k: v for k, v in data.items()
                                         if k != "education"})
            profile.education = edu
            profiles.append(profile)

    log.info(f"Loaded {len(profiles)} profiles from {path}")
    return profiles


# ---------------------------------------------------------------------------
# Run summary
# ---------------------------------------------------------------------------

def _print_run_summary(results: list[FirmResult], elapsed: float) -> None:
    total_discovered = sum(r.discovered for r in results)
    total_enriched = sum(r.enriched for r in results)
    total_failed = sum(r.failed for r in results)
    total_errors = sum(len(r.errors) for r in results)

    print("\n" + "=" * 65)
    print(f"  PIPELINE COMPLETE")
    print(f"  Firms processed : {len(results)}")
    print(f"  URLs discovered : {total_discovered}")
    print(f"  Profiles enriched: {total_enriched}")
    print(f"  Failed profiles : {total_failed}")
    print(f"  Errors          : {total_errors}")
    print(f"  Total elapsed   : {elapsed:.1f}s")
    print("=" * 65)

    # Per-firm table
    print(f"\n{'Firm':<35} {'Disc':>6} {'Enriched':>9} {'Failed':>7} {'Errors':>7} {'Time':>7}")
    print("-" * 75)
    for r in sorted(results, key=lambda x: -x.enriched):
        print(
            f"{r.firm.name[:34]:<35} "
            f"{r.discovered:>6} {r.enriched:>9} {r.failed:>7} "
            f"{len(r.errors):>7} {r.elapsed_seconds:>6.1f}s"
        )
    print()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="AmLaw200 Attorney Extraction Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument(
        "--firms",
        nargs="+",
        metavar="NAME",
        help="Partial firm name(s) to filter (case-insensitive). Omit to run all 200.",
    )
    p.add_argument(
        "--max-profiles",
        type=int,
        default=None,
        metavar="N",
        help="Cap the number of profiles enriched per firm (useful for testing).",
    )
    p.add_argument(
        "--workers",
        type=int,
        default=1,
        metavar="N",
        help="Number of parallel firm workers (default 1 = sequential).",
    )
    p.add_argument(
        "--discover-only",
        action="store_true",
        help="Run discovery only; skip enrichment.",
    )
    p.add_argument(
        "--skip-discovery",
        action="store_true",
        help="Skip discovery; run enrichment only (requires --resume).",
    )
    p.add_argument(
        "--resume",
        metavar="FILE",
        help="Path to existing JSONL to append to (used with --skip-discovery).",
    )
    p.add_argument(
        "--output",
        metavar="FILE",
        help="Override output file base name (without extension).",
    )
    p.add_argument(
        "--rate-delay",
        type=float,
        default=float(os.getenv("PIPELINE_RATE_DELAY", "0.5")),
        metavar="SECONDS",
        help="Seconds between HTTP requests per firm (default 0.5).",
    )
    p.add_argument(
        "--timeout",
        type=int,
        default=int(os.getenv("PIPELINE_TIMEOUT", "15")),
        metavar="SECONDS",
        help="HTTP request timeout in seconds (default 15).",
    )
    p.add_argument(
        "--no-playwright",
        action="store_true",
        default=(os.getenv("PIPELINE_NO_PLAYWRIGHT", "") == "1"),
        help="Disable Playwright entirely (static requests only).",
    )
    p.add_argument(
        "--verbose",
        action="store_true",
        default=False,
        help="Enable DEBUG-level console output.",
    )
    p.add_argument(
        "--structure-type",
        metavar="TYPE",
        help=(
            "Only run firms with this structure_type (e.g. SITEMAP_XML, "
            "HTML_DIRECTORY_FLAT). Comma-separated for multiple."
        ),
    )
    p.add_argument(
        "--stop-after",
        metavar="DURATION",
        default=None,
        help="Gracefully stop after DURATION (e.g. 30m, 2h, 1h30m, 90s). "
             "Finishes current firm, saves partial output with _partial suffix.",
    )
    return p


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    # --- Graceful stop setup ---
    stop_after_secs: int | None = None
    if args.stop_after:
        try:
            stop_after_secs = _parse_duration(args.stop_after)
        except ValueError as e:
            parser.error(str(e))

    if _STOP_FILE.exists():
        log.warning("STOP file found at startup — deleting it and continuing.")
        _STOP_FILE.unlink()

    _stop_event = threading.Event()

    def _handle_sigint(signum: int, frame: object) -> None:
        log.warning("Ctrl+C received — will stop after current firm completes.")
        _stop_event.set()

    signal.signal(signal.SIGINT, _handle_sigint)
    # --- End graceful stop setup ---

    # Configure root logger
    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        datefmt="%H:%M:%S",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True,
    )

    # Timestamp for output file names
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%S")
    base_name = args.output or f"attorneys_{ts}"
    out_xlsx = OUTPUT_DIR / f"{base_name}.xlsx"
    out_jsonl = OUTPUT_DIR / f"{base_name}.jsonl"

    # ----------------------------------------------------------------
    # Resume mode: load existing profiles, re-write outputs, exit
    # ----------------------------------------------------------------
    if args.skip_discovery:
        if not args.resume:
            parser.error("--skip-discovery requires --resume <file>")
        existing = load_existing_jsonl(Path(args.resume))
        _write_excel(existing, out_xlsx)
        _write_jsonl(existing, out_jsonl)
        return 0

    # ----------------------------------------------------------------
    # Load firms
    # ----------------------------------------------------------------
    firms = load_firms(filter_names=args.firms)
    if not firms:
        log.error("No firms matched the filter. Exiting.")
        return 1

    # Filter by structure type if requested
    if args.structure_type:
        allowed_types = {t.strip().upper() for t in args.structure_type.split(",")}
        filtered = []
        for firm in firms:
            info = lookup_structure(firm.url)
            stype = (info or {}).get("structure_type", "UNKNOWN").upper()
            if stype in allowed_types:
                filtered.append(firm)
        log.info(
            f"Structure-type filter '{args.structure_type}': "
            f"{len(filtered)}/{len(firms)} firms retained"
        )
        firms = filtered
        if not firms:
            log.error("No firms matched the structure-type filter. Exiting.")
            return 1

    t_run_start = time.time()  # wall clock — survives laptop suspend
    all_profiles: list[AttorneyProfile] = []
    results: list[FirmResult] = []

    # ----------------------------------------------------------------
    # Run per-firm (sequential or parallel)
    # ----------------------------------------------------------------
    firm_kwargs = dict(
        max_profiles=args.max_profiles,
        discover_only=args.discover_only,
        rate_delay=args.rate_delay,
        timeout=args.timeout,
        enable_playwright=not args.no_playwright,
    )

    if args.workers > 1:
        log.info(f"Running {len(firms)} firms with {args.workers} parallel workers")
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futures: dict = {}
            for firm in firms:
                if _stop_event.is_set():
                    break
                futures[pool.submit(run_firm, firm, **firm_kwargs)] = firm
            for future in as_completed(futures):
                firm = futures[future]
                try:
                    result = future.result()
                except Exception as exc:
                    log.error(f"[{firm.name}] Unhandled exception: {exc}", exc_info=True)
                    result = FirmResult(firm=firm, errors=[str(exc)])
                results.append(result)
                all_profiles.extend(result.profiles)
                # Stop check after each future
                if _stop_event.is_set():
                    break
                if _STOP_FILE.exists():
                    log.info("STOP file detected during parallel run.")
                    _stop_event.set()
                    _STOP_FILE.unlink()
                elif stop_after_secs is not None and (time.time() - t_run_start) >= stop_after_secs:
                    log.info("--stop-after limit reached during parallel run — stopping.")
                    _stop_event.set()
    else:
        for idx, firm in enumerate(firms, start=1):
            # --- Stop check (firm boundary only) ---
            if _stop_event.is_set():
                log.info(f"Stop flag set — skipping remaining {len(firms) - idx + 1} firms.")
                break
            if _STOP_FILE.exists():
                log.info("STOP file detected — stopping after this check.")
                _stop_event.set()
                _STOP_FILE.unlink()
                break
            if stop_after_secs is not None:
                elapsed_now = time.time() - t_run_start
                if elapsed_now >= stop_after_secs:
                    log.info(f"--stop-after limit reached ({elapsed_now:.0f}s >= {stop_after_secs}s) — stopping.")
                    _stop_event.set()
                    break
            # --- End stop check ---
            log.info(f"── Firm {idx}/{len(firms)}: {firm.name} ──")
            result = run_firm(firm, **firm_kwargs)
            results.append(result)
            all_profiles.extend(result.profiles)

    # ----------------------------------------------------------------
    # Write outputs
    # ----------------------------------------------------------------
    # Clean up any leftover STOP file
    if _STOP_FILE.exists():
        _STOP_FILE.unlink()

    elapsed = time.time() - t_run_start
    stopped_early = _stop_event.is_set()

    if not args.discover_only and all_profiles:
        if stopped_early and not args.output:
            base_name = base_name + "_partial"
            out_xlsx = OUTPUT_DIR / f"{base_name}.xlsx"
            out_jsonl = OUTPUT_DIR / f"{base_name}.jsonl"
        _write_excel(all_profiles, out_xlsx)
        _write_jsonl(all_profiles, out_jsonl)
        if stopped_early:
            log.warning(
                f"⚠  Stopped early — {len(all_profiles)} profiles saved to "
                f"{out_jsonl.name} ({len(results)}/{len(firms)} firms completed)"
            )
    elif not args.discover_only and not all_profiles and stopped_early:
        log.warning("⚠  Stopped early — no profiles collected, no output files written.")
    elif args.discover_only:
        # Write a discovery-only summary JSON
        summary_path = OUTPUT_DIR / f"{base_name}_discovery_summary.json"
        summary = {
            firm_result.firm.name: {
                "url": firm_result.firm.url,
                "discovered": firm_result.discovered,
                "errors": firm_result.errors,
            }
            for firm_result in results
        }
        summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
        log.info(f"Discovery summary → {summary_path}")

    # ----------------------------------------------------------------
    # Run summary
    # ----------------------------------------------------------------
    _print_run_summary(results, elapsed)
    return 0


if __name__ == "__main__":
    sys.exit(main())
