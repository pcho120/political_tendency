from openpyxl import load_workbook
import sys

files = ["Company list.xlsx", "Company list_with_websites.xlsx"]

for f in files:
    print(f"Checking {f}...")
    try:
        wb = load_workbook(f)
        print(f"  OK. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"  ERROR: {e}")
